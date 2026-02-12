import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from scipy.stats import pointbiserialr
import streamlit.components.v1 as components
import re
from io import BytesIO
from datetime import datetime
import json
import pathlib
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from modules.statistics import calculate_kr20_reliability

# ═══════════════════════════════════════════════════════════════════
# Session State 초기화 (통합 구조)
# ═══════════════════════════════════════════════════════════════════

if 'app_config' not in st.session_state:
    st.session_state.app_config = {
        # ============= 평가 기본 설정 =============
        'eval': {
            'eval_type': 'achievement',
            'criterion_rate': 66.7,
            'target_rate': 70,
            'level_type': "5수준+미도달 (A, B, C, D, E, 미도달)"
        },
        
        # ============= 평가 계획 =============
        'plan': {
            'regular': {},
            'performance': []
        },
        
        # ============= 분석 선택 =============
        'selected': {
            'analysis_type': None,
            'analysis_category': None,
            'analysis_basis': '분할점수 기반',
            'exam_category': '정기고사'
        }
    }


# ═══════════════════════════════════════════════════════════════════
# 헬퍼 함수들
# ═══════════════════════════════════════════════════════════════════

def get_config(path: str, default=None):
    """
    세션 설정에서 값을 안전하게 가져옵니다.
    
    사용 예시:
        get_config('eval.eval_type')           → 'achievement'
        get_config('eval.criterion_rate')      → 66.7
        get_config('plan.regular')             → {...}
    """
    keys = path.split('.')
    value = st.session_state.app_config
    
    for key in keys:
        if isinstance(value, dict) and key in value:
            value = value[key]
        else:
            return default
    
    return value


def set_config(path: str, value):
    """
    세션 설정을 안전하게 변경합니다.
    
    사용 예시:
        set_config('eval.eval_type', 'selection')
        set_config('plan.regular', {})
    """
    keys = path.split('.')
    config = st.session_state.app_config
    
    # 마지막 키 전까지 순회
    for key in keys[:-1]:
        if key not in config:
            config[key] = {}
        config = config[key]
    
    # 마지막 키에 값 할당
    config[keys[-1]] = value


# 하위 호환성을 위한 속성 접근 (기존 코드와 호환)
if 'eval_type' not in st.session_state:
    st.session_state.eval_type = get_config('eval.eval_type')
if 'criterion_rate' not in st.session_state:
    st.session_state.criterion_rate = get_config('eval.criterion_rate')
if 'target_rate' not in st.session_state:
    st.session_state.target_rate = get_config('eval.target_rate')
if 'level_type' not in st.session_state:
    st.session_state.level_type = get_config('eval.level_type')
if 'eval_plan' not in st.session_state:
    st.session_state.eval_plan = get_config('plan')
if 'selected_analysis_type' not in st.session_state:
    st.session_state.selected_analysis_type = get_config('selected.analysis_type')
if 'selected_analysis_category' not in st.session_state:
    st.session_state.selected_analysis_category = get_config('selected.analysis_category')
if 'analysis_basis' not in st.session_state:
    st.session_state.analysis_basis = get_config('selected.analysis_basis')
if 'exam_category' not in st.session_state:
    st.session_state.exam_category = get_config('selected.exam_category')


# ═══════════════════════════════════════════════════════════════════
# 분석 유틸리티 함수
# ═══════════════════════════════════════════════════════════════════

def determine_item_level(achievement_rates: dict, criterion: float = 66.7) -> dict:
    """
    성취평가제 문항의 수준을 판정합니다.
    
    Parameters:
    - achievement_rates: {'A': 100.0, 'B': 87.5, 'C': 65.2, 'D': 45.0, 'E': 30.0, '미도달': 20.0}
    - criterion: 기준 정답률 (기본 66.7%)
    
    Returns:
    - dict: {'level': 'B', 'meets': ['A', 'B'], 'below': ['C', 'D', 'E', '미도달'], 'description': '...'}
    """
    levels = ['A', 'B', 'C', 'D', 'E', '미도달']
    meets_criterion = []
    below_criterion = []
    
    for level in levels:
        rate = achievement_rates.get(level, 0)
        if rate >= criterion:
            meets_criterion.append(level)
        else:
            below_criterion.append(level)
    
    if not meets_criterion:
        item_level = 'A이상'
    elif '미도달' in meets_criterion and len(meets_criterion) == len(levels):
        item_level = 'E'
    else:
        meets_without_below = [l for l in meets_criterion if l != '미도달']
        item_level = meets_without_below[-1] if meets_without_below else 'E'
    
    return {
        'level': item_level,
        'meets': meets_criterion,
        'below': below_criterion,
        'description': f"{item_level}수준 문항 ({', '.join(meets_criterion)}이(가) 기준 {criterion}% 충족)"
    }


def analyze_achievement_pattern(achievement_rates: dict, student_counts: dict = None) -> dict:
    """
    성취수준별 정답률 패턴을 분석합니다.
    
    ⚠️ 중요: 패턴을 "관찰된 현상"으로 표현하고, 가능한 원인을 안내합니다.
    
    Parameters:
    - achievement_rates: {'A': 100.0, 'B': 87.5, 'C': 89.5, ...}
    - student_counts: {'A': 5, 'B': 8, 'C': 19, ...} (선택)
    
    Returns:
    - dict: 패턴 분석 결과
    """
    levels = ['A', 'B', 'C', 'D', 'E', '미도달']
    observations = []
    
    for i in range(len(levels) - 1):
        upper = levels[i]
        lower = levels[i + 1]
        upper_rate = achievement_rates.get(upper, 0)
        lower_rate = achievement_rates.get(lower, 0)
        gap = upper_rate - lower_rate
        
        if gap < 0:
            abs_gap = abs(gap)
            
            if abs_gap <= 5:
                interpretation = 'normal'
                icon = '📊'
                title = '통계적 변동 범위'
                message = "이 정도의 차이는 통계적 변동 범위 내입니다."
            elif abs_gap <= 10:
                interpretation = 'notable'
                icon = '📋'
                title = '참고 사항'
                message = "다소 큰 차이가 관찰되었습니다. 다른 문항에서도 유사 패턴이 있는지 확인해 보세요."
            else:
                interpretation = 'significant'
                icon = '🔍'
                title = '검토 권고'
                message = "상당한 차이가 관찰되었습니다. 문항 내용 및 해당 성취수준 집단의 특성을 검토해 보세요."
            
            count_info = ""
            if student_counts:
                count_info = f" ({upper}수준 {student_counts.get(upper, '?')}명, {lower}수준 {student_counts.get(lower, '?')}명)"
            
            observations.append({
                'upper': upper,
                'lower': lower,
                'upper_rate': upper_rate,
                'lower_rate': lower_rate,
                'gap': gap,
                'abs_gap': abs_gap,
                'interpretation': interpretation,
                'icon': icon,
                'title': title,
                'message': message,
                'count_info': count_info,
                'disclaimer': "※ 이 현상이 문항의 결함을 의미하지는 않습니다."
            })
    
    if not observations:
        overall = "전반적으로 성취수준이 높을수록 정답률이 높은 정상적인 경향을 보입니다. ✓"
        overall_status = 'good'
    else:
        overall = f"일부 구간({len(observations)}개)에서 정답률 패턴이 관찰되었습니다."
        overall_status = 'observed'
    
    return {
        'has_observations': len(observations) > 0,
        'observations': observations,
        'overall': overall,
        'overall_status': overall_status
    }


def evaluate_difficulty(correct_rate: float) -> dict:
    """난이도 구간을 판정합니다. (CTT 기반)"""
    if correct_rate >= 80:
        return {'level': '매우 쉬움', 'icon': '🔵', 'description': '대부분의 학생이 맞힐 수 있는 문항'}
    elif correct_rate >= 60:
        return {'level': '쉬움', 'icon': '🟢', 'description': '다수의 학생이 맞힐 수 있는 문항'}
    elif correct_rate >= 40:
        return {'level': '적정', 'icon': '🟢', 'description': '중간 수준의 난이도'}
    elif correct_rate >= 20:
        return {'level': '어려움', 'icon': '🟡', 'description': '소수의 학생만 맞힐 수 있는 문항'}
    else:
        return {'level': '매우 어려움', 'icon': '🔴', 'description': '극소수만 맞힐 수 있는 문항'}


def evaluate_discrimination(discrimination: float) -> dict:
    """변별도 구간을 판정합니다. (CTT 기반)"""
    if discrimination >= 0.40:
        return {'level': '매우 높음 (0.4+)', 'icon': '🟢', 'description': '상·하위 집단 구분 명확'}
    elif discrimination >= 0.30:
        return {'level': '높음 (0.3+)', 'icon': '🟢', 'description': '높은 변별력'}
    elif discrimination >= 0.20:
        return {'level': '보통 (0.2+)', 'icon': '🟡', 'description': '중간 수준의 변별력'}
    elif discrimination >= 0.10:
        return {'level': '낮음 (0.1+)', 'icon': '🟠', 'description': '낮은 변별력'}
    else:
        return {'level': '매우 낮음 (0.1 미만)', 'icon': '🔴', 'description': '변별 기능 제한적'}


def analyze_distractor(response_dist: dict, correct_answer: int) -> list:
    """
    오답 매력도를 분석합니다.
    
    Parameters:
    - response_dist: {1: 11.5, 2: 82.0, 3: 3.3, 4: 3.3, 5: 0} (각 선택지 선택률)
    - correct_answer: 정답 번호
    
    Returns:
    - list: 각 선택지 분석 결과
    """
    results = []
    
    for option, rate in response_dist.items():
        if option == correct_answer:
            results.append({
                'option': option,
                'rate': rate,
                'type': 'correct',
                'status': '정답',
                'icon': '✓'
            })
        elif str(option).lower() in ['noresponse', '무응답']:
            results.append({
                'option': '무응답',
                'rate': rate,
                'type': 'no_response',
                'status': '무응답',
                'icon': '-'
            })
        else:
            if rate >= 5:
                results.append({
                    'option': option,
                    'rate': rate,
                    'type': 'functional',
                    'status': '적절',
                    'icon': '○',
                    'description': '기능적 오답 (매력도 있음)'
                })
            else:
                results.append({
                    'option': option,
                    'rate': rate,
                    'type': 'non_functional',
                    'status': '낮음',
                    'icon': '△',
                    'description': '낮은 선택률 오답 (5% 미만)'
                })
    
    return results

# ═══════════════════════════════════════════════════════════════════
# 엑셀 포매팅 함수
# ═══════════════════════════════════════════════════════════════════
# ===== 엑셀 포매팅 함수 =====
def format_excel_file(main_display, exam_name, basis_str, max_score, ratio):
    """
    엑셀 파일을 정렬된 동양으로 포매팅 (처음부터 새로 작성)
    """
    output = BytesIO()
    
    # 1. 새로운 엑셀 워크북 생성
    wb = Workbook()
    ws = wb.active
    ws.title = '학생성적데이터'
    
    # 2. 문서 제목 및 정보 섹션 작성
    row_num = 1
    
    # 제목
    title_cell = ws.cell(row=row_num, column=1, value="📊 학생 성적 데이터")
    title_cell.font = Font(name='맑은 고딕', size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1A5C9E", end_color="1A5C9E", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f"A{row_num}:H{row_num}")
    ws.row_dimensions[row_num].height = 24
    
    # 정보 행 1
    row_num += 1
    info_row1 = row_num
    ws.cell(row=info_row1, column=1, value="평가명:").font = Font(bold=True, size=10)
    ws.cell(row=info_row1, column=2, value=exam_name)
    ws.cell(row=info_row1, column=3, value="만점:").font = Font(bold=True, size=10)
    ws.cell(row=info_row1, column=4, value=max_score)
    ws.cell(row=info_row1, column=5, value="반영비율:").font = Font(bold=True, size=10)
    ws.cell(row=info_row1, column=6, value=f"{ratio}%")
    
    # 정보 행 2
    row_num += 1
    info_row2 = row_num
    ws.cell(row=info_row2, column=1, value="분석 기준:").font = Font(bold=True, size=10)
    ws.cell(row=info_row2, column=2, value=basis_str)
    ws.cell(row=info_row2, column=3, value="출력일시:").font = Font(bold=True, size=10)
    ws.cell(row=info_row2, column=4, value=datetime.now().strftime('%Y년 %m월 %d일 %H:%M:%S'))
    
    # 정보 행 스타일 적용
    for row in [info_row1, info_row2]:
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(name='맑은 고딕', size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 18
    
    # 공백 행
    row_num += 2
    blank_row = row_num
    ws.row_dimensions[blank_row].height = 8
    
    # 헤더가 시작되는 행
    row_num += 1
    header_row = row_num
    
    # 3. 테이블 스타일 정의
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(name='맑은 고딕', size=11, bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    data_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font = Font(name='맑은 고딕', size=10)
    
    # 4. 헤더 작성
    for col_num, col_name in enumerate(main_display.columns, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = col_name
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
    
    ws.row_dimensions[header_row].height = 20
    
    # 5. 데이터 작성
    data_start_row = header_row + 1
    for data_row_idx, (idx, row_data) in enumerate(main_display.iterrows()):
        excel_row = data_start_row + data_row_idx
        for col_num, col_name in enumerate(main_display.columns, 1):
            cell = ws.cell(row=excel_row, column=col_num)
            cell.value = row_data[col_name]
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
        ws.row_dimensions[excel_row].height = 18
    
    # 6. 셀 폭 자동 조정
    for col_num, col_name in enumerate(main_display.columns, 1):
        max_length = len(str(col_name)) + 2
        for row_data in main_display[col_name]:
            max_length = max(max_length, len(str(row_data)) + 2)
        
        # 최대 35, 최소 8로 제한
        adjusted_width = min(35, max(8, max_length))
        ws.column_dimensions[get_column_letter(col_num)].width = adjusted_width
    
    # 워크북 저장
    wb.save(output)
    output.seek(0)
    return output.getvalue()

# [DataTables 렌더링 함수]
def render_datatables(html_content, unique_id):
    datatables_html = f"""
    <!DOCTYPE html>
    <html lang="ko">
    <head>
        <meta charset="UTF-8">
        <link rel="stylesheet" type="text/css" href="https://cdn.datatables.net/1.13.6/css/jquery.dataTables.min.css">
        <style>
            table.dataTable thead th {{ text-align: center !important; vertical-align: middle !important; background-color: #f8f9fa !important; border: 1px solid #e0e0e0 !important; font-size: 0.9rem; }}
            table.dataTable tbody td {{ text-align: center !important; vertical-align: middle !important; border: 1px solid #e0e0e0 !important; font-size: 0.9rem; padding: 4px !important; }}
            table.dataTable thead .sorting:before, table.dataTable thead .sorting:after {{ bottom: 0.5em !important; }}
            .dataTables_wrapper .dataTables_paginate .paginate_button.current {{ background: #e0e0e0 !important; border: 1px solid #bdbdbd !important; }}
            body {{ font-family: 'Pretendard', sans-serif; }}
        </style>
        <script type="text/javascript" charset="utf8" src="https://code.jquery.com/jquery-3.7.0.js"></script>
        <script type="text/javascript" charset="utf8" src="https://cdn.datatables.net/1.13.6/js/jquery.dataTables.min.js"></script>
    </head>
    <body style="margin: 0;">
        {html_content}
        <script>
            $(document).ready(function() {{
                $('table').attr('id', 'example_{unique_id}');
                $('#example_{unique_id}').DataTable({{
                    "paging": false, 
                    "lengthChange": false, 
                    "searching": false, 
                    "ordering": true, 
                    "info": false, 
                    "autoWidth": false, 
                    "responsive": true, 
                    "order": [], 
                    "language": {{
                        "zeroRecords": "데이터가 없습니다.",
                        "infoEmpty": "데이터 없음"
                    }}
                }});
            }});
        </script>
    </body>
    </html>
    """
    return components.html(datatables_html, height=600, scrolling=True)

# [문항 통계 일괄 계산 함수 - 성능 최적화]
@st.cache_data(ttl=3600)
def calculate_all_item_statistics(main_df, info_df, available_levels):
    """
    모든 선택형 문항(1-16)의 통계를 한 번에 계산하여 반환
    Tab 3 로딩 성능 최적화용
    """
    item_stats = {}
    
    for i in range(1, 17):
        col = f'Item_{i}'
        
        # 기본 정보
        exp_diff_val = info_df[info_df['No'] == i]['Exp_Diff'].values
        exp_diff = exp_diff_val[0] if len(exp_diff_val) > 0 else '-'
        
        score_val = info_df[info_df['No'] == i]['Score'].values
        score = score_val[0] if len(score_val) > 0 else 0
        
        # 정답 번호
        try:
            ans_val = info_df[info_df['No'] == i]['Correct_Ans'].values[0]
            correct_ans = str(int(ans_val))
        except:
            correct_ans = ''
        
        # 정답률
        correct_rate = (main_df[col].astype(str) == '.').mean() * 100
        
        # 선택지별 응답분포
        item_responses = main_df[col].astype(str).value_counts()
        choice_counts = {str(j): item_responses.get(str(j), 0) for j in range(1, 6)}
        choice_counts['.'] = item_responses.get('.', 0)
        
        # 변별도 계산 (상위/하위 27%)
        total_students = len(main_df)
        top_n = int(total_students * 0.27)
        
        if 'Total_Score' in main_df.columns:
            sorted_df = main_df.sort_values('Total_Score', ascending=False)
        else:
            item_cols = [f'Item_{j}' for j in range(1, 17) if f'Item_{j}' in main_df.columns]
            main_df['_temp_score'] = main_df[item_cols].apply(
                lambda row: sum([1 if str(val) == '.' else 0 for val in row]), axis=1
            )
            sorted_df = main_df.sort_values('_temp_score', ascending=False)
        
        top_group = sorted_df.head(top_n)
        bottom_group = sorted_df.tail(top_n)
        
        top_correct_rate = (top_group[col].astype(str) == '.').mean()
        bottom_correct_rate = (bottom_group[col].astype(str) == '.').mean()
        discrimination = top_correct_rate - bottom_correct_rate
        
        # 성취수준별 정답률
        achievement_rates = {}
        for level in available_levels:
            level_data = main_df[main_df['Achievement'] == level]
            if len(level_data) > 0:
                lv_rate = (level_data[col].astype(str) == '.').mean() * 100
                achievement_rates[level] = lv_rate
            else:
                achievement_rates[level] = 0.0
        
        # 통계 저장
        item_stats[i] = {
            'exp_diff': exp_diff,
            'score': score,
            'correct_ans': correct_ans,
            'correct_rate': correct_rate,
            'discrimination': discrimination,
            'choice_counts': choice_counts,
            'achievement_rates': achievement_rates
        }
    
    return item_stats

# [시각화 함수 정의]
def custom_bar_style(val, threshold):
    try:
        v = float(val)
        if pd.isna(v): return ''
        # 정답률은 모든 셀을 흰색으로
        bg_color = '#ffffff'
        return f"background: linear-gradient(90deg, #90caf9 {v}%, {bg_color} {v}%); color: black;"
    except:
        return ''

# [배경색 스타일 함수 - 성취수준별 정답률]
def style_background_level_v2(val, threshold):
    try:
        if isinstance(val, str): return ''
        v = float(val)
        # 기준 미만이면 회색(#eeeeee), 이상이면 흰색(#ffffff)
        bg_color = '#eeeeee' if v < threshold else '#ffffff'
        return f'background-color: {bg_color}; color: black;'
    except:
        return ''

# [HTML 후처리] 헤더 병합
def merge_headers(html_content, target_cols):
    thead_match = re.search(r'(<thead[^>]*>)(.*?)(</thead>)', html_content, re.DOTALL)
    if not thead_match: return html_content
    thead_open, thead_body, thead_close = thead_match.groups()
    rows = re.findall(r'(<tr[^>]*>)(.*?)(</tr>)', thead_body, re.DOTALL)
    if len(rows) < 2: return html_content
    tr1_open, tr1_content, tr1_close = rows[0]
    tr2_open, tr2_content, tr2_close = rows[1]
    for col in target_cols:
        pattern = re.compile(r'(<th\b[^>]*>)(\s*' + re.escape(col) + r'\s*)(</th>)')
        if pattern.search(tr1_content):
            def add_rowspan(match):
                tag_open = match.group(1)
                if 'rowspan' not in tag_open:
                    return tag_open.replace('<th', '<th rowspan="2"') + match.group(2) + match.group(3)
                return match.group(0)
            tr1_content = pattern.sub(add_rowspan, tr1_content)
        if pattern.search(tr2_content):
            tr2_content = pattern.sub('', tr2_content)
    new_thead = f"{thead_open}\n{tr1_open}{tr1_content}{tr1_close}\n{tr2_open}{tr2_content}{tr2_close}\n{thead_close}"
    return html_content.replace(thead_match.group(0), new_thead)

# --- 페이지 설정 (모바일 친화적 설정) ---
st.set_page_config(
    page_title="성취평가 문항 분석 시스템",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="expanded"
)

# --- 디자인 커스텀 (CSS) - 외부 파일에서 로드 ---
try:
    css_path = pathlib.Path(__file__).parent / "styles" / "main.css"
    with open(css_path, "r", encoding="utf-8") as css_file:
        css_content = css_file.read()
    st.markdown(f"<style>\n{css_content}\n</style>", unsafe_allow_html=True)
except Exception as e:
    st.warning(f"CSS 파일 로드 실패: {e}")
    # Fallback: 기본 스타일만 적용
    st.markdown("""
    <style>
    @import url("https://cdn.jsdelivr.net/gh/orioncactus/pretendard@v1.3.9/dist/web/static/pretendard.min.css");
    html, body, [class*="css"] {
        font-family: 'Pretendard', sans-serif;
    }
    .stApp {
        background-color: #FFFFFF;
        color: #1E293B;
    }
    </style>
    """, unsafe_allow_html=True)

# --- 데이터 처리 로직 ---
# --- 데이터 처리 로직 ---
@st.cache_data(ttl=3600)
def load_and_merge_data(info_file, ans_files, grade_files):
    try:
        # 1. 문항정보표 파싱
        # dtype={'No': str}로 읽어서 문항 번호가 숫자로 자동 변환되는 것 방지
        info = pd.read_excel(info_file, skiprows=10, engine='openpyxl', dtype={'No': str}).iloc[:22]
        info = info.iloc[:, [1, 3, 14, 16, 18, 19, 21]]
        info.columns = ['No', 'Standard', 'Hard', 'Medium', 'Easy', 'Score', 'Correct_Ans']
        
        # 유효한 문항 번호만 필터링 (숫자인 행만)
        info = info[info['No'].apply(lambda x: str(x).replace('.0','').strip().isdigit())].copy()
        info['No'] = info['No'].astype(float).astype(int) # 정수형으로 변환
        
        # 배점(Score) 숫자 변환 (에러 발생 시 0점 처리)
        info['Score'] = pd.to_numeric(info['Score'], errors='coerce').fillna(0)
        
        # 난이도(Exp_Diff) 계산
        info['Exp_Diff'] = info.apply(lambda r: '상' if r['Hard']=='○' else ('중' if r['Medium']=='○' else '하'), axis=1)

        # 2. 정오표 병합 (문항 번호 기준 상대 위치 파악)
        all_ans = []
        
        # 파일 데이터에서 강의실 번호 추출 함수
        def extract_classroom_from_data(raw_preview):
            """정오표 파일의 상단 데이터에서 강의실 번호 추출"""
            import re
            # 상용 20행 탐색
            for row_idx in range(min(10, len(raw_preview))):
                row_str = ' '.join([str(val) for val in raw_preview.iloc[row_idx].values])
                # "강의실" 다음의 숫자 찾기 (예: "4 강의실", "강의실 1", "강의실1")
                match = re.search(r'(\d+)\s*강의실|강의실\s*(\d+)', row_str)
                if match:
                    classroom = match.group(1) if match.group(1) else match.group(2)
                    return classroom.strip()
            return None
        
        for f in ans_files:
            # 상위 20행 미리보기
            raw_preview = pd.read_excel(f, nrows=20, header=None, engine='openpyxl', dtype=str)
            
            # 파일의 데이터에서 강의실 번호 추출
            classroom_no = extract_classroom_from_data(raw_preview)
            
            header_row_idx = -1
            item_start_col_idx = -1
            name_col_idx = -1
            id_col_idx = -1
            
            # 1. 문항 번호(1, 2, 3...)가 있는 행 찾기
            item_col_map = {}  # {문항번호: 컬럼인덱스}
            for r_idx, row in raw_preview.iterrows():
                row_str = row.astype(str).values
                # 1, 2, 3, 4, 5가 연속해서 등장하거나 포함된 행 찾기
                # 간단하게 '1', '2', '3', '4'가 모두 포함되어 있는지 확인
                if '1' in row_str and '2' in row_str and '3' in row_str and '4' in row_str:
                    header_row_idx = r_idx
                    # 각 문항 번호(1~16)의 컬럼 인덱스 찾기
                    for c_idx, val in enumerate(row_str):
                        val_clean = str(val).strip().replace('.0', '')
                        if val_clean.isdigit():
                            item_num = int(val_clean)
                            if 1 <= item_num <= 16 and item_num not in item_col_map:
                                item_col_map[item_num] = c_idx
                    # '1'의 컬럼 인덱스 (기존 호환성)
                    if 1 in item_col_map:
                        item_start_col_idx = item_col_map[1]
                    if len(item_col_map) >= 4:  # 최소 4개 문항 발견
                        break
            
            if header_row_idx == -1 or item_start_col_idx == -1:
                # 문항 번호 헤더를 찾지 못한 경우, 기존 방식 (성명/이름 찾기) 시도
                st.warning(f"'{f.name}' 파일에서 문항 번호 헤더를 찾을 수 없습니다. '성명' 또는 '이름' 열을 기준으로 데이터를 파싱합니다.")
                
                name_col_idx_fallback = -1
                score_col_idx_fallback = -1
                header_row_idx_fallback = -1

                for r_idx, row in raw_preview.iterrows():
                    row_str = row.astype(str).values
                    if any('성명' in str(x) for x in row_str) or any('이름' in str(x) for x in row_str):
                        header_row_idx_fallback = r_idx
                        for c_idx, val in enumerate(row_str):
                            if '성명' in str(val) or '이름' in str(val): name_col_idx_fallback = c_idx
                            if '번호' in str(val) or 'ID' in str(val): id_col_idx = c_idx # Use global id_col_idx
                            if '점수' in str(val) or 'Total' in str(val) or '합계' in str(val): score_col_idx_fallback = c_idx
                        break
                
                if header_row_idx_fallback == -1:
                    st.error(f"'{f.name}' 파일에서 '성명' 또는 '이름' 열을 찾을 수 없습니다.")
                    with st.expander(f"❌ '{f.name}' 원본 데이터 미리보기 (상위 20행)"):
                        st.dataframe(raw_preview)
                    return None, None

                # Fallback: 데이터 다시 읽기 (기존 로직)
                raw = pd.read_excel(f, skiprows=header_row_idx_fallback + 1, header=None, engine='openpyxl', dtype=str)
                
                # 컬럼 매핑
                if score_col_idx_fallback == -1: score_col_idx_fallback = len(raw.columns) - 1 # 맨 마지막
                
                data = raw.copy()
                data = data.rename(columns={name_col_idx_fallback: 'Name'})
                
                if id_col_idx != -1:
                    data = data.rename(columns={id_col_idx: 'ID'})
                else:
                    data['ID'] = data.index # 임시
                
                data = data.rename(columns={score_col_idx_fallback: 'Total_Score'})

                # 문항 컬럼 추출 (Name 컬럼 뒤쪽으로 2칸 띄우고 시작한다고 가정 - 기존 3->5 패턴)
                # Name이 found 안되면... error.
                start_col = name_col_idx_fallback + 2
                for i in range(1, 17):
                    if start_col + i - 1 < len(data.columns):
                         data[f'Item_{i}'] = data.iloc[:, start_col + i - 1]
                    else:
                         data[f'Item_{i}'] = '.'
                
                # 불필요한 행 제거 (정답, 배점 등 문자열이 이름에 있는 경우)
                data = data[~data['Name'].isin(['정답', '배점', '합계', '평균', 'None', 'nan'])]
                
                # 강의실 정보 추가
                if classroom_no:
                    data['강의실'] = classroom_no
                
                # 필요한 컬럼만 추출
                cols = ['ID', 'Name', 'Total_Score'] + [f'Item_{i}' for i in range(1, 17)]
                if classroom_no:
                    cols.append('강의실')
                final_cols = [c for c in cols if c in data.columns]
                all_ans.append(data[final_cols])
                continue # 다음 파일로 넘어감
            
            # 문항 번호 기준으로 데이터 로드 (새로운 로직)
            # 데이터 시작: 헤더 + 3 (정답, 배점 행 제외) - NEIS 표준
            data_start_row = header_row_idx + 3 
            raw = pd.read_excel(f, skiprows=data_start_row, header=None, engine='openpyxl', dtype=str)
            
            # 컬럼 매핑 (상대 위치)
            # 문항 1번이 item_start_col_idx에 있음.
            # 성명은 보통 문항 1번 보다 앞쪽 2칸 (item_start_col_idx - 2)
            # 번호/ID는 문항 1번 보다 앞쪽 4칸 (item_start_col_idx - 4)
            
            # 초기 추정
            name_col_idx_candidate = item_start_col_idx - 2
            id_col_idx_candidate = item_start_col_idx - 4 
            
            # Name 컬럼 유효성 검사 및 보정
            def looks_like_korean_name(s):
                if pd.isna(s) or not isinstance(s, str): return False
                s = s.strip()
                if len(s) < 2 or len(s) > 5: return False # 일반적인 이름 길이
                return all('가' <= char <= '힣' for char in s) # 한글 여부
            
            # name_col_idx_candidate가 유효한지 확인
            if name_col_idx_candidate >= 0 and name_col_idx_candidate < len(raw.columns):
                sample_names = raw.iloc[:10, name_col_idx_candidate].dropna().tolist()
                korean_name_count = sum(1 for s in sample_names if looks_like_korean_name(s))
                
                if korean_name_count < 3: # 충분히 한글 이름 같지 않으면
                    # item_start_col_idx - 1 위치 확인
                    if item_start_col_idx - 1 >= 0 and item_start_col_idx - 1 < len(raw.columns):
                        sample_names_alt = raw.iloc[:10, item_start_col_idx - 1].dropna().tolist()
                        korean_name_count_alt = sum(1 for s in sample_names_alt if looks_like_korean_name(s))
                        if korean_name_count_alt >= 3:
                            name_col_idx = item_start_col_idx - 1
                        else: # 둘 다 아니면 초기 추정 사용 (최악의 경우)
                            name_col_idx = name_col_idx_candidate
                    else:
                        name_col_idx = name_col_idx_candidate
                else:
                    name_col_idx = name_col_idx_candidate
            else: # name_col_idx_candidate가 범위 밖이면
                name_col_idx = -1 # 찾지 못함
            
            # ID 컬럼 (name_col_idx - 2 또는 id_col_idx_candidate)
            if name_col_idx != -1 and name_col_idx - 2 >= 0:
                id_col_idx = name_col_idx - 2
            elif id_col_idx_candidate >= 0:
                id_col_idx = id_col_idx_candidate
            else:
                id_col_idx = -1 # 찾지 못함

            score_col_idx = len(raw.columns) - 1 # 맨 뒤 컬럼을 점수로 가정
            
            data = raw.copy()
            
            # [수정] 반/번호 컬럼 찾기 - 이름 컬럼 왼쪽에서 숫자/숫자 패턴 찾기
            import re
            def is_class_num_format(s):
                if pd.isna(s): return False
                return bool(re.match(r'^\d+[/\-]\d+$', str(s).strip()))
            
            # 이름 컬럼 왼쪽에서 반/번호 컬럼 찾기
            class_num_col_idx = -1
            for col_offset in range(1, min(name_col_idx + 1, 4)):  # 최대 3칸 왼쪽까지 탐색
                check_idx = name_col_idx - col_offset
                if check_idx >= 0 and check_idx < len(data.columns):
                    sample_vals = data.iloc[:10, check_idx].tolist()
                    valid_count = sum(1 for x in sample_vals if is_class_num_format(x))
                    if valid_count >= 3:
                        class_num_col_idx = check_idx
                        break
            
            # 컬럼 이름 변경
            col_mapping = {}
            if name_col_idx != -1 and name_col_idx < len(data.columns): 
                col_mapping[name_col_idx] = 'Name'
            else:
                st.warning(f"'{f.name}' 파일에서 'Name' 컬럼을 찾을 수 없습니다. 데이터 처리에 문제가 있을 수 있습니다.")
                data['Name'] = 'Unknown_' + data.index.astype(str) # 임시 이름
            
            if class_num_col_idx != -1 and class_num_col_idx < len(data.columns): 
                col_mapping[class_num_col_idx] = 'ClassNum'
            
            if score_col_idx < len(data.columns):
                col_mapping[score_col_idx] = 'Total_Score' # 마지막 컬럼 점수
            else:
                st.warning(f"'{f.name}' 파일에서 'Total_Score' 컬럼을 찾을 수 없습니다. 0으로 처리됩니다.")
                data['Total_Score'] = 0
            
            data = data.rename(columns=col_mapping)
            
            # [신규] 반/번호 -> 학번 변환
            def parse_class_num_to_id(s):
                """'1/1' -> '20101' (2학년 01반 01번)"""
                if pd.isna(s): return ''
                s = str(s).strip()
                match = re.match(r'^(\d+)[/\-](\d+)$', s)
                if match:
                    class_no = match.group(1).zfill(2)
                    student_no = match.group(2).zfill(2)
                    return f'2{class_no}{student_no}'
                return ''
            
            if 'ClassNum' in data.columns:
                data['ID'] = data['ClassNum'].apply(parse_class_num_to_id)
                data = data.drop(columns=['ClassNum'])
            else:
                data['ID'] = ''
            
            # 문항 컬럼 매핑 (item_col_map 사용)
            for item_num in range(1, 17):
                if item_num in item_col_map:
                    q_idx = item_col_map[item_num]
                else:
                    # 매핑되지 않은 경우 순차 오프셋 사용 (fallback)
                    q_idx = item_start_col_idx + (item_num - 1)
                
                if q_idx < len(raw.columns):
                    data[f'Item_{item_num}'] = raw.iloc[:, q_idx]
                else:
                    data[f'Item_{item_num}'] = '.'

            # 불필요한 행 제거 (정답, 배점 등 문자열이 이름에 있는 경우)
            data = data[~data['Name'].isin(['정답', '배점', '합계', '평균', 'None', 'nan'])]
            
            # 강의실 정보 추가
            if classroom_no:
                data['강의실'] = classroom_no
            
            # 필요한 컬럼만 추출
            cols = ['ID', 'Name', 'Total_Score'] + [f'Item_{i}' for i in range(1, 17)]
            if classroom_no:
                cols.append('강의실')
            final_cols = [c for c in cols if c in data.columns]
            all_ans.append(data[final_cols])
        
        if not all_ans:
            return None, None

        ans_df = pd.concat(all_ans)
        ans_df = ans_df.dropna(subset=['Name']) # 이름 없는 행 제거
        ans_df['Name'] = ans_df['Name'].astype(str).str.strip() # 공백 제거

        # 3. 성적일람표 병합 (다중 파일 지원 & 동적 헤더 탐색 & 컬럼 보정)
        all_grades = []
        if not isinstance(grade_files, list):
            grade_files = [grade_files]
            
        for f in grade_files:
            # 1. 헤더 위치 찾기 (상위 30행 탐색)
            raw_preview = pd.read_excel(f, nrows=30, header=None, engine='openpyxl', dtype=str)
            
            name_row_idx = -1
            grade_row_idx = -1
            name_col_idx = -1
            grade_col_idx = -1
            
            # 전체 셀을 순회하며 키워드 찾기
            for r_idx, row in raw_preview.iterrows():
                row_str = row.astype(str).values
                for c_idx, val in enumerate(row_str):
                    val_str = str(val)
                    # 성명 컬럼 찾기
                    if name_col_idx == -1 and ('성명' in val_str or '이름' in val_str):
                        name_row_idx = r_idx
                        name_col_idx = c_idx
                    # 성취도 컬럼 찾기
                    if grade_col_idx == -1 and ('성취도' in val_str or '등급' in val_str):
                        grade_row_idx = r_idx
                        grade_col_idx = c_idx
            
            if name_col_idx != -1 and grade_col_idx != -1:
                # 데이터 시작 행: 헤더 아래
                data_start_row = max(name_row_idx, grade_row_idx) + 1
                
                # 데이터 로드
                g_raw = pd.read_excel(f, skiprows=data_start_row, header=None, engine='openpyxl', dtype=str)
                
                # [신규] 반/번호 컬럼 찾기
                class_num_col_idx = -1
                for r_idx, row in raw_preview.iterrows():
                    row_str = row.astype(str).values
                    for c_idx, val in enumerate(row_str):
                        val_str = str(val)
                        if '반' in val_str and '번' in val_str:  # "반/번호" 또는 "반번"
                            class_num_col_idx = c_idx
                            break
                    if class_num_col_idx != -1:
                        break
                
                # [중요] Name 컬럼 보정 로직
                # 찾아낸 name_col_idx가 실제 이름이 아니라 ID(숫자) 등일 수 있음 (Merge Cell 문제)
                # 해당 컬럼의 데이터가 한글 이름인지 확인
                def looks_like_name(s):
                    # 길이가 2~5이고, 숫자가 포함되지 않아야 함
                    if pd.isna(s) or len(str(s)) < 2: return False
                    return not any(char.isdigit() for char in str(s))

                # 현재 컬럼 데이터 확인 (상위 5개)
                sample_data = g_raw.iloc[:10, name_col_idx].tolist()
                valid_count = sum(looks_like_name(x) for x in sample_data)
                
                # 만약 유효한 이름이 적다면, 오른쪽으로 이동하며 탐색 (최대 3칸)
                if valid_count < 3: 
                    found_better = False
                    for offset in range(1, 4):
                        if name_col_idx + offset < len(g_raw.columns):
                            sample_next = g_raw.iloc[:10, name_col_idx + offset].tolist()
                            if sum(looks_like_name(x) for x in sample_next) >= 3:
                                name_col_idx += offset
                                found_better = True
                                break
                
                # 찾은 인덱스로 데이터 선택
                # [수정] 반/번호 컬럼 찾기 - 성명 컬럼 왼쪽에 있음
                class_num_col_idx = name_col_idx - 1 if name_col_idx > 0 else -1
                
                # 반/번호 컬럼 유효성 확인 (숫자/숫자 형태인지)
                import re
                def is_class_num_format(s):
                    if pd.isna(s): return False
                    return bool(re.match(r'^\d+[/\-]\d+$', str(s).strip()))
                
                if class_num_col_idx >= 0 and class_num_col_idx < len(g_raw.columns):
                    sample_class = g_raw.iloc[:10, class_num_col_idx].tolist()
                    valid_class_count = sum(1 for x in sample_class if is_class_num_format(x))
                    
                    if valid_class_count >= 3:
                        # 반/번호 컬럼 포함하여 선택
                        g_raw = g_raw.iloc[:, [class_num_col_idx, name_col_idx, grade_col_idx]]
                        g_raw.columns = ['ClassNum', 'Name', 'Achievement']
                        
                        # 학번(ID) 생성: "1/1" -> "20101"
                        def parse_class_num(s):
                            if pd.isna(s): return ''
                            s = str(s).strip()
                            match = re.match(r'^(\d+)[/\-](\d+)$', s)
                            if match:
                                class_no = match.group(1).zfill(2)
                                student_no = match.group(2).zfill(2)
                                return f'2{class_no}{student_no}'
                            return ''
                        
                        g_raw['ID'] = g_raw['ClassNum'].apply(parse_class_num)
                        g_raw = g_raw.drop(columns=['ClassNum'])
                    else:
                        # 반/번호 컬럼이 없으면 기존 방식 (이름만)
                        g_raw = g_raw.iloc[:, [name_col_idx, grade_col_idx]]
                        g_raw.columns = ['Name', 'Achievement']
                        g_raw['ID'] = ''
                else:
                    g_raw = g_raw.iloc[:, [name_col_idx, grade_col_idx]]
                    g_raw.columns = ['Name', 'Achievement']
                    g_raw['ID'] = ''
                
                all_grades.append(g_raw)
            else:
                st.error(f"'{f.name}' 파일에서 '성명'과 '성취도' 열을 찾을 수 없습니다.")
                with st.expander(f"❌ '{f.name}' 원본 데이터 미리보기"):
                    st.dataframe(raw_preview.head(10))
                return None, None
        
        if not all_ans:
            # 정오표가 없는 경우 (수행평가)
            # 성적일람표에서만 데이터 생성
            if not all_grades:
                st.error("❌ 데이터를 로드할 수 없습니다. 필요한 파일을 올바르게 업로드했는지 확인하세요.")
                return None, None
            
            grade = pd.concat(all_grades)
            grade = grade.dropna(subset=['Name'])
            grade['Name'] = grade['Name'].astype(str).str.strip()
            
            # 수행평가용 기본 데이터 생성 (정오표 대신 성적일람표만 사용)
            merged = grade.copy()
            merged['Total_Score'] = 0  # 임시
            
            # 16개 문항 컬럼 추가 (더미 데이터)
            for i in range(1, 17):
                merged[f'Item_{i}'] = '.'
        else:
            # 정오표가 있는 경우 (정기고사)
            ans_df = pd.concat(all_ans)
            ans_df = ans_df.dropna(subset=['Name'])
            ans_df['Name'] = ans_df['Name'].astype(str).str.strip()
            
            if not all_grades:
                # 성적일람표가 없으면 정오표만 사용
                merged = ans_df.copy()
                # Total_Score가 없으면 생성
                if 'Total_Score' not in merged.columns:
                    merged['Total_Score'] = 0
                # Achievement가 없으면 임시 생성
                if 'Achievement' not in merged.columns:
                    merged['Achievement'] = 'E'
            else:
                # 성적일람표와 병합
                grade = pd.concat(all_grades)
                grade = grade.dropna(subset=['Name'])
                grade['Name'] = grade['Name'].astype(str).str.strip()
                
                # 학생 수 비교 (학적 변동 확인)
                ans_students = set(ans_df['Name'].unique())
                grade_students = set(grade['Name'].unique())
                
                excluded_students = ans_students - grade_students  # 정오표에만 있는 학생
                
                if excluded_students:
                    st.warning(
                        f"⚠️ **학생 수 불일치 감지**\n\n"
                        f"• 정오표 학생 수: {len(ans_students)}명\n"
                        f"• 성적일람표 학생 수: {len(grade_students)}명\n\n"
                        f"**성적일람표를 기준으로 {len(excluded_students)}명 제외** (학적변동: 전출, 자퇴 등)\n\n"
                        f"분석에는 성적일람표에 있는 {len(grade_students)}명만 포함됩니다."
                    )
                    with st.expander(f"제외된 학생 목록 ({len(excluded_students)}명) - 클릭하여 확인"):
                        st.write("**다음 학생들이 정오표에는 있지만 성적일람표에는 없어서 제외되었습니다:**")
                        st.write(", ".join(sorted(list(excluded_students))))
                
                # 정오표와 성적일람표 병합 (성적일람표 기준)
                merged = pd.merge(ans_df, grade[['Name', 'Achievement', 'ID']], on='Name', how='inner', suffixes=('', '_grade'))
                
                # ID 우선순위: 정오표 ID > 성적일람표 ID
                if 'ID' in merged.columns and 'ID_grade' in merged.columns:
                    merged['ID'] = merged.apply(
                        lambda row: row['ID'] if row['ID'] and str(row['ID']).strip() else row['ID_grade'], 
                        axis=1
                    )
                    merged = merged.drop(columns=['ID_grade'])
                elif 'ID_grade' in merged.columns:
                    merged['ID'] = merged['ID_grade']
                    merged = merged.drop(columns=['ID_grade'])
        
        # 병합 결과 확인
        if merged.empty:
            st.warning("⚠️ **분석할 데이터가 없습니다.** 정오표와 성적일람표의 '이름'이 일치하는지 확인해주세요.")
            with st.expander("🔍 데이터 불일치 상세 정보 확인 (클릭)"):
                st.write(f"**정오표(Answer Sheet) 학생 수:** {len(ans_df)}명")
                st.write(f"**성적일람표(Grade Report) 학생 수:** {len(grade)}명")
                
                c1, c2 = st.columns(2)
                with c1:
                    st.write("#### 📋 정오표 이름 예시 (상위 5명)")
                    st.dataframe(ans_df[['Name']].head())
                with c2:
                    st.write("#### 📋 성적일람표 이름 예시 (상위 5명)")
                    st.dataframe(grade[['Name']].head())
                    
                st.info("Tip: 이름 사이에 공백이 다르거나(예: '홍길동' vs '홍 길 동'), 오타가 있는지 확인해보세요.")
            return info, pd.DataFrame() # 빈 데이터프레임 반환

        merged['Total_Score'] = pd.to_numeric(merged['Total_Score'], errors='coerce').fillna(0)
        return info, merged.dropna(subset=['Achievement'])

    except Exception as e:
        st.error(f"데이터 처리 중 오류가 발생했습니다: {e}")
        import traceback
        st.error(f"```\n{traceback.format_exc()}\n```")
        return None, None

# level_type 초기값 설정 (widget 생성 전)
if 'level_type' not in st.session_state:
    st.session_state.level_type = "5수준+미도달 (A, B, C, D, E, 미도달)"

# 대기 중인 설정 먼저 적용 (widget 생성 전)
if '_pending_settings' in st.session_state and st.session_state._pending_settings:
    pending = st.session_state._pending_settings
    st.session_state.analysis_basis = pending.get('analysis_basis', '분할점수 기반')
    st.session_state.exam_category = pending.get('exam_category', '정기고사')
    st.session_state.level_type = pending.get('level_type', st.session_state.level_type)
    st.session_state.eval_plan = pending.get('eval_plan', {})
    st.session_state.selected_analysis_type = pending.get('selected_analysis_type', '')
    st.session_state.selected_analysis_category = pending.get('selected_analysis_category', '')
    # 즉시 삭제하여 다시 적용되지 않도록 방지
    del st.session_state._pending_settings
    st.rerun()
    st.rerun()

# --- 사이드바 UI ---
with st.sidebar:
    st.markdown("### 📂 성적 분석 설정")
    
    # 보안 문구 (강조)
    st.info(
        "🔒 **데이터 보안 안내**\n\n"
        "본 서비스는 **사용자의 데이터를 서버로 전송하지 않고, "
        "브라우저에서 직접 읽고 분석합니다.**"
    )
    
    st.markdown("---")
    
    # 1단계: 분석 기준 선택
    st.subheader("1️⃣ 분석 기준 선택")
    # 분석 방식을 선택하세요
    options = ["분할점수 기반", "학기말 성취도 기반"]
    default_basis = st.session_state.get('analysis_basis', '분할점수 기반')
    default_index = options.index(default_basis) if default_basis in options else 0
    
    analysis_basis = st.radio(
        "📋 분석 방식을 선택하세요",
        options,
        index=default_index,
        help="분할점수: 각 평가의 점수로 성취도 판정\n학기말 성취도: 기존 성적일람표의 성취도 사용"
    )
    
    st.markdown("---")
    
    # 문항 해석 기준 선택 (Context 부여)
    st.subheader("🎯 문항 해석 기준 설정")
    st.caption("평가의 목적에 따라 문항 해석 기준이 달라집니다.")
    
    # Radio 버튼 - session_state에 자동 저장
    test_type = st.radio(
        "평가 목적을 선택하세요",
        ["석차 5등급제 (상대평가)", "성취평가제 (절대평가)"],
        key="test_type_radio",
        help="석차 5등급제: 변별도가 핵심. 학생들을 서열화하는 것이 목적\n"
             "성취평가제: 학습 목표 달성이 핵심. 변별도가 낮아도 정답률이 높으면 '성공적 수업'"
    )
    
    # 평가 유형에 따라 session_state에 저장
    if test_type == "성취평가제 (절대평가)":
        st.session_state.eval_type = 'achievement'
        set_config('eval.eval_type', 'achievement')
        
        criterion_rate = st.number_input(
            "📊 기준 정답률 (성취수준별 최소 도달 비율)",
            min_value=50.0,
            max_value=100.0,
            value=st.session_state.criterion_rate,
            step=0.1,
            key="criterion_rate_input",
            help="각 성취수준 학생의 최소 정답률 기준 (기본값: 66.7% = KICE 2/3 기준)\n"
                 "예) A수준 학생의 66.7% 이상이 맞춰야 'A수준 문항'으로 판정"
        )
        st.session_state.criterion_rate = criterion_rate
        st.session_state.target_rate = criterion_rate
        set_config('eval.criterion_rate', criterion_rate)
        set_config('eval.target_rate', criterion_rate)
        
        # 문항 수준 판정 기준 안내
        with st.expander("📖 문항 수준 판정 기준", expanded=False):
            st.markdown("""
            | 문항 수준 | 기준 충족 조건 |
            |-----------|----------------|
            | A수준 문항 | A만 기준 충족 |
            | B수준 문항 | A, B가 기준 충족 |
            | C수준 문항 | A, B, C가 기준 충족 |
            | D수준 문항 | A, B, C, D가 기준 충족 |
            | E수준 문항 | A~E 모두 기준 충족 |
            """)
        
        st.caption("💡 각 성취수준 학생의 2/3(66.7%)가 맞힐 수 있는 문항을 해당 수준 문항으로 판정합니다.")
        
    else:  # 석차 5등급제 (상대평가)
        st.session_state.eval_type = 'selection'
        set_config('eval.eval_type', 'selection')
        
        target_rate = st.slider(
            "🎯 목표 정답률 (%)",
            min_value=30.0,
            max_value=90.0,
            value=st.session_state.target_rate,
            step=5.0,
            key="target_rate_slider",
            help="문항의 적정 난이도를 판단하는 기준"
        )
        st.session_state.target_rate = target_rate
        st.session_state.criterion_rate = 66.7
        set_config('eval.target_rate', target_rate)
        set_config('eval.criterion_rate', 66.7)
        
        st.caption("💡 상대평가에서는 변별도와 난이도 분포가 중요합니다.")
    
    st.markdown("---")
    
    # 2단계: 평가 계획 및 분석 대상 설정
    st.subheader("2️⃣ 평가 계획 및 분석 대상 설정")
    
    # 과목명 입력 (공통)
    subject_name = st.text_input("📚 과목명", value="", placeholder="예) 공통국어1, 대수, 물리학", help="분석할 과목명을 입력하세요")

    # 세션 상태 초기화: 평가 항목 (빈 상태로 시작) - 이미 app_config에 초기화됨
    if 'eval_plan' not in st.session_state:
        st.session_state.eval_plan = get_config('plan')
    
    # 선택된 분석 대상을 저장할 세션 변수 (기본값 없음) - 이미 app_config에 초기화됨
    if 'selected_analysis_type' not in st.session_state:
        st.session_state.selected_analysis_type = get_config('selected.analysis_type')
    if 'selected_analysis_category' not in st.session_state:
        st.session_state.selected_analysis_category = get_config('selected.analysis_category')

    # 탭 구성: 평가 계획 관리 / 분석 대상 선택
    tab_plan, tab_target = st.tabs(["📝 평가 계획 관리", "🎯 분석 대상 선택"])

    with tab_plan:
        # 정기고사 관리
        # 정기고사 반영비율 합계 계산
        total_regular_ratio = sum([exam['ratio'] for exam in st.session_state.eval_plan['regular'].values()])
        
        # 헤더: 제목과 추가/항목선택/삭제 버튼 (한 줄)
        st.markdown(f"##### 📌 정기시험 구성 관계: **{total_regular_ratio}%**")
        col_add, col_item, col_del = st.columns([1.0, 1.8, 1.0])
        with col_add:
            if st.button("추가", key="btn_add_regular", use_container_width=True, help="정기시험 추가", ):
                # 새 정기고사 회차 추가
                new_exam_num = len(st.session_state.eval_plan['regular']) + 1
                st.session_state.eval_plan['regular'][f"{new_exam_num}회"] = {
                    'max_score': 100,
                    'ratio': 0
                }
                set_config('plan', st.session_state.eval_plan)  # 통합 구조 동기화
                st.rerun()
        
        with col_item:
            if len(st.session_state.eval_plan['regular']) > 0:
                exam_to_delete = st.selectbox(
                    "",
                    sorted(st.session_state.eval_plan['regular'].keys(), 
                          key=lambda x: int(x.split('회')[0])),
                    format_func=lambda x: f"{x} 정기시험",
                    key="delete_exam_select_header",
                    label_visibility="collapsed"
                )
            else:
                exam_to_delete = None
        
        with col_del:
            if len(st.session_state.eval_plan['regular']) > 0 and exam_to_delete:
                if st.button("삭제", key="btn_del_regular_header", type="secondary", use_container_width=True):
                    del st.session_state.eval_plan['regular'][exam_to_delete]
                    st.rerun()
        
        # 선택된 정기고사 정보를 파란색 박스로 표시 (항목 선택 바로 아래)
        if len(st.session_state.eval_plan['regular']) > 0 and exam_to_delete:
            selected_exam_info = st.session_state.eval_plan['regular'][exam_to_delete]
            st.info(
                f"**{exam_to_delete} 정기시험** | "
                f"만점: {selected_exam_info['max_score']}점 | "
                f"반영비율: {selected_exam_info['ratio']}%"
            )
        
        # 정기고사가 있을 때만 테이블 표시
        if st.session_state.eval_plan['regular']:
            # 정기고사 항목들을 테이블 형식으로 표시
            regular_data = []
            regular_key_map = {}  # 표시용 -> 실제 키 매핑
            for exam_num in sorted(st.session_state.eval_plan['regular'].keys(), 
                                  key=lambda x: int(x.split('회')[0])):
                exam_info = st.session_state.eval_plan['regular'][exam_num]
                display_name = f"{exam_num} 정기시험"
                regular_key_map[display_name] = exam_num
                regular_data.append({
                    '구분': display_name,
                    '만점': exam_info['max_score'],
                    '반영비율(%)': exam_info['ratio']
                })
            
            # 정기고사 데이터 에디터
            df_regular = pd.DataFrame(regular_data)
            edited_regular = st.data_editor(
                df_regular,
                column_config={
                    '구분': st.column_config.TextColumn('구분', disabled=True),
                    '만점': st.column_config.NumberColumn('만점', min_value=1, step=1),
                    '반영비율(%)': st.column_config.NumberColumn('반영비율(%)', min_value=0, max_value=100, step=1)
                },
                hide_index=True,
                use_container_width=True,
                key="editor_regular",
                height=min(len(regular_data) * 35 + 40, 200)
            )
            
            # 정기고사 데이터 다시 세션 상태에 저장
            for idx, row in edited_regular.iterrows():
                display_name = row['구분']
                exam_num = regular_key_map[display_name]  # 실제 키로 매핑
                st.session_state.eval_plan['regular'][exam_num]['max_score'] = row['만점']
                st.session_state.eval_plan['regular'][exam_num]['ratio'] = row['반영비율(%)']
            set_config('plan', st.session_state.eval_plan)  # 통합 구조 동기화
        else:
            st.info("추가 버튼을 눌러 정기시험을 추가하세요.")

        st.markdown("---")

        # 수행평가 관리
        # 수행평가 반영비율 합계 계산
        total_perf_ratio = sum(item['ratio'] for item in st.session_state.eval_plan['performance'])
        
        # 헤더: 제목
        st.markdown(f"##### 📌 수행평가 구성 관계: **{total_perf_ratio}%**")
        
        # 버튼: 추가/항목선택/삭제 (정기시험과 동일한 구조) - 단축비 조정
        col_add_perf, col_item_perf, col_del_perf = st.columns([1.0, 1.8, 1.0])
        with col_add_perf:
            if st.button("추가", key="btn_add_perf", use_container_width=True, help="수행평가 추가"):
                # 새 수행평가 항목 추가
                new_perf_count = len(st.session_state.eval_plan['performance']) + 1
                st.session_state.eval_plan['performance'].append({
                    'name': f'수행평가{new_perf_count}',
                    'max_score': 100,
                    'ratio': 0
                })
                st.rerun()
        
        with col_item_perf:
            if len(st.session_state.eval_plan['performance']) > 0:
                perf_to_delete = st.selectbox(
                    "",
                    [item['name'] for item in st.session_state.eval_plan['performance']],
                    format_func=lambda x: x[:20] if len(x) > 20 else x,
                    key="delete_perf_select_header",
                    label_visibility="collapsed"
                )
            else:
                perf_to_delete = None
        
        with col_del_perf:
            if len(st.session_state.eval_plan['performance']) > 0 and perf_to_delete:
                if st.button("삭제", key="btn_del_perf_header", type="secondary", use_container_width=True):
                    st.session_state.eval_plan['performance'] = [
                        item for item in st.session_state.eval_plan['performance']
                        if item['name'] != perf_to_delete
                    ]
                    set_config('plan', st.session_state.eval_plan)  # 통합 구조 동기화
                    st.rerun()

        # 선택된 수행평가 정보를 파란색 박스로 표시 (항목 선택 바로 아래)
        if len(st.session_state.eval_plan['performance']) > 0 and perf_to_delete:
            selected_perf_info = next((item for item in st.session_state.eval_plan['performance'] if item['name'] == perf_to_delete), None)
            if selected_perf_info:
                st.info(
                    f"**{selected_perf_info['name']}** | "
                    f"만점: {selected_perf_info['max_score']}점 | "
                    f"반영비율: {selected_perf_info['ratio']}%"
                )
        
        # 수행평가가 있을 때만 테이블 표시
        if st.session_state.eval_plan['performance']:
            # 수행평가 데이터 에디터
            df_perf = pd.DataFrame(st.session_state.eval_plan['performance'])
            
            edited_perf = st.data_editor(
                df_perf,
                column_config={
                    'name': st.column_config.TextColumn('영역명', required=True),
                    'max_score': st.column_config.NumberColumn('만점', min_value=1, required=True, step=1),
                    'ratio': st.column_config.NumberColumn('반영비율(%)', min_value=0, max_value=100, required=True, step=1)
                },
                hide_index=True,
                use_container_width=True,
                key="editor_perf",
                height=min(len(df_perf) * 35 + 40, 200)
            )
            
            # 수행평가 데이터 다시 세션 상태에 저장
            st.session_state.eval_plan['performance'] = edited_perf.to_dict('records')
            set_config('plan', st.session_state.eval_plan)  # 통합 구조 동기화
        else:
            st.info("추가 버튼을 눌러 수행평가를 추가하세요.")

        st.markdown("---")

        # 총 반영비율 검증
        try:
            total_regular_ratio = sum([
                st.session_state.eval_plan['regular'][exam]['ratio']
                for exam in st.session_state.eval_plan['regular']
            ])
            total_perf_ratio = sum(item['ratio'] for item in st.session_state.eval_plan['performance'])
            total_ratio = total_regular_ratio + total_perf_ratio
            
            if total_ratio != 100:
                st.warning(
                    f"⚠️ **반영비율 합계: {total_ratio}% (100%가 되어야 합니다!)**\n\n"
                    f"정기시험: {total_regular_ratio}% | 수행평가: {total_perf_ratio}%"
                )
            else:
                st.success(f"✅ **반영비율 합계: {total_ratio}% (완료)**")
        except Exception as e:
            st.error(f"❌ 반영비율 데이터 오류: {str(e)}")

    with tab_target:
        st.markdown("##### 🎯 분석 대상 평가 선택")
        
        # 정기고사와 수행평가가 모두 없을 때
        if not st.session_state.eval_plan['regular'] and not st.session_state.eval_plan['performance']:
            st.warning("⚠️ 설정된 평가가 없습니다. '📝 평가 계획 관리' 탭에서 평가를 먼저 추가해주세요.")
            # 기본값 설정
            max_score, ratio, exam_category, exam_name = 100, 0, "정기고사", "평가 없음"
        else:
            # 평가 유형 선택 (한 줄 레이아웃)
            col_label, col_select, col_spacer = st.columns([1.2, 2.0, 0.8])
            with col_label:
                st.markdown("🔹 평가 유형")
            with col_select:
                # pending 설정이 있으면 그에 맞게 초기값 설정
                default_method = "정기시험"
                if st.session_state.get('selected_analysis_category') == 'performance':
                    default_method = "수행평가"
                elif st.session_state.get('selected_analysis_category') != 'regular' and not st.session_state.eval_plan['regular']:
                    default_method = "수행평가"
                
                default_method_index = ["정기시험", "수행평가"].index(default_method)
                selection_method = st.selectbox(
                    "평가 유형",
                    ["정기시험", "수행평가"],
                    index=default_method_index,
                    key="eval_category_select",
                    label_visibility="collapsed"
                )
            with col_spacer:
                st.write("")
            
            if selection_method == "정기시험":
                if not st.session_state.eval_plan['regular']:
                    st.error("설정된 정기시험이 없습니다. '📝 평가 계획 관리' 탭에서 추가해주세요.")
                    max_score, ratio, exam_category, exam_name = 100, 0, "정기고사", "평가 없음"
                else:
                    # 정기고사 목록 표시
                    regular_exams = sorted(st.session_state.eval_plan['regular'].keys(), 
                                          key=lambda x: int(x.split('회')[0]))
                    
                    # 선택 UI: 넓은 selectbox
                    selected_exam = st.selectbox(
                        "회차",
                        regular_exams,
                        format_func=lambda x: f"{x} 정기시험",  # "1회 정기시험"으로 표시
                        key="regular_exam_select",
                        label_visibility="collapsed"
                    )
                    
                    # 선택된 값 적용
                    max_score = st.session_state.eval_plan['regular'][selected_exam]['max_score']
                    ratio = st.session_state.eval_plan['regular'][selected_exam]['ratio']
                    exam_name = f"{selected_exam} 정기시험"
                    exam_category = "정기고사"
                    
                    st.session_state.selected_analysis_type = selected_exam
                    st.session_state.selected_analysis_category = 'regular'
                    set_config('selected.analysis_type', selected_exam)
                    set_config('selected.analysis_category', 'regular')
                    
                    # 선택된 평가 정보 카드 표시 (선택 바로 아래)
                    st.info(f"**{exam_name}** | 만점: {max_score}점 | 반영비율: {ratio}%")
                
            else:  # 수행평가
                if not st.session_state.eval_plan['performance']:
                    st.error("설정된 수행평가가 없습니다. '📝 평가 계획 관리' 탭에서 추가해주세요.")
                    max_score, ratio, exam_category, exam_name = 100, 0, "수행평가", "평가 없음"
                else:
                    perf_names = [item['name'] for item in st.session_state.eval_plan['performance']]
                    
                    # 선택 UI: 넓은 selectbox
                    selected_perf_name = st.selectbox(
                        "항목",
                        perf_names,
                        format_func=lambda x: x[:20] if len(x) > 20 else x,  # 20자로 축약
                        key="perf_select",
                        label_visibility="collapsed"
                    )
                    
                    selected_perf_item = next((item for item in st.session_state.eval_plan['performance'] 
                                              if item['name'] == selected_perf_name), None)
                    
                    if selected_perf_item:
                        max_score = selected_perf_item['max_score']
                        ratio = selected_perf_item['ratio']
                        exam_name = selected_perf_item['name']
                        exam_category = "수행평가"
                        
                        st.session_state.selected_analysis_type = selected_perf_name
                        st.session_state.selected_analysis_category = 'performance'
                        set_config('selected.analysis_type', selected_perf_name)
                        set_config('selected.analysis_category', 'performance')
                        
                        # 선택된 평가 정보 카드 표시 (선택 바로 아래)
                        st.info(f"**{exam_name}** | 만점: {max_score}점 | 반영비율: {ratio}%")

    
    st.markdown("---")
    
    # tab_target에서 선택한 값을 세션에서 가져오기 (폴백 처리)
    # 평가가 설정되어 있지 않으면 기본값 사용
    if st.session_state.selected_analysis_category == 'regular' and st.session_state.selected_analysis_type:
        selected_exam = st.session_state.selected_analysis_type
        if selected_exam in st.session_state.eval_plan['regular']:
            max_score = st.session_state.eval_plan['regular'][selected_exam]['max_score']
            ratio = st.session_state.eval_plan['regular'][selected_exam]['ratio']
            exam_name = f"{selected_exam} 정기시험"
            exam_category = "정기고사"
        else:
            max_score, ratio, exam_category, exam_name = 100, 0, "정기고사", "평가 없음"
    elif st.session_state.selected_analysis_category == 'performance' and st.session_state.selected_analysis_type:
        selected_perf = st.session_state.selected_analysis_type
        perf_item = next((item for item in st.session_state.eval_plan['performance'] 
                         if item['name'] == selected_perf), None)
        if perf_item:
            max_score = perf_item['max_score']
            ratio = perf_item['ratio']
            exam_name = perf_item['name']
            exam_category = "수행평가"
        else:
            max_score, ratio, exam_category, exam_name = 100, 0, "수행평가", "평가 없음"
    else:
        # 아무것도 선택되지 않았을 때: 첫 번째 평가 사용 또는 기본값
        if st.session_state.eval_plan['regular']:
            first_exam = sorted(st.session_state.eval_plan['regular'].keys(), 
                               key=lambda x: int(x.split('회')[0]))[0]
            max_score = st.session_state.eval_plan['regular'][first_exam]['max_score']
            ratio = st.session_state.eval_plan['regular'][first_exam]['ratio']
            exam_name = f"{first_exam} 정기시험"
            exam_category = "정기고사"
            st.session_state.selected_analysis_type = first_exam
            st.session_state.selected_analysis_category = 'regular'
            set_config('selected.analysis_type', first_exam)
            set_config('selected.analysis_category', 'regular')
        elif st.session_state.eval_plan['performance']:
            first_perf = st.session_state.eval_plan['performance'][0]
            max_score = first_perf['max_score']
            ratio = first_perf['ratio']
            exam_name = first_perf['name']
            exam_category = "수행평가"
            st.session_state.selected_analysis_type = first_perf['name']
            st.session_state.selected_analysis_category = 'performance'
            set_config('selected.analysis_type', first_perf['name'])
            set_config('selected.analysis_category', 'performance')
        else:
            max_score, ratio, exam_category, exam_name = 100, 0, "정기고사", "평가 없음"
    
    # 3단계: 분할점수 설정 (분할점수 기반 선택시만 표시)
    if analysis_basis == "분할점수 기반":
        st.subheader("3️⃣ 성취수준 분할점수")
        
        # 성취수준 수 선택 (3수준, 3수준+미도달, 5수준, 5수준+미도달)
        level_options = ["3수준 (A, B, C)", "3수준+미도달 (A, B, C, 미도달)", "5수준 (A, B, C, D, E)", "5수준+미도달 (A, B, C, D, E, 미도달)"]
        
        level_type = st.selectbox(
            "🎯 성취수준 구분",
            level_options,
            key="level_type"
        )
        
        st.caption("📊 등급 간 분할점수를 설정하세요 (총점 기준)")
        
        if level_type == "3수준 (A, B, C)":
            col1, col2 = st.columns(2)
            with col1:
                cut_AB = st.number_input("A/B 분할점수(점)", value=int(max_score * 0.8), min_value=0, max_value=max_score, 
                                        key="cut_AB_3", 
                                        help=f"이 점수 이상이면 A (80% = {int(max_score * 0.8)}점)")
            with col2:
                cut_BC = st.number_input("B/C 분할점수(점)", value=int(max_score * 0.6), min_value=0, max_value=max_score, 
                                        key="cut_BC_3",
                                        help=f"이 점수 이상이면 B (60% = {int(max_score * 0.6)}점)")
            cut_CD = None
            cut_DE = None
            cut_EI = None
        elif level_type == "3수준+미도달 (A, B, C, 미도달)":
            col1, col2 = st.columns(2)
            with col1:
                cut_AB = st.number_input("A/B 분할점수(점)", value=int(max_score * 0.8), min_value=0, max_value=max_score, 
                                        key="cut_AB_3m", 
                                        help=f"이 점수 이상이면 A (80% = {int(max_score * 0.8)}점)")
                cut_EI = st.number_input("C/미도달 분할점수(점)", value=int(max_score * 0.4), min_value=0, max_value=max_score, key="cut_EI_3m",
                                        help=f"이 점수 이상이면 C, 미만이면 미도달 (40% = {int(max_score * 0.4)}점)")
            with col2:
                cut_BC = st.number_input("B/C 분할점수(점)", value=int(max_score * 0.6), min_value=0, max_value=max_score, 
                                        key="cut_BC_3m",
                                        help=f"이 점수 이상이면 B (60% = {int(max_score * 0.6)}점)")
            cut_CD = None
            cut_DE = None
        elif level_type == "5수준 (A, B, C, D, E)":
            col1, col2 = st.columns(2)
            with col1:
                cut_AB = st.number_input("A/B 분할점수(점)", value=90, min_value=0, max_value=max_score, key="cut_AB", 
                                        help="이 점수 이상이면 A, 미만이면 B")
                cut_CD = st.number_input("C/D 분할점수(점)", value=70, min_value=0, max_value=max_score, key="cut_CD",
                                        help="이 점수 이상이면 C, 미만이면 D")
            with col2:
                cut_BC = st.number_input("B/C 분할점수(점)", value=80, min_value=0, max_value=max_score, key="cut_BC",
                                        help="이 점수 이상이면 B, 미만이면 C")
                cut_DE = st.number_input("D/E 분할점수(점)", value=60, min_value=0, max_value=max_score, key="cut_DE",
                                        help="이 점수 이상이면 D, 미만이면 E")
            cut_EI = None
        else:  # 5수준+미도달
            col1, col2 = st.columns(2)
            with col1:
                cut_AB = st.number_input("A/B 분할점수(점)", value=90, min_value=0, max_value=max_score, key="cut_AB_5i",
                                        help="이 점수 이상이면 A, 미만이면 B")
                cut_CD = st.number_input("C/D 분할점수(점)", value=70, min_value=0, max_value=max_score, key="cut_CD_5i",
                                        help="이 점수 이상이면 C, 미만이면 D")
                cut_EI = st.number_input("E/미도달 분할점수(점)", value=40, min_value=0, max_value=max_score, key="cut_EI",
                                        help="이 점수 이상이면 E, 미만이면 미도달")
            with col2:
                cut_BC = st.number_input("B/C 분할점수(점)", value=80, min_value=0, max_value=max_score, key="cut_BC_5i",
                                        help="이 점수 이상이면 B, 미만이면 C")
                cut_DE = st.number_input("D/E 분할점수(점)", value=60, min_value=0, max_value=max_score, key="cut_DE_5i",
                                        help="이 점수 이상이면 D, 미만이면 E")
        
        st.markdown("---")
    else:
        # 변수 초기화
        subject_name = ""
        cut_AB = 90
        cut_BC = 80
        cut_CD = 70
        cut_DE = 60
        cut_EI = 0
        level_type = "5수준 (A, B, C, D, E)"
    
    # 4단계: 파일 업로드 (분석 기준과 평가 유형에 따라 다름)
    st.subheader("4️⃣ 데이터 파일 업로드")
    
    # 파일 업로드 key 생성 (exam_name 대신 category와 type 기반)
    if exam_category == "정기고사":
        file_key_prefix = f"regular_{st.session_state.selected_analysis_type}"
    else:
        file_key_prefix = f"perf_{st.session_state.selected_analysis_type.replace(' ', '_')}"
    
    if analysis_basis == "분할점수 기반":
        # 분할점수 기반: 성적일람표 불필요
        if exam_category == "정기고사":
            st.caption(f"📌 {exam_name} 필수 파일")
            info_f = st.file_uploader(
                "📑 문항정보표 (Excel)",
                type=['xlsx'],
                key=f"info_{file_key_prefix}_score",
                help="NEIS에서 다운로드한 문항정보표를 선택하세요"
            )
            
            ans_fs = st.file_uploader(
                "✍️ 학생답정오표 (Excel)",
                type=['xlsx'],
                accept_multiple_files=True,
                key=f"ans_{file_key_prefix}_score",
                help="여러 학급의 정오표를 한 번에 선택할 수 있습니다"
            )
            
            grade_fs = []  # 성적일람표 불필요
            
            st.info("💡 **팁:** 학생 정오표에서 자동으로 성취도를 판정합니다.")
            
        else:  # 수행평가
            st.caption(f"📌 {exam_name} 필수 파일")
            info_f = st.file_uploader(
                "📑 평가기준표 (Excel)",
                type=['xlsx'],
                key=f"info_{file_key_prefix}_score",
                help="수행평가 항목과 배점이 포함된 평가기준표"
            )
            ans_fs = []
            grade_fs = []
            
            st.info("💡 **팁:** 수행평가는 평가기준표만 필요합니다.")
    
    else:  # 학기말 성취도 기반
        # 학기말 성취도 기반: 성적일람표 필수
        if exam_category == "정기고사":
            st.caption(f"📌 {exam_name} 필수 파일")
            info_f = st.file_uploader(
                "📑 문항정보표 (Excel)",
                type=['xlsx'],
                key=f"info_{file_key_prefix}_term",
                help="NEIS에서 다운로드한 문항정보표를 선택하세요"
            )
            
            ans_fs = st.file_uploader(
                "✍️ 학생답정오표 (Excel)",
                type=['xlsx'],
                accept_multiple_files=True,
                key=f"ans_{file_key_prefix}_term",
                help="여러 학급의 정오표를 한 번에 선택할 수 있습니다"
            )
            
            grade_fs = st.file_uploader(
                "📊 성적일람표 (Excel)",
                type=['xlsx'],
                accept_multiple_files=True,
                key=f"grade_{file_key_prefix}_term",
                help="성취도가 포함된 성적일람표를 선택하세요"
            )
            
        else:  # 수행평가
            st.caption(f"📌 {exam_name} 필수 파일")
            info_f = st.file_uploader(
                "📑 평가기준표 (Excel)",
                type=['xlsx'],
                key=f"info_{file_key_prefix}_term",
                help="수행평가 항목과 배점이 포함된 평가기준표"
            )
            ans_fs = []
            grade_fs = st.file_uploader(
                "📊 성적일람표 (Excel)",
                type=['xlsx'],
                accept_multiple_files=True,
                key=f"grade_{file_key_prefix}_term",
                help="수행평가 점수와 성취도가 포함된 성적일람표"
            )
    
    st.markdown("---")
    
    # 5단계: 분석 필터
    st.subheader("5️⃣ 분석 필터")
    
    # 기본 선택값 동적 설정
    if analysis_basis == "분할점수 기반":
        if level_type == "3수준 (A, B, C)":
            default_grades = ['A', 'B', 'C']
        elif level_type == "3수준+미도달 (A, B, C, 미도달)":
            default_grades = ['A', 'B', 'C', '미도달']
        elif level_type == "5수준 (A, B, C, D, E)":
            default_grades = ['A', 'B', 'C', 'D', 'E']
        else:  # 5수준+미도달
            default_grades = ['A', 'B', 'C', 'D', 'E', '미도달']
    else:
        default_grades = ['A', 'B', 'C', 'D', 'E']
        
    target_grade = st.multiselect(
        "🎯 분석 대상 성취도",
        ['A', 'B', 'C', 'D', 'E', '미도달'],
        default=default_grades,
        help="분석에 포함할 성취수준을 선택하세요"
    )
    
    # I(미도달) 표시를 'I'로 변환 -> 미도달은 그대로 미도달
    # target_grade = ['I' if x == 'I(미도달)' else x for x in target_grade]
    
    st.markdown("<br><br>", unsafe_allow_html=True)
    st.caption("ⓒ 2026. Data Analysis Pro for Teachers")

# --- 메인 대시보드 ---
st.title("🎓 성취평가 문항 분석 시스템")
st.markdown("#### 데이터 기반의 정확하고 세련된 문항 분석 보고서")

# 데이터 저장/불러오기 섹션
with st.expander("💾 분석 설정 저장/불러오기", expanded=False):
    col_save, col_load = st.columns(2)
    
    with col_save:
        # 현재 설정 저장
        if st.button("📥 설정 저장", use_container_width=True):
            settings = {
                'analysis_basis': st.session_state.get('analysis_basis', analysis_basis),
                'exam_category': st.session_state.get('exam_category', exam_category),
                'level_type': st.session_state.get('level_type'),
                'eval_plan': st.session_state.get('eval_plan', {}),
                'selected_analysis_type': st.session_state.get('selected_analysis_type', ''),
                'selected_analysis_category': st.session_state.get('selected_analysis_category', ''),
            }
            settings_json = json.dumps(settings, ensure_ascii=False, indent=2)
            st.download_button(
                label="⬇️ JSON 다운로드",
                data=settings_json,
                file_name=f"분석설정_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                mime="application/json",
                use_container_width=True
            )
            st.success("✅ 설정이 준비되었습니다. 위 버튼을 클릭하여 다운로드하세요.")
    
    with col_load:
        # 이전 설정 불러오기
        uploaded_settings = st.file_uploader("📤 설정 파일 업로드", type="json", key="settings_upload")
        if uploaded_settings:
            try:
                settings = json.load(uploaded_settings)
                # 임시 session_state에 저장 후 rerun 트리거
                st.session_state._pending_settings = settings
                st.success("✅ 설정을 불러왔습니다. 페이지를 새로고침합니다...")
                st.info("📋 **안내:** 설정 불러오기 후 문항정보표와 학생 정오표를 다시 업로드해 주세요.")
                st.rerun()
            except Exception as e:
                st.error(f"❌ 설정 파일을 읽을 수 없습니다: {str(e)}")

# 분석 기준에 따른 필요 파일 확인
if analysis_basis == "분할점수 기반":
    # 분할점수 기반: 성적일람표 불필요
    if exam_category == "정기고사":
        files_ready = info_f and ans_fs
    else:  # 수행평가
        files_ready = info_f
else:  # 학기말 성취도 기반
    # 학기말 성취도 기반: 성적일람표 필수
    if exam_category == "정기고사":
        files_ready = info_f and ans_fs and grade_fs
    else:  # 수행평가
        files_ready = info_f and grade_fs

if files_ready:
    with st.spinner('데이터를 분석 중입니다...'):
        try:
            if analysis_basis == "분할점수 기반":
                # 분할점수 기반: 성적일람표 없이 정오표만 사용
                result_pkg = load_and_merge_data(info_f, ans_fs, [])
            else:
                # 학기말 성취도 기반: 성적일람표 사용
                result_pkg = load_and_merge_data(info_f, ans_fs, grade_fs)
        except Exception as e:
            st.error(f"❌ 데이터 처리 중 오류 발생: {str(e)}")
            result_pkg = None
        
    if result_pkg and result_pkg[0] is not None and result_pkg[1] is not None and not result_pkg[1].empty:
        info_df, main_df = result_pkg
        
        # 분할점수 기반일 때 Achievement 컬럼 생성 (존재하지 않으면)
        if analysis_basis == "분할점수 기반":
            # 총점에 따라 성취도 판정
            def get_achievement_score_based(score):
                score = pd.to_numeric(score, errors='coerce')
                if pd.isna(score):
                    return '미도달' if cut_EI is not None else 'E' if cut_CD is not None else 'C'
                
                # 반올림 적용 (예: 89.5 → 90)
                score = round(score)
                
                if cut_CD is None and cut_EI is None:  # 3수준
                    if score >= cut_AB:
                        return 'A'
                    elif score >= cut_BC:
                        return 'B'
                    else:
                        return 'C'
                elif cut_CD is None and cut_EI is not None:  # 3수준+미도달
                    if score >= cut_AB:
                        return 'A'
                    elif score >= cut_BC:
                        return 'B'
                    elif score >= cut_EI:
                        return 'C'
                    else:
                        return '미도달'
                elif cut_CD is not None and cut_EI is None:  # 5수준
                    if score >= cut_AB:
                        return 'A'
                    elif score >= cut_BC:
                        return 'B'
                    elif score >= cut_CD:
                        return 'C'
                    elif score >= cut_DE:
                        return 'D'
                    else:
                        return 'E'
                else:  # 5수준+미도달
                    if score >= cut_AB:
                        return 'A'
                    elif score >= cut_BC:
                        return 'B'
                    elif score >= cut_CD:
                        return 'C'
                    elif score >= cut_DE:
                        return 'D'
                    elif score >= cut_EI:
                        return 'E'
                    else:
                        return '미도달'
            
            main_df['Achievement'] = main_df['Total_Score'].apply(get_achievement_score_based)
            
            # 분할점수 정보 표시
            if cut_CD is None and cut_EI is None:  # 3수준
                cut_info = f"A/B:{cut_AB}점 ({cut_AB/max_score*100:.0f}%), B/C:{cut_BC}점 ({cut_BC/max_score*100:.0f}%)"
            elif cut_CD is None and cut_EI is not None:  # 3수준+미도달
                cut_info = f"A/B:{cut_AB}점, B/C:{cut_BC}점, C/미도달:{cut_EI}점"
            elif cut_CD is not None and cut_EI is None:  # 5수준
                cut_info = f"A/B:{cut_AB}점, B/C:{cut_BC}점, C/D:{cut_CD}점, D/E:{cut_DE}점"
            else:  # 5수준+미도달
                cut_info = f"A/B:{cut_AB}점, B/C:{cut_BC}점, C/D:{cut_CD}점, D/E:{cut_DE}점, E/미도달:{cut_EI}점"
            
            st.success(f"✅ 분할점수 기반으로 성취도 판정 완료\n({cut_info})")
        else:
            # 학기말 성취도 기반: 기존 Achievement 컬럼 사용
            st.success(f"✅ 학기말 성취도를 기준으로 분석합니다")
        
        # 필터링
        main_df = main_df[main_df['Achievement'].isin(target_grade)]
        
        if main_df.empty:
            st.warning("선택한 성취도에 해당하는 학생이 없습니다.")
        else:
            # 통계 계산
            # '.' 문자나 기타 문자를 처리하기 위해 1/0 매핑 시 오류 방지
            item_cols = [f'Item_{i}' for i in range(1, 17)]
            
            # 안전한 이진 행렬 변환 (Applymap 대신 apply 사용 권장)
            def safe_binary(x):
                return 1 if str(x).strip() == '.' else 0
                
            binary_matrix = main_df[item_cols].applymap(safe_binary)
            
            # 신뢰도(KR-20) 계산 - 분모 0 방지
            var_sum = binary_matrix.var().sum()
            total_var = binary_matrix.sum(axis=1).var()
            
            if total_var == 0 or np.isnan(total_var):
                alpha = 0.0
            else:
                alpha = (16/15) * (1 - var_sum / total_var)

            # [지표 계산] 문항 통계 (정답률, 변별도)
            top_len = max(1, int(len(main_df)*0.25))
            top_25 = main_df.nlargest(top_len, 'Total_Score')
            bot_25 = main_df.nsmallest(top_len, 'Total_Score')
            
            discrimination_scores = {}
            item_p_scores = {}

            for i in range(1, 17):
                col = f'Item_{i}'
                # 상위권 정답률 - 하위권 정답률
                p_top = (top_25[col].astype(str) == '.').mean()
                p_bot = (bot_25[col].astype(str) == '.').mean()
                discrimination_scores[i] = p_top - p_bot
                item_p_scores[i] = (main_df[col].astype(str) == '.').mean()
            
            # 문항 분석 DataFrame 생성 (공통 사용)
            item_stats_list = []
            for i in range(1, 17):
                item_stats_list.append({
                    'No': i, 
                    '정답률(P)': item_p_scores[i], 
                    '변별도(D)': discrimination_scores[i]
                })
            res_df = pd.merge(pd.DataFrame(item_stats_list), info_df[['No', 'Exp_Diff', 'Score', 'Standard']], on='No')
            res_df['Score'] = pd.to_numeric(res_df['Score'], errors='coerce').fillna(0)

            # 탭 구성 (6개)
            tab_data, tab_summary, tab_item, tab_dist, tab_std, tab_report = st.tabs([
                "데이터", "전체 성취도 분석", "문항 분석", 
                "성취수준별 답지반응-부분점수 분포", "성취기준 분석 결과", "분석 리포트"
            ])

            # --- [Tab 1] Data ---
            with tab_data:
                st.subheader("📊 데이터 미리보기")
                st.caption("업로드된 문항정보표와 병합된 학생 성적 데이터입니다.")
                
                # 문항정보표 (위)
                st.write("#### 📑 문항정보표")
                info_display = info_df.copy()
                info_rename = {
                    'No': '문항번호', 'Score': '배점', 'Correct_Ans': '정답', 
                    'Exp_Diff': '예상난이도', 'Standard': '성취기준',
                    'Hard': '상', 'Medium': '중', 'Easy': '하'
                }
                info_display = info_display.rename(columns={k: v for k, v in info_rename.items() if k in info_display.columns})
                info_display = info_display.fillna('')
                info_display = info_display.replace('None', '')
                
                # 성취기준 컬럼에 왼쪽 정렬 클래스 적용
                def make_html_table(df, left_align_cols=None):
                    """DataFrame을 HTML 테이블로 변환 (특정 컬럼 왼쪽 정렬)"""
                    left_align_cols = left_align_cols or []
                    html = '<table class="styled-table">'
                    # Header
                    html += '<thead><tr>'
                    for col in df.columns:
                        html += f'<th>{col}</th>'
                    html += '</tr></thead>'
                    # Body
                    html += '<tbody>'
                    for _, row in df.iterrows():
                        html += '<tr>'
                        for col in df.columns:
                            val = row[col]
                            if col in left_align_cols:
                                html += f'<td class="left-align">{val}</td>'
                            else:
                                html += f'<td>{val}</td>'
                        html += '</tr>'
                    html += '</tbody></table>'
                    return html
                
                info_html = make_html_table(info_display, left_align_cols=['성취기준'])
                st.markdown(f'<div class="table-container">{info_html}</div>', unsafe_allow_html=True)
                
                st.divider()
                
                # 학생 성적 데이터 (아래)
                st.write("#### 🧑‍🎓 학생 성적 데이터")
                
                # 선택된 평가 정보 표시
                basis_str = "분할점수 기반" if analysis_basis == "분할점수 기반" else "학기말 성취도 기반"
                st.caption(f"📌 **선택된 평가:** {exam_name} | **분석 기준:** {basis_str} | **만점:** {max_score}점 | **반영비율:** {ratio}%")
                
                # 성취도 분포 시각화 (테이블 위에 표시)
                st.write("**점수 분포 분석**")
                
                # 그래프 유형 선택 (기본값을 "총점"으로 설정)
                score_type = st.selectbox("표시할 점수 유형을 선택하세요", ["총점", "학기말 원점수"], index=0)
                
                # 성취수준 색상 정의 (미도달 I 추가)
                achievement_colors = {
                    'A': '#1DD1A1',  # 초록색
                    'B': '#54A0FF',  # 파랑색
                    'C': '#FFD93D',  # 노랑색
                    'D': '#FF6348',  # 주황색
                    'E': '#EE5A6F',  # 빨강색
                    '미도달': '#868E96'   # 회색 (미도달)
                }
                
                # 분석용 데이터 준비
                dist_df = main_df.copy()
                
                if score_type == "학기말 원점수":
                    # 학기말 원점수 계산 (총점 × 반영비율%)
                    dist_df['Total_Score_Num'] = pd.to_numeric(dist_df['Total_Score'], errors='coerce').fillna(0)
                    dist_df['학기말 원점수'] = (dist_df['Total_Score_Num'] * ratio / 100).round(1)
                    score_df = dist_df[['학기말 원점수', 'Achievement']].dropna()
                    score_df = score_df.rename(columns={'학기말 원점수': '점수', 'Achievement': '성취수준'})
                    x_axis = '점수'
                    title_text = "<b>학기말 원점수 분포 (성취수준별)</b>"
                    max_semester_score = (max_score * ratio / 100)
                    nbins = max(3, int(max_semester_score / 10))  # 10점 간격으로 변경 (더 넓은 막대)
                    xaxis_range = [0, max_semester_score]
                else:  # 총점
                    dist_df['총점'] = pd.to_numeric(dist_df['Total_Score'], errors='coerce')
                    score_df = dist_df[['총점', 'Achievement']].dropna()
                    score_df = score_df.rename(columns={'총점': '점수', 'Achievement': '성취수준'})
                    x_axis = '점수'
                    title_text = "<b>총점 분포 (성취수준별)</b>"
                    nbins = 10  # 100점 ÷ 10점 간격 = 10개
                    xaxis_range = [0, 100]
                
                # 성취수준 순서 정렬 (level_type에 따라 다름)
                if level_type == "3수준 (A, B, C)":
                    all_levels = ['A', 'B', 'C']
                elif level_type == "5수준+미도달 (A, B, C, D, E, 미도달)":
                    all_levels = ['A', 'B', 'C', 'D', 'E', '미도달']
                else:  # 5수준
                    all_levels = ['A', 'B', 'C', 'D', 'E']
                
                available_levels = [level for level in all_levels if level in score_df['성취수준'].unique()]
                score_df['성취수준'] = pd.Categorical(score_df['성취수준'], categories=available_levels, ordered=True)
                
                # 점수 범위별로 binning (10점 간격)
                bins = np.arange(int(xaxis_range[0]), int(xaxis_range[1]) + 10, 10)
                score_df['bin'] = pd.cut(score_df['점수'], bins=bins)
                
                # 각 bin별로 성취수준 카운트
                bin_counts = score_df.groupby(['bin', '성취수준']).size().unstack(fill_value=0)
                bin_labels = [f"{int(interval.left)}-{int(interval.right)}" for interval in bin_counts.index]
                
                # go.Figure로 그룹 막대 그래프 생성
                fig_dist = go.Figure()
                
                for level in available_levels:
                    if level in bin_counts.columns:
                        hover_texts = [f"성취수준: {level}\n점수 범위: {label}\n학생 수: {int(count)}명" 
                                      for label, count in zip(bin_labels, bin_counts[level])]
                        fig_dist.add_trace(go.Bar(
                            x=bin_labels,
                            y=bin_counts[level],
                            name=level,
                            hovertext=hover_texts,
                            hoverinfo="text",
                            marker=dict(
                                color=achievement_colors[level],
                                line=dict(color='rgba(0,0,0,0.4)', width=2)
                            )
                        ))
                
                fig_dist.update_layout(
                    title=title_text,
                    paper_bgcolor="rgba(0,0,0,0)", 
                    plot_bgcolor="rgba(240,242,246,0.3)", 
                    font_family="Pretendard",
                    height=400,
                    showlegend=True,
                    xaxis_title="점수",
                    yaxis_title="학생수",
                    barmode='group',
                    bargap=0.0,
                    bargroupgap=0.0,
                    margin=dict(l=60, r=120, t=80, b=60),
                    legend=dict(
                        title="성취수준",
                        orientation="v",
                        yanchor="top",
                        y=0.99,
                        xanchor="right",
                        x=0.99,
                        traceorder="normal"
                    )
                )
                
                st.plotly_chart(fig_dist, use_container_width=True)
                
                main_display = main_df.copy()
                
                # 선택형 점수 계산 (문항 1-9번: 선택형으로 가정)
                select_cols = [f'Item_{i}' for i in range(1, 10)]
                select_cols = [c for c in select_cols if c in main_display.columns]
                
                # 서답형 점수 계산 (문항 10-16번: 서답형으로 가정)
                essay_cols = [f'Item_{i}' for i in range(10, 17)]
                essay_cols = [c for c in essay_cols if c in main_display.columns]
                
                # 점수 계산을 위한 info_df의 배점 참조
                score_map = {}
                for _, row in info_df.iterrows():
                    item_no = int(row['No']) if pd.notna(row['No']) else 0
                    item_score = pd.to_numeric(row['Score'], errors='coerce')
                    if item_no > 0 and pd.notna(item_score):
                        score_map[f'Item_{item_no}'] = item_score
                
                # 선택형 점수 합계
                def calc_select_score(row):
                    total = 0
                    for col in select_cols:
                        if col in score_map and str(row[col]).strip() == '.':
                            total += score_map[col]
                    return total
                    
                # 서답형 점수 합계
                def calc_essay_score(row):
                    total = 0
                    for col in essay_cols:
                        if col in score_map and str(row[col]).strip() == '.':
                            total += score_map[col]
                    return total
                
                main_display['Select_Score'] = main_display.apply(calc_select_score, axis=1)
                main_display['Essay_Score'] = main_display.apply(calc_essay_score, axis=1)
                
                # 학기말 원점수 계산 (총점 × 반영비율%)
                main_display['Total_Score_Num'] = pd.to_numeric(main_display['Total_Score'], errors='coerce').fillna(0)
                main_display['Semester_Score'] = (main_display['Total_Score_Num'] * ratio / 100).round(1)
                
                # 컬럼 순서 재정렬: 학번, 이름, 문항들, 선택형점수, 서답형점수, 총점, 학기말 원점수, 성취수준
                col_order = ['강의실', 'ID', 'Name'] + [f'Item_{i}' for i in range(1, 17)] + ['Select_Score', 'Essay_Score', 'Total_Score', 'Semester_Score', 'Achievement']
                col_order = [c for c in col_order if c in main_display.columns]
                main_display = main_display[col_order]
                
                main_rename = {
                    '강의실': '수강반',
                    'ID': '학번', 'Name': '이름', 
                    'Select_Score': '선택형점수', 'Essay_Score': '서답형점수',
                    'Total_Score': '총점', 'Semester_Score': '학기말 원점수', 
                    'Achievement': '성취수준'
                }
                for i in range(1, 17):
                    main_rename[f'Item_{i}'] = f'문{i}'
                main_display = main_display.rename(columns={k: v for k, v in main_rename.items() if k in main_display.columns})
                
                # 숫자 컬럼 소수점 처리
                if '총점' in main_display.columns:
                    main_display['총점'] = pd.to_numeric(main_display['총점'], errors='coerce').round(1)
                
                main_display = main_display.fillna('')
                main_display = main_display.replace('None', '')
                
                # 학생 성적 데이터 표 위 설명
                st.info("ℹ️ **성취수준은 총점을 반올림한 값을 기준으로 판단됩니다.**")
                
                # 정렬 가능한 데이터프레임으로 표시
                st.dataframe(
                    main_display,
                    use_container_width=True,
                    height=450
                )
                
                # 다운로드 및 출력 버튼
                col_download, col_print = st.columns([1, 1])
                
                with col_download:
                    # 엑셀 다운로드 버튼 (포매팅 적용)
                    excel_data = format_excel_file(main_display, exam_name, basis_str, max_score, ratio)
                    
                    filename = f"학생성적데이터_{exam_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                    st.download_button(
                        label="📥 엑셀 다운로드",
                        data=excel_data,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                
                with col_print:
                    # 인쇄 HTML 다운로드
                    print_html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>학생 성적 데이터 - {exam_name}</title>
    <style>
        @media print {{
            body {{ margin: 1cm; }}
            table {{ page-break-inside: auto; }}
            tr {{ page-break-inside: avoid; page-break-after: auto; }}
        }}
        body {{
            font-family: 'Malgun Gothic', '맑은 고딕', sans-serif;
            font-size: 10pt;
            margin: 20px;
        }}
        h1 {{
            text-align: center;
            font-size: 18pt;
            margin-bottom: 10px;
            margin-top: 0;
        }}
        .info {{
            text-align: center;
            margin-bottom: 20px;
            font-size: 11pt;
            border-bottom: 1px solid #ccc;
            padding-bottom: 10px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 9pt;
        }}
        th, td {{
            border: 1px solid #999;
            padding: 4px 6px;
            text-align: center;
        }}
        th {{
            background-color: #d9d9d9;
            font-weight: bold;
        }}
        tr:nth-child(even) {{
            background-color: #f5f5f5;
        }}
        .print-button {{
            text-align: center;
            margin-top: 20px;
        }}
        button {{
            padding: 10px 20px;
            font-size: 12pt;
            background-color: #4CAF50;
            color: white;
            border: none;
            border-radius: 4px;
            cursor: pointer;
        }}
        button:hover {{
            background-color: #45a049;
        }}
    </style>
</head>
<body>
    <h1>📊 학생 성적 데이터</h1>
    <div class="info">
        <p><strong>평가명:</strong> {exam_name} | <strong>분석 기준:</strong> {basis_str}</p>
        <p><strong>만점:</strong> {max_score}점 | <strong>반영비율:</strong> {ratio}% | <strong>출력일시:</strong> {datetime.now().strftime('%Y년 %m월 %d일 %H:%M:%S')}</p>
    </div>
    {main_display.to_html(index=False, classes='data-table')}
    <div class=\"print-button\">
        <button onclick=\"window.print()\">🖨️ 인쇄</button>
    </div>
    <script>
        document.addEventListener('DOMContentLoaded', function() {{
            // 사용자가 버튼을 클릭하지 않으면 자동으로 인쇄 대화상자 열기 (선택사항)
            // window.print();
        }});
    </script>
</body>
</html>"""
                    
                    html_filename = f"학생성적데이터_인쇄_{exam_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html"
                    st.download_button(
                        label="🖨️ 인쇄용 HTML 다운로드",
                        data=print_html,
                        file_name=html_filename,
                        mime="text/html",
                        use_container_width=True
                    )

            # --- [Tab 2] 전체 성취도 분석 ---
            with tab_summary:
                # 강의실 정보 추출
                if '강의실' in main_df.columns and not main_df['강의실'].isna().all():
                    classroom_list = sorted([str(x) for x in main_df['강의실'].dropna().unique()])
                    
                    # 전체/강의실별 선택
                    filter_type = st.radio(
                        "분석 범위 선택",
                        options=["전체"] + [f"{r}수강반" for r in classroom_list],
                        horizontal=True,
                        key="classroom_filter"
                    )
                    
                    # 선택에 따른 데이터 필터링
                    if filter_type == "전체":
                        analysis_df = main_df
                        analysis_title = "전체"
                    else:
                        selected_room = filter_type.replace("수강반", "")
                        analysis_df = main_df[main_df['강의실'].astype(str) == selected_room]
                        analysis_title = filter_type
                else:
                    # 강의실 정보가 없으면 전체만 표시
                    analysis_df = main_df
                    analysis_title = "전체"
                
                st.divider()
                
                # 1. 상단 메트릭
                m1, m2, m3 = st.columns(3)
                m1.metric("학생 수", f"{len(analysis_df)}명")
                m2.metric("평가 종류", exam_name)
                m3.metric("분석 범위", analysis_title)

                st.divider()

                # 2. 성취도 분포 차트
                # 데이터 준비
                dist = analysis_df['Achievement'].value_counts().reset_index()
                dist.columns = ['성취수준', '학생 수']
                
                # 성취수준 순서 정렬 (level_type에 따라 다름)
                if level_type == "3수준 (A, B, C)":
                    level_order = ['A', 'B', 'C']
                elif level_type == "5수준+미도달 (A, B, C, D, E, I)":
                    level_order = ['A', 'B', 'C', 'D', 'E', 'I']
                else:  # 5수준
                    level_order = ['A', 'B', 'C', 'D', 'E']
                
                dist['성취수준'] = pd.Categorical(dist['성취수준'], categories=level_order, ordered=True)
                dist = dist.sort_values('성취수준', ascending=False)
                
                # 비율 계산
                total_students = dist['학생 수'].sum()
                dist['비율(%)'] = (dist['학생 수'] / total_students * 100).round(1)
                
                # 텍스트 라벨 생성
                text_labels = [f"{pct:.1f}% ({cnt}명)" for pct, cnt in zip(dist['비율(%)'], dist['학생 수'])]
                
                # 동적 우측 margin 계산
                max_label_length = max(len(label) for label in text_labels)
                right_margin = 80 + max_label_length * 10
                
                # X축 범위 동적 계산
                max_ratio = dist['비율(%)'].max()
                xaxis_max = max(60, max_ratio * 1.4)
                
                # 좌우 2개 그래프
                col_left, col_right = st.columns(2)
                
                with col_left:
                    # 수평 막대 그래프 (성취수준별 학생 수)
                    fig_count = go.Figure()
                    
                    fig_count.add_trace(go.Bar(
                        x=dist['비율(%)'],
                        y=dist['성취수준'],
                        orientation='h',
                        marker=dict(
                            color=[achievement_colors.get(level, '#999999') for level in dist['성취수준']]
                        ),
                        text=text_labels,
                        textposition='outside',
                        hovertemplate="<b>%{y}</b><br>비율: %{x:.1f}%<br>학생 수: %{customdata}명<extra></extra>",
                        customdata=dist['학생 수']
                    ))
                    
                    fig_count.update_layout(
                        title="<b>성취수준별 학생 수</b>",
                        paper_bgcolor="rgba(0,0,0,0)", 
                        plot_bgcolor="rgba(240,242,246,0.2)",
                        font_family="Pretendard",
                        height=400,
                        showlegend=False,
                        font=dict(size=11),
                        xaxis_title="비율(%)",
                        yaxis_title="성취수준",
                        margin=dict(l=80, r=right_margin),
                        xaxis=dict(range=[0, xaxis_max], showgrid=True, gridwidth=1, gridcolor='rgba(200,200,200,0.2)'),
                        yaxis=dict(tickfont=dict(size=12))
                    )
                    
                    st.plotly_chart(fig_count, use_container_width=True, key="tab2_count")
                
                with col_right:
                    # 평가 점수 선택
                    score_type = st.radio(
                        "평가 점수 선택",
                        options=["1회 정기시험", "학기말 원점수"],
                        horizontal=True,
                        key="score_type_radio"
                    )
                    
                    # 수직 막대 그래프 (성취수준별 평균)
                    analysis_df['Total_Score_Num'] = pd.to_numeric(analysis_df['Total_Score'], errors='coerce').fillna(0)
                    
                    # 선택된 유형에 따라 데이터 계산
                    if score_type == "1회 정기시험":
                        analysis_df['선택점수'] = analysis_df['Total_Score_Num']  # 반영비율 미적용 (100점 만점)
                        y_title = "1회 정기시험 평균"
                        graph_title = "<b>성취수준별 평균 (1회 정기시험)</b>"
                        y_max = 100  # 정기시험은 100점 만점
                    else:
                        analysis_df['선택점수'] = (analysis_df['Total_Score_Num'] * ratio / 100).round(1)  # 반영비율 적용
                        y_title = "학기말 원점수 평균"
                        graph_title = "<b>성취수준별 평균 (학기말 원점수)</b>"
                        y_max = (max_score * ratio / 100)  # 반영비율 적용된 만점
                    
                    avg_data = analysis_df.groupby('Achievement')['선택점수'].agg(['mean', 'std']).reset_index()
                    avg_data.columns = ['성취수준', '평균', '표준편차']
                    avg_data['평균'] = avg_data['평균'].round(2)
                    avg_data['표준편차'] = avg_data['표준편차'].round(2)
                    
                    # 성취수준 순서 정렬 (level_type에 따라)
                    if level_type == "3수준 (A, B, C)":
                        level_order = ['A', 'B', 'C']
                    elif level_type == "5수준+미도달 (A, B, C, D, E, I)":
                        level_order = ['A', 'B', 'C', 'D', 'E', 'I']
                    else:  # 5수준
                        level_order = ['A', 'B', 'C', 'D', 'E']
                    
                    avg_data['성취수준'] = pd.Categorical(avg_data['성취수준'], categories=level_order, ordered=True)
                    avg_data = avg_data.sort_values('성취수준')
                    
                    fig_avg = go.Figure()
                    
                    # 호버 텍스트 준비
                    hover_text = [
                        f"<b style='font-size:14px'>{row['성취수준']}</b><br>" +
                        f"<span style='font-size:13px'>평균: <b>{row['평균']:.2f}</b>점</span><br>" +
                        f"<span style='font-size:13px'>표준편차: <b>{row['표준편차']:.2f}</b></span>"
                        for _, row in avg_data.iterrows()
                    ]
                    
                    fig_avg.add_trace(go.Bar(
                        x=avg_data['성취수준'],
                        y=avg_data['평균'],
                        marker=dict(
                            color=[achievement_colors.get(level, '#999999') for level in avg_data['성취수준']],
                            line=dict(color='rgba(0,0,0,0.3)', width=1.5)
                        ),
                        text=[f"<b>{val:.2f}</b>" for val in avg_data['평균']],
                        textposition='outside',
                        textfont=dict(size=13, color='black'),
                        hovertemplate="%{customdata}<extra></extra>",
                        customdata=hover_text
                    ))
                    
                    fig_avg.update_layout(
                        title=graph_title,
                        paper_bgcolor="rgba(0,0,0,0)", 
                        plot_bgcolor="rgba(240,242,246,0.2)",
                        font_family="Pretendard",
                        height=400,
                        showlegend=False,
                        font=dict(size=11),
                        xaxis_title="성취수준",
                        yaxis_title=y_title,
                        margin=dict(l=60, r=60, t=80, b=60),
                        yaxis=dict(range=[0, y_max * 1.15], showgrid=True, gridwidth=1, gridcolor='rgba(200,200,200,0.2)'),
                        xaxis=dict(tickfont=dict(size=13))
                    )
                    
                    st.plotly_chart(fig_avg, use_container_width=True, key="tab2_avg")

                st.divider()
                
                # 3. 전체 통계 및 수강반별 통계
                st.subheader("📊 성적 통계 요약")
                
                # 3-1. 전체 통계 (분석 범위에 따라)
                st.markdown(f"### 📈 {analysis_title} 통계")
                overall_cols = st.columns(4)
                
                total_score_data = analysis_df['Total_Score_Num']
                overall_mean = total_score_data.mean()
                overall_std = total_score_data.std()
                overall_mean_ratio = (overall_mean * ratio / 100)
                overall_std_ratio = (overall_std * ratio / 100)
                
                overall_cols[0].metric("1회 정기시험 평균", f"{overall_mean:.2f}점")
                overall_cols[1].metric("1회 정기시험 표준편차", f"{overall_std:.2f}")
                overall_cols[2].metric("학기말 원점수 평균", f"{overall_mean_ratio:.2f}점")
                overall_cols[3].metric("학기말 원점수 표준편차", f"{overall_std_ratio:.2f}")
                
                st.divider()
                
                # 3-2. 수강반별 통계 (강의실 정보가 있고, 전체 분석일 때만)
                if '강의실' in main_df.columns and not main_df['강의실'].isna().all() and analysis_title == "전체":
                    st.markdown("### 📚 수강반별 통계")
                    
                    # 수강반별 집계
                    class_stats_list = []
                    for classroom in sorted([str(x) for x in main_df['강의실'].dropna().unique()]):
                        class_subset = pd.to_numeric(
                            main_df[main_df['강의실'].astype(str) == classroom]['Total_Score'], 
                            errors='coerce'
                        ).fillna(0)
                        
                        if len(class_subset) > 0:
                            class_mean = class_subset.mean()
                            class_std = class_subset.std()
                            class_mean_ratio = (class_mean * ratio / 100)
                            class_std_ratio = (class_std * ratio / 100)
                            
                            class_stats_list.append({
                                '수강반': f"{classroom}수강반",
                                '학생수': len(class_subset),
                                '정기시험평균': class_mean,
                                '정기시험표준편차': class_std,
                                '원점수평균': class_mean_ratio,
                                '원점수표준편차': class_std_ratio
                            })
                    
                    if class_stats_list:
                        class_stats_df = pd.DataFrame(class_stats_list).round(2)
                        
                        # HTML 테이블 생성
                        def make_class_stats_table(df):
                            html = '<table class="styled-table" style="width:100%; border-collapse: collapse;">'
                            # 2-level Header
                            html += '<thead>'
                            html += '<tr style="text-align: center;">'
                            html += '<th rowspan="2" style="vertical-align: middle; border: 1px solid #ddd; padding: 12px;">수강반</th>'
                            html += '<th rowspan="2" style="vertical-align: middle; border: 1px solid #ddd; padding: 12px;">학생수</th>'
                            html += '<th colspan="2" style="border: 1px solid #ddd; padding: 12px; background-color: #e3f2fd;">1회 정기시험</th>'
                            html += '<th colspan="2" style="border: 1px solid #ddd; padding: 12px; background-color: #e3f2fd;">학기말 원점수</th>'
                            html += '</tr>'
                            html += '<tr style="text-align: center;">'
                            html += '<th style="border: 1px solid #ddd; padding: 12px; background-color: #e3f2fd;">평균</th>'
                            html += '<th style="border: 1px solid #ddd; padding: 12px; background-color: #e3f2fd;">표준편차</th>'
                            html += '<th style="border: 1px solid #ddd; padding: 12px; background-color: #e3f2fd;">평균</th>'
                            html += '<th style="border: 1px solid #ddd; padding: 12px; background-color: #e3f2fd;">표준편차</th>'
                            html += '</tr>'
                            html += '</thead>'
                            html += '<tbody>'
                            for _, row in df.iterrows():
                                html += '<tr style="text-align: center;">'
                                html += f'<td style="border: 1px solid #ddd; padding: 10px;"><b>{row["수강반"]}</b></td>'
                                html += f'<td style="border: 1px solid #ddd; padding: 10px;">{int(row["학생수"])}</td>'
                                html += f'<td style="border: 1px solid #ddd; padding: 10px;">{row["정기시험평균"]:.2f}</td>'
                                html += f'<td style="border: 1px solid #ddd; padding: 10px;">{row["정기시험표준편차"]:.2f}</td>'
                                html += f'<td style="border: 1px solid #ddd; padding: 10px;">{row["원점수평균"]:.2f}</td>'
                                html += f'<td style="border: 1px solid #ddd; padding: 10px;">{row["원점수표준편차"]:.2f}</td>'
                                html += '</tr>'
                            html += '</tbody></table>'
                            return html
                        
                        class_stats_html = make_class_stats_table(class_stats_df)
                        st.markdown(class_stats_html, unsafe_allow_html=True)
                    
                    st.divider()
                
                # 4. 성취수준별 통계 요약
                st.subheader(f"📊 성취수준별 통계 ({analysis_title})")
                
                # 각 성취수준별 통계 계산 (명확하게)
                stat_list = []
                for achievement in sorted(analysis_df['Achievement'].unique()):
                    subset = analysis_df[analysis_df['Achievement'] == achievement]['Total_Score_Num']
                    환산점수_subset = (subset * ratio / 100)  # 반영비율 적용
                    
                    stat_dict = {
                        '성취수준': achievement,
                        '학생수': len(subset),
                        '비율(%)': (len(subset) / len(analysis_df) * 100),
                        '정기시험평균': subset.mean(),  # 반영비율 미적용 (100점 만점 기준)
                        '정기시험표준편차': subset.std(),  # 반영비율 미적용
                        '원점수평균': 환산점수_subset.mean(),  # 반영비율 적용 (ratio점 만점 기준)
                        '원점수표준편차': 환산점수_subset.std()  # 반영비율 적용
                    }
                    stat_list.append(stat_dict)
                
                stat_summary = pd.DataFrame(stat_list).round(2)
                
                # level_type에 따라 표시할 수준 결정
                if level_type == "3수준 (A, B, C)":
                    level_order_stat = ['A', 'B', 'C']
                elif level_type == "5수준+미도달 (A, B, C, D, E, I)":
                    level_order_stat = ['A', 'B', 'C', 'D', 'E', 'I']
                else:  # 5수준
                    level_order_stat = ['A', 'B', 'C', 'D', 'E']
                
                stat_summary = stat_summary[stat_summary['성취수준'].isin(level_order_stat)]
                
                # 성취수준 순서 정렬 (A가 위)
                stat_summary['성취수준'] = pd.Categorical(stat_summary['성취수준'], categories=level_order_stat, ordered=True)
                stat_summary = stat_summary.sort_values('성취수준')
                
                # 컬럼 순서: 성취수준, 학생수, 비율, 정기시험평균, 정기시험표준편차, 원점수평균, 원점수표준편차
                stat_summary = stat_summary[['성취수준', '학생수', '비율(%)', '정기시험평균', '정기시험표준편차', '원점수평균', '원점수표준편차']]
                
                # 숫자 컬럼의 NaN 값을 0으로 대체
                numeric_cols = ['학생수', '비율(%)', '정기시험평균', '정기시험표준편차', '원점수평균', '원점수표준편차']
                stat_summary[numeric_cols] = stat_summary[numeric_cols].fillna(0)
                
                # 멀티레벨 헤더를 가진 HTML 테이블 생성
                def make_multi_header_table(df):
                    html = '<table class="styled-table" style="width:100%; border-collapse: collapse;">'
                    # 2-level Header
                    html += '<thead>'
                    # 첫 번째 행 (상위 헤더)
                    html += '<tr style="text-align: center;">'
                    html += '<th rowspan="2" style="vertical-align: middle; border: 1px solid #ddd; padding: 12px;">성취수준</th>'
                    html += '<th rowspan="2" style="vertical-align: middle; border: 1px solid #ddd; padding: 12px;">학생수</th>'
                    html += '<th rowspan="2" style="vertical-align: middle; border: 1px solid #ddd; padding: 12px;">비율(%)</th>'
                    html += '<th colspan="2" style="border: 1px solid #ddd; padding: 12px; background-color: #f8f9fa;">1회 정기시험</th>'
                    html += '<th colspan="2" style="border: 1px solid #ddd; padding: 12px; background-color: #f8f9fa;">학기말 원점수</th>'
                    html += '</tr>'
                    # 두 번째 행 (하위 헤더)
                    html += '<tr style="text-align: center;">'
                    html += '<th style="border: 1px solid #ddd; padding: 12px; background-color: #f8f9fa;">평균</th>'
                    html += '<th style="border: 1px solid #ddd; padding: 12px; background-color: #f8f9fa;">표준편차</th>'
                    html += '<th style="border: 1px solid #ddd; padding: 12px; background-color: #f8f9fa;">평균</th>'
                    html += '<th style="border: 1px solid #ddd; padding: 12px; background-color: #f8f9fa;">표준편차</th>'
                    html += '</tr>'
                    html += '</thead>'
                    # Body
                    html += '<tbody>'
                    for _, row in df.iterrows():
                        html += '<tr style="text-align: center;">'
                        html += f'<td style="border: 1px solid #ddd; padding: 10px;"><b>{row["성취수준"]}</b></td>'
                        html += f'<td style="border: 1px solid #ddd; padding: 10px;">{int(row["학생수"])}</td>'
                        html += f'<td style="border: 1px solid #ddd; padding: 10px;">{row["비율(%)"]:.1f}</td>'
                        html += f'<td style="border: 1px solid #ddd; padding: 10px;">{row["정기시험평균"]:.2f}</td>'
                        html += f'<td style="border: 1px solid #ddd; padding: 10px;">{row["정기시험표준편차"]:.2f}</td>'
                        html += f'<td style="border: 1px solid #ddd; padding: 10px;">{row["원점수평균"]:.2f}</td>'
                        html += f'<td style="border: 1px solid #ddd; padding: 10px;">{row["원점수표준편차"]:.2f}</td>'
                        html += '</tr>'
                    html += '</tbody></table>'
                    return html
                
                stat_html = make_multi_header_table(stat_summary)
                st.markdown(stat_html, unsafe_allow_html=True)

            # --- [Tab 3] 문항 분석 ---
            with tab_item:
                # 평가 유형 및 criterion_rate 변수 확인 (사이드바에서 정의됨)
                # test_type, criterion_rate, target_mastery는 사이드바에서 이미 정의되어 있음
                
                # 크론바흐 알파 (신뢰도) 계산
                # 선택형 문항 1-16의 이진 행렬 생성 (1=정답, 0=오답)
                binary_matrix = pd.DataFrame()
                for i in range(1, 17):
                    col = f'Item_{i}'
                    # '.'을 1(정답)로, 그 외는 0(오답)으로 변환
                    binary_matrix[col] = (main_df[col].astype(str) == '.').astype(int)
                
                # KR-20 신뢰도 계산 (이분형 문항에 대한 크론바흐 알파)
                reliability = calculate_kr20_reliability(binary_matrix)
                
                # 신뢰도 표시
                st.markdown(f"""
                <div style="background-color: #f0f8ff; padding: 15px; border-radius: 10px; border-left: 5px solid #4682b4; margin-bottom: 20px;">
                    <h4 style="margin: 0; color: #2c3e50;">📊 신뢰도 (Cronbach's α)</h4>
                    <p style="font-size: 28px; font-weight: bold; margin: 10px 0; color: #2980b9;">{reliability:.3f}</p>
                    <p style="margin: 0; color: #7f8c8d; font-size: 14px;">KR-20 계수를 사용한 내적 일관성 신뢰도 (0.7 이상 권장)</p>
                </div>
                """, unsafe_allow_html=True)
                
                # 해석 안내 배너 (성취평가제일 때만 표시)
                if st.session_state.get('eval_type') == 'achievement':
                    st.info("""
                    ℹ️ **성취수준별 정답률 해석 안내**
                    
                    성취평가제에서 각 문항은 특정 성취수준을 판별하기 위해 출제되지만, 
                    개별 학생의 응답은 다양한 요인에 의해 영향을 받습니다.
                    
                    • C수준 학생도 A수준 문항을 맞힐 수 있습니다.
                    • 상위-하위 수준 간 정답률 차이가 작거나 정답률 패턴이 관찰되는 것은 **자연스러운 현상**일 수 있습니다.
                    • 특히 해당 수준의 학생 수가 적을 경우 통계적 변동이 큽니다.
                    
                    아래 **"참고 정보"**는 문항의 결함을 판단하는 것이 아니라, 출제 의도와 실제 응답 패턴을 비교하여 참고 정보를 제공합니다.
                    """)
                
                st.subheader("📋 선택형 문항 분석 (문항 1-16)")
                st.caption("각 문항의 정답률, 변별도, 선택지별 응답분포, 성취수준별 정답률을 분석합니다.")
                st.info("🎨 **선택형 문항의 색상 범례**: 2/3 이상은 흰색, 2/3 미만은 회색입니다.")
                
                # 사용 가능한 성취수준 결정
                available_levels = ['A', 'B', 'C', 'D', 'E']
                if level_type == "5수준+미도달 (A, B, C, D, E, 미도달)":
                    available_levels.append('미도달')
                
                # 문항 통계 일괄 계산 (성능 최적화)
                item_stats = calculate_all_item_statistics(main_df, info_df, available_levels)
                
                # 문항별 상세 분석 데이터 생성 - MultiIndex 구조 적용
                item_data_list = []
                for i in range(1, 17):
                    col = f'Item_{i}'
                    stats = item_stats[i]
                    
                    # 기본 정보 (캐시된 데이터 사용)
                    exp_diff = stats['exp_diff']
                    correct_ans = stats['correct_ans']
                    correct_rate = stats['correct_rate']
                    abandon_rate = discrimination_scores.get(i, 0)
                    
                    # 수집 데이터 - MultiIndex 구조를 위해 튜플 키 사용
                    # 그룹되지 않는 컬럼은 L1을 비우지 않고 '기본정보', '분석결과' 등으로 그룹화하여 UI 일관성 유지
                    # [수정] 헤더 단일행 처리를 위해 첫번째 레벨을 빈 문자열로 설정
                    row = {
                        ('', '문항'): i,
                        ('', '예상난이도'): exp_diff,
                        ('', '정답률(%)'): correct_rate,
                        ('', '변별도'): abandon_rate
                    }
                    
                    # 선택지별 응답분포 (캐시된 데이터 사용)
                    item_responses = main_df[col].astype(str).value_counts()
                    choice_counts = stats['choice_counts'].copy()
                    
                    # [수정] 정답 표기('.')를 해당 정답 번호의 카운트에 합산
                    if correct_ans in choice_counts:
                        # '.'는 이미 choice_counts에 포함되어 있지만, 정답 번호에 합산
                        choice_counts[correct_ans] = choice_counts.get(correct_ans, 0) + choice_counts.get('.', 0)
                        
                    no_response = item_responses.get('nan', 0) + item_responses.get('', 0)
                    
                    total_students = len(main_df)
                    
                    # 답지반응분포 (1~5, 무응답) - [수정] 비율(%)로 변경 + NaN 처리
                    for choice_key in sorted([k for k in choice_counts.keys() if k != '.']):
                        val_count = choice_counts[choice_key]
                        if val_count > 0 and total_students > 0:
                            row[('답지반응비율분포(%)', choice_key)] = round((val_count / total_students) * 100, 1)
                        else:
                            row[('답지반응비율분포(%)', choice_key)] = np.nan  # 응답이 없으면 빈칸
                    
                    if no_response > 0 and total_students > 0:
                        row[('답지반응비율분포(%)', '무응답')] = round((no_response / total_students) * 100, 1)
                    else:
                        row[('답지반응비율분포(%)', '무응답')] = np.nan  # 무응답이 없으면 빈칸
                    
                    # 성취수준별 정답률 (캐시된 데이터 사용)
                    achievement_rates = stats['achievement_rates']
                    for level in available_levels:
                        lv_rate = achievement_rates.get(level, 0)
                        if lv_rate > 0:
                            row[('성취수준별 정답률(%)', level)] = lv_rate
                        else:
                            row[('성취수준별 정답률(%)', level)] = np.nan  # 해당 성취수준 학생이 없으면 빈칸
                        
                    item_data_list.append(row)
                
                # DataFrame 생성
                analysis_df_multi = pd.DataFrame(item_data_list)
                
                # 컬럼 순서 정렬
                # 1. 문항, 예상난이도, 정답률, 변별도 - [수정] 1행처럼 보이게 하기 위해 첫번째 레벨을 빈 문자열로 설정
                # 컬럼 순서 정렬
                # 1. 문항, 예상난이도, 정답률, 변별도 - [수정] 1행, 2행 제목을 동일하게 설정하여 to_html() 시 자동 병합 유도
                cols_basic = [('문항', '문항'), ('예상난이도', '예상난이도'), ('정답률', '정답률(%)'), ('변별도', '변별도')]
                # 2. 답지반응 (1~5, 무응답)
                cols_response = [('답지반응비율분포(%)', str(k)) for k in range(1, 6)] + [('답지반응비율분포(%)', '무응답')]
                # 3. 성취수준별
                cols_level = [('성취수준별 정답률(%)', lv) for lv in available_levels]
                
                final_cols = cols_basic + cols_response + cols_level
                # 존재하는 컬럼만 필터링 (컬럼 매칭 로직 수정 필요)
                # 데이터 생성 시 키 값도 변경해야 함
                
                # 데이터 리스트 재구성 (키 불일치 방지)
                new_item_data_list = []
                for idx, row_data in analysis_df.iterrows():
                    # 기존 로직에서 계산된 값들 가져오기 (이미 위에서 계산됨)
                    # 여기서는 analysis_df의 값을 기반으로 새로운 키의 딕셔너리 생성
                    
                    # analysis_df는 기존 컬럼명을 가짐
                    # item_data_list 생성 로직을 여기서 다시 구현하는 대신, 
                    # 위쪽의 데이터 생성 루프를 수정하는 것이 안전함.
                    pass 

                # [주의] 위의 item_data_list 생성 루프는 이 코드 블록 이전에 위치함.
                # 따라서 이 replace_file_content로는 루프 내부를 수정할 수 없으므로,
                # analysis_df_multi를 생성한 후 컬럼명을 변경하는 방식을 사용하거나,
                # 범위를 넓혀서 루프까지 수정해야 함.
                # 현재 범위(1660-1720) 내에 루프 생성 코드는 없음.
                # 따라서 analysis_df_multi 컬럼을 매핑으로 변경.

                # 기존 컬럼 매핑 ('', '문항') -> ('문항', '문항')
                # 하지만 현재 analysis_df_multi는 아직 MultiIndex가 아님 (리스트 딕셔너리에서 생성됨)
                
                # 전략: 데이터프레임 생성 후 MultiIndex 설정 시 names를 조정하는 것이 아니라,
                # 컬럼 파이프라인을 재정의.
                
                # cols_basic 정의만 바꾸고, 실제 데이터의 키는 아래 로직에서 처리되도록 함.
                # DataFrame 컬럼 재설정 (Renaming approach)
                rename_map = {
                    ('', '문항'): ('문항', '문항'),
                    ('문항', '문항'): ('문항', '문항'), # 재실행시 안전장치
                    ('', '예상난이도'): ('예상난이도', '예상난이도'),
                    ('예상난이도', '예상난이도'): ('예상난이도', '예상난이도'),
                    ('', '정답률(%)'): ('정답률', '정답률'), # [변경] 병합을 위해 이름 통일
                    ('정답률', '정답률(%)'): ('정답률', '정답률'),
                    ('', '변별도'): ('변별도', '변별도'),
                    ('변별도', '변별도'): ('변별도', '변별도')
                }
                
                new_cols = []
                for c in analysis_df_multi.columns:
                    if c in rename_map:
                        new_cols.append(rename_map[c])
                    else:
                        new_cols.append(c)
                analysis_df_multi.columns = new_cols

                # MultiIndex 변환
                analysis_df_multi.columns = pd.MultiIndex.from_tuples(
                    analysis_df_multi.columns, 
                    names=[None, None]
                )

                # 숫자형 변환 (키 변경 반영)
                analysis_df_multi[('정답률', '정답률')] = pd.to_numeric(analysis_df_multi[('정답률', '정답률')], errors='coerce').round(1)
                analysis_df_multi[('변별도', '변별도')] = pd.to_numeric(analysis_df_multi[('변별도', '변별도')], errors='coerce').round(1)
                
                response_cols = [c for c in analysis_df_multi.columns if c[0] == '답지반응비율분포(%)']
                level_cols = [c for c in analysis_df_multi.columns if c[0] == '성취수준별 정답률(%)']
                
                # 스타일링
                styler = analysis_df_multi.style.format(precision=1, subset=level_cols, na_rep='') \
                    .hide(axis='index') \
                    .format(precision=1, subset=[('정답률', '정답률')]) \
                    .format(precision=1, subset=[('변별도', '변별도')]) \
                    .format(precision=1, subset=response_cols, na_rep='') \
                    .map(lambda x: style_background_level_v2(x, 66.7), subset=level_cols) \
                    .map(lambda x: custom_bar_style(x, 66.7), subset=[('정답률', '정답률')]) \
                    .bar(subset=[('변별도', '변별도')], color='#a5d6a7', vmin=-0.2, vmax=1.0) \
                    .set_table_styles([
                        {'selector': 'th', 'props': [
                            ('text-align', 'center'), 
                            ('font-weight', 'bold'), 
                            ('color', 'black'), 
                            ('vertical-align', 'middle'), 
                            ('border', '1px solid #e0e0e0'),
                            ('background-color', '#f8f9fa')
                        ]},
                        {'selector': 'td', 'props': [
                            ('text-align', 'center'), 
                            ('vertical-align', 'middle'), 
                            ('border', '1px solid #e0e0e0')
                        ]}
                    ], overwrite=False)

                # HTML 생성
                html = styler.to_html(escape=False)
                html = merge_headers(html, ['문항', '예상난이도', '정답률', '변별도'])
                
                # [DataTables 렌더링 호출]
                render_datatables(html, unique_id='item_analysis')
                
                st.divider()
                
                # 난이도별 범주화 분석
                st.subheader("📊 난이도별 범주화 및 성취수준별 정답률 분석")
                st.caption("난이도(상/중/하) 범주별로 해당 문항들의 성취수준별 평균 정답률을 표시합니다.")
                
                # 난이도별 분류
                difficulty_categories = {'상': [], '중': [], '하': []}
                for i in range(1, 17):
                    exp_diff_val = info_df[info_df['No'] == i]['Exp_Diff'].values
                    exp_diff = exp_diff_val[0] if len(exp_diff_val) > 0 else '중'
                    difficulty_categories[exp_diff].append(i)
                
                # 난이도별 성취수준별 정답률 계산
                difficulty_analysis = []
                for difficulty in ['상', '중', '하']:
                    items_in_category = difficulty_categories[difficulty]
                    if len(items_in_category) == 0:
                        continue
                    
                    row = {'난이도': difficulty, '문항 수': len(items_in_category)}
                    
                    # 각 성취수준별 정답률 계산
                    for level in available_levels:
                        level_data = main_df[main_df['Achievement'] == level]
                        if len(level_data) > 0:
                            # 해당 난이도 문항들에 대한 정답률
                            correct_counts = []
                            for item_no in items_in_category:
                                col = f'Item_{item_no}'
                                correct_count = (level_data[col].astype(str) == '.').sum()
                                correct_counts.append(correct_count)
                            
                            avg_correct_rate = (sum(correct_counts) / (len(level_data) * len(items_in_category)) * 100) if len(items_in_category) > 0 else 0
                            row[f'{level} 정답률(%)'] = round(avg_correct_rate, 1)
                        else:
                            row[f'{level} 정답률(%)'] = np.nan
                    
                    difficulty_analysis.append(row)
                
                # DataFrame 생성
                diff_analysis_df = pd.DataFrame(difficulty_analysis)
                
                # 숫자 컬럼을 명시적으로 float로 변환하고, NaN을 빈 문자열로 대체
                for col in diff_analysis_df.columns:
                    if '정답률' in col:
                        diff_analysis_df[col] = pd.to_numeric(diff_analysis_df[col], errors='coerce')
                
                # NaN을 빈 문자열로 변환 (표시용)
                diff_display_df = diff_analysis_df.fillna('')
                
                # 스타일링
                def style_difficulty_table(val):
                    if val == '' or pd.isna(val):
                        return 'text-align: center; font-size: 16px; padding: 10px;'
                    
                    # 숫자 값 확인
                    try:
                        num_val = float(val) if isinstance(val, str) else val
                        # 66.7% 미만이면 회색, 이상이면 흰색
                        bg_color = '#eeeeee' if num_val < 66.7 else '#ffffff'
                        return f'background-color: {bg_color}; text-align: center; font-size: 16px; padding: 10px;'
                    except:
                        return 'text-align: center; font-size: 16px; padding: 10px;'
                
                # 숫자 포맷팅 함수 (빈 값은 그대로 유지)
                def format_number(val):
                    if val == '' or pd.isna(val):
                        return ''
                    if isinstance(val, (int, float)):
                        return f'{val:.1f}'
                    return val
                
                styled_diff = diff_display_df.style \
                    .format(format_number) \
                    .applymap(style_difficulty_table) \
                    .set_table_styles([
                        {'selector': 'th', 'props': [
                            ('text-align', 'center'),
                            ('font-weight', 'bold'),
                            ('color', 'black'),
                            ('vertical-align', 'middle'),
                            ('border', '1px solid #e0e0e0'),
                            ('background-color', '#f8f9fa'),
                            ('font-size', '16px'),
                            ('padding', '12px')
                        ]},
                        {'selector': 'td', 'props': [
                            ('text-align', 'center'),
                            ('vertical-align', 'middle'),
                            ('border', '1px solid #e0e0e0'),
                            ('font-size', '16px'),
                            ('padding', '10px')
                        ]}
                    ], overwrite=False).hide(axis='index')
                
                st.dataframe(styled_diff, use_container_width=True)
                
                st.divider()
                
                # P-D Chart (보조 분석)
                # 문항 수준 판정 함수 (성취평가제용) - 로컬 버전 (tuple 반환)
                def determine_item_level_local(achievement_rates, criterion_rate=66.7):
                    """
                    성취수준별 정답률을 분석하여 문항의 목표 수준을 판정
                    
                    판정 규칙:
                    - A수준 문항: A만 기준(66.7%) 충족, B~E 및 미도달은 미충족
                    - B수준 문항: A, B가 기준 충족, C~E 및 미도달은 미충족
                    - C수준 문항: A, B, C가 기준 충족, D, E, 미도달은 미충족
                    - D수준 문항: A, B, C, D가 기준 충족, E, 미도달은 미충족
                    - E수준 문항: A, B, C, D, E 모두 기준 충족 (미도달은 무관)
                    """
                    levels = ['A', 'B', 'C', 'D', 'E']
                    if '미도달' in achievement_rates:
                        levels.append('미도달')
                    
                    # 각 수준별 기준 충족 여부 확인
                    meets_criterion = {}
                    for level in levels:
                        rate = achievement_rates.get(level, 0)
                        meets_criterion[level] = rate >= criterion_rate
                    
                    # 패턴 매칭으로 문항 수준 판정
                    if meets_criterion.get('A', False) and not meets_criterion.get('B', True):
                        return 'A', meets_criterion, True
                    elif meets_criterion.get('A', False) and meets_criterion.get('B', False) and not meets_criterion.get('C', True):
                        return 'B', meets_criterion, True
                    elif meets_criterion.get('A', False) and meets_criterion.get('B', False) and meets_criterion.get('C', False) and not meets_criterion.get('D', True):
                        return 'C', meets_criterion, True
                    elif meets_criterion.get('A', False) and meets_criterion.get('B', False) and meets_criterion.get('C', False) and meets_criterion.get('D', False) and not meets_criterion.get('E', True):
                        return 'D', meets_criterion, True
                    elif all([meets_criterion.get(l, False) for l in ['A', 'B', 'C', 'D', 'E']]):
                        return 'E', meets_criterion, True
                    else:
                        return '판정불가', meets_criterion, False
                
                # 정답률 패턴 탐지 함수
                def detect_reversals(achievement_rates):
                    """
                    성취수준이 높은 집단의 정답률이 낮은 집단보다 낮은 경우를 탐지
                    (정답률 패턴 관찰)
                    """
                    levels = ['A', 'B', 'C', 'D', 'E']
                    if '미도달' in achievement_rates:
                        levels.append('미도달')
                    
                    reversals = []
                    for i in range(len(levels) - 1):
                        upper_level = levels[i]
                        lower_level = levels[i + 1]
                        
                        upper_rate = achievement_rates.get(upper_level, 0)
                        lower_rate = achievement_rates.get(lower_level, 0)
                        
                        if upper_rate < lower_rate:
                            gap = lower_rate - upper_rate
                            severity = 'high' if gap > 5 else 'medium' if gap > 2 else 'low'
                            reversals.append({
                                'upper_level': upper_level,
                                'lower_level': lower_level,
                                'gap': gap,
                                'severity': severity
                            })
                    
                    return reversals
                
                # 양호도 판정 함수 정의 (평가 유형별 분기)
                def judge_item_quality(p, d, test_type, criterion_rate=66.7, achievement_rates=None):
                    """난이도(P)와 변별돀(D)에 따른 양호도 판정 (평가 목적별 분기)"""
                    if pd.isna(p) or pd.isna(d):
                        return '미평가', '#CCCCCC', 'unknown', ''
                    
                    p_val = float(p)
                    d_val = float(d)
                    p_percent = p_val * 100  # 퍼센트로 변환
                    
                    # === 석차 5등급제 (상대평가): 변별도가 핵심 ===
                    if test_type == "석차 5등급제 (상대평가)":
                        # 1. 변별도가 매우 높은 문항
                        if d_val >= 0.4:
                            if 0.4 <= p_val <= 0.7:
                                return '✅ 높은 변별도 (0.4+)', '#90EE90', 'excellent', '높은 변별력과 적정 난이도'
                            else:
                                return '✅ 높은 변별도 (0.4+)', '#B8E6B8', 'good_disc', '높은 변별력'
                        
                        # 2. 극단적으로 어려운 문항 (상위권도 다 틀림)
                        elif p_val < 0.2:
                            if d_val < 0.0:  # 역방향 변별
                                return '📋 초고난도/역변별', '#FF6B6B', 'error', '상위권도 오답 선택 - 문항 내용 확인 권장'
                            else:
                                return '📚 초고난도', '#DDA0DD', 'very_hard', '출제 의도와 실제 데이터 비교 권장'
                        
                        # 3. 변별도가 낮은 문항
                        elif d_val < 0.2:
                            return '📊 낮은 변별도 (0.2 미만)', '#FFB6C6', 'poor_disc', '선발 기능 저하가 관찰됨'
                        
                        # 4. 난이도 관찰
                        elif p_val > 0.8:
                            return '📝 매우 쉬움', '#ADD8E6', 'easy', '정답률 80% 초과'
                        elif p_val < 0.3:
                            return '📚 어려움', '#DDA0DD', 'hard', '정답률 30% 미만'
                        
                        # 5. 보통
                        else:
                            return '📋 보통', '#FFF9E6', 'fair', '중간 수준의 지표'
                    
                    # === 성취평가제 (절대평가): 성취수준별 패턴 분석 ===
                    else:
                        # achievement_rates가 있으면 패턴 기반 분석
                        if achievement_rates:
                            # 문항 수준 판정
                            item_level_result = determine_item_level(achievement_rates, criterion_rate)
                            item_level = item_level_result['level']
                            is_clear = len(item_level_result['below']) == 0
                            
                            # 정답률 패턴 관찰
                            pattern_result = analyze_achievement_pattern(achievement_rates)
                            reversals = pattern_result.get('observations', [])
                            
                            # 판정
                            if is_clear and len(reversals) == 0:
                                # 최상의 경우: 명확한 수준 판정 + 정상적인 패턴
                                if d_val >= 0.3:
                                    return f'✅ {item_level}수준+변별', '#90EE90', 'excellent', f'{item_level}수준 문항으로 기준 충족. 변별도도 높음.'
                                else:
                                    return f'✅ {item_level}수준', '#B8E6B8', 'good', f'{item_level}수준 문항으로 기준 충족. 변별도는 낮지만 성취평가제에서는 문제없음.'
                            elif is_clear and len(reversals) > 0:
                                # 수준은 판정되나 정답률 패턴 관찰됨
                                return f'📋 {item_level}수준(패턴)', '#D4E6F1', 'pattern', f'{item_level}수준 문항이나 {len(reversals)}건의 정답률 패턴이 관찰됨. 참고하세요.'
                            else:
                                # 수준 판정 불가
                                return '📊 판정불가', '#FFB6C6', 'unclear', '성취수준별 패턴이 불명확합니다. 문항 내용 확인 권장.'
                        else:
                            # achievement_rates가 없으면 전체 정답률 기반 (호환성)
                            if p_percent >= 70:
                                return '✅ 기초달성', '#B8E6B8', 'basic', f'전체 정답률 {p_percent:.0f}% - 기초 학습 달성'
                            elif p_val < 0.3:
                                return '⛔ 학습미달', '#FF6B6B', 'not_mastered', '전체 정답률이 매우 낮음. 교수학습 방안 검토 필요'
                            else:
                                return '📋 보통', '#FFF9E6', 'fair', '전체 정답률 중간 수준'
                
                st.subheader("📊 문항 양호도 맵 및 판정")
                
                # 그래프 위 안내 설명
                with st.expander("📌 문항 양호도 판정 기준 보기", expanded=False):
                    col1, col2 = st.columns(2)
                    with col1:
                        st.write("""
                        **정답률(P) - X축**
                        - 응시한 학생들 중 정답한 학생의 비율
                        - 좌측(낮음): 어려운 문항
                        - 우측(높음): 쉬운 문항
                        
                        **변별도(D) - Y축**
                        - 상위 27% 학생과 하위 27% 학생의 정답률 차이
                        - 하단(낮음): 변별력이 낮은 문항
                        - 상단(높음): 변별력이 높은 문항
                        """)
                    with col2:
                        if test_type == "석차 5등급제 (상대평가)":
                            st.write("""
                            **판정 기준 - 석차 5등급제 (변별도 중심)**
                            - ✅ **높은 변별도 (0.4+)**: 변별도 ≥ 0.4, 정답률 40-70%
                            - ✅ **높은 변별도 (0.4+)**: 변별도 ≥ 0.4
                            - 📋 **초고난도/역변별**: 정답률 < 20%, 역변별
                            - 📊 **낮은 변별도 (0.2 미만)**: 변별도 < 0.2
                            - 📝 **매우 쉬움**: 정답률 > 80%
                            - 📚 **어려움**: 정답률 < 30%
                            """)
                        else:
                            st.write(f"""
                            **판정 기준 - 성취평가제 (성취수준별 패턴)**
                            - ✅ **기준 충족**: 문항 수준 명확하고 정상적인 패턴
                            - 📋 **참고**: 문항 수준은 판정되나 정답률 패턴 관찰됨
                            - 📊 **판정불가**: 성취수준별 패턴 불명확
                            
                            **기준 정답률**: {criterion_rate}%
                            - 해당 수준 학생의 {criterion_rate}% 이상이 정답
                            - 한국교육과정평가원(KICE) 2/3 기준
                            
                            ℹ️ 변별도보다 성취수준별 패턴이 핵심
                            """)
                
                st.markdown("##### 📈 P-D Chart (난이도-변별력 분포)")
                
                # res_df 컬럼명 한글화 (표시용)
                res_display = res_df.rename(columns={
                    'No': '문항', 'Exp_Diff': '예상난이도', 'Score': '배점', 'Standard': '성취기준'
                })
                
                fig_pd = px.scatter(res_display, x='정답률(P)', y='변별도(D)', text='문항', color='예상난이도',
                                size='배점', title="<b>문항 난이도 및 변별력 분석</b>",
                                labels={'정답률(P)': '정답률(난이도) - 어려움 ⟵ ⟶ 쉬움', '변별도(D)': '변별도(변별력) - 낮음 ⟵ ⟶ 높음'},
                                color_discrete_map={'상': '#FF9F43', '중': '#54A0FF', '하': '#1DD1A1'})
                
                # 평가 유형에 따른 참조선과 영역 추가
                if test_type == "석차 5등급제 (상대평가)":
                    # 높은 변별도 기준선
                    fig_pd.add_hline(y=0.4, line_dash="dash", line_color="red", 
                                    annotation_text="높은 변별도 기준 (0.4)", annotation_position="right")
                    fig_pd.add_hline(y=0.2, line_dash="dot", line_color="orange", 
                                    annotation_text="최소 변별 (0.2)", annotation_position="right")
                    # 이상적 난이도 영역
                    fig_pd.add_vrect(x0=0.4, x1=0.7, fillcolor="lightgreen", opacity=0.1, 
                                    annotation_text="이상적 난이도", annotation_position="top left")
                else:
                    # 성취수준별 기준선 (성취평가제)
                    criterion_line = criterion_rate / 100
                    fig_pd.add_hline(y=criterion_line, line_dash="dash", line_color="orange",
                                    annotation_text=f"성취수준별 기준({criterion_rate}%)", annotation_position="right")
                    st.info(f"📌 성취평가제: 각 성취수준 학생의 {criterion_rate}% 이상이 정답을 맞춰야 기준 충족")
                
                fig_pd.update_layout(
                    paper_bgcolor="rgba(0,0,0,0)", 
                    plot_bgcolor="rgba(240,242,246,0.5)", 
                    font_family="Pretendard",
                    height=400,
                    hovermode='closest'
                )
                st.plotly_chart(fig_pd, use_container_width=True)
                
                # 그래프 아래 데이터 해석 결과
                st.write("**📈 데이터 해석 결과**")
                
                # 평가 유형에 따른 기준 정답률 가져오기
                criterion_rate = st.session_state.get('criterion_rate', 66.7)
                
                # 양호도 데이터 생성 (평가 유형별 판정 기준 사용)
                quality_data = []
                item_level_data = []  # 성취평가제용 문항 수준 판정 결과
                
                for idx, row in res_df.iterrows():
                    item_no = int(row['No'])
                    p = row['정답률(P)']
                    d = row['변별도(D)']
                    
                    # 성취확인형일 때 성취수준별 정답률 계산
                    achievement_rates = {}
                    if test_type == "성취확인형 (절대평가)":
                        col = f'Item_{item_no}'
                        for level in available_levels:
                            level_data = main_df[main_df['Achievement'] == level]
                            if len(level_data) > 0 and col in level_data.columns:
                                level_rate = (level_data[col].astype(str) == '.').mean() * 100
                                achievement_rates[level] = level_rate
                            else:
                                achievement_rates[level] = 0
                        
                        # 문항 수준 판정
                        item_level_result = determine_item_level(achievement_rates, criterion_rate)
                        item_level = item_level_result['level']
                        is_clear = len(item_level_result['below']) == 0
                        
                        # 정답률 패턴 탐지 (analyze_achievement_pattern 함수 사용)
                        pattern_result = analyze_achievement_pattern(achievement_rates)
                        reversals = pattern_result.get('observations', [])
                        
                        item_level_data.append({
                            '문항': item_no,
                            '판정수준': item_level,
                            '명확성': '✓' if is_clear else '△',
                            '패턴관찰': len(reversals),
                            '성취수준별정답률': achievement_rates,
                            '패턴상세': reversals
                        })
                    
                    status, color, category, reason = judge_item_quality(p, d, test_type, criterion_rate, achievement_rates if achievement_rates else None)
                    
                    quality_data.append({
                        '문항': item_no,
                        '정답률(%)': round(p * 100, 1),
                        '변별도': round(d, 3),
                        '양호도': status,
                        '분류': category,
                        '해석': reason
                    })
                
                quality_df = pd.DataFrame(quality_data)
                
                # 평가 유형별 메트릭 및 해석
                if test_type == "석차 5등급제 (상대평가)":
                    # 석차 5등급제 카테고리
                    excellent_items = quality_df[quality_df['분류'].isin(['excellent', 'good_disc'])]
                    easy_items = quality_df[quality_df['분류'] == 'easy']
                    hard_items = quality_df[quality_df['분류'].isin(['hard', 'very_hard'])]
                    poor_disc_items = quality_df[quality_df['분류'] == 'poor_disc']
                    error_items = quality_df[quality_df['분류'] == 'error']
                    
                    # 메트릭
                    col1, col2, col3, col4, col5 = st.columns(5)
                    
                    with col1:
                        st.metric("✅ 높은 변별도", len(excellent_items), 
                                  help="변별도 ≥ 0.4인 문항")
                    
                    with col2:
                        st.metric("📊 낮은 변별도", len(poor_disc_items),
                                  help="변별도 < 0.2")
                    
                    with col3:
                        st.metric("📝 매우 쉬움", len(easy_items),
                                  help="정답률 > 80%")
                    
                    with col4:
                        st.metric("📚 어려움", len(hard_items),
                                  help="정답률 < 30%")
                    
                    with col5:
                        st.metric("📋 역변별 관찰", len(error_items),
                                  help="역변별 현상 관찰됨")
                    
                    # 상세 해석
                    if len(excellent_items) > 0:
                        st.success(f"✅ **높은 변별도 문항**: {', '.join(map(str, excellent_items['문항'].unique()))} " + 
                                  f"(총 {len(excellent_items)}개) - {excellent_items.iloc[0]['해석']}")
                    
                    if len(poor_disc_items) > 0:
                        st.warning(f"📊 **낮은 변별도**: {', '.join(map(str, poor_disc_items['문항'].unique()))} " +
                                  f"(총 {len(poor_disc_items)}개) - 선발 기능 저하가 관찰됩니다.")
                    
                    if len(easy_items) > 0:
                        st.info(f"📝 **매우 쉬움**: {', '.join(map(str, easy_items['문항'].unique()))} " +
                                 f"(총 {len(easy_items)}개) - 정답률 80% 초과.")
                    
                    if len(hard_items) > 0:
                        st.info(f"📚 **어려운 문항**: {', '.join(map(str, hard_items['문항'].unique()))} " +
                              f"(총 {len(hard_items)}개) - 정답률이 낮습니다.")
                    
                    if len(error_items) > 0:
                        st.warning(f"📋 **역변별 문항**: {', '.join(map(str, error_items['문항'].unique()))} " +
                                f"(총 {len(error_items)}개) - 상위권 학생도 오답 선택. 문항 내용 확인 권장.")
                
                else:  # 성취평가제
                    # 성취평가제 카테고리
                    excellent_items = quality_df[quality_df['분류'].isin(['excellent', 'good'])]
                    reversal_items = quality_df[quality_df['분류'] == 'pattern']
                    unclear_items = quality_df[quality_df['분류'] == 'unclear']
                    basic_items = quality_df[quality_df['분류'] == 'basic']
                    
                    # 메트릭
                    col1, col2, col3, col4 = st.columns(4)
                    
                    with col1:
                        st.metric("✅ 기준 충족", len(excellent_items), 
                                  help="문항 수준 명확하고 정상적인 패턴")
                    
                    with col2:
                        st.metric("📋 패턴관찰", len(reversal_items),
                                  help="문항 수준은 판정되나 정답률 패턴 관찰됨")
                    
                    with col3:
                        st.metric("📊 판정불가", len(unclear_items),
                                  help="성취수준별 패턴 불명확")
                    
                    with col4:
                        st.metric("ℹ️ 기타", len(basic_items),
                                  help="기타 문항")
                    
                    # 상세 해석
                    if len(excellent_items) > 0:
                        st.success(f"✅ **기준 충족 문항**: {', '.join(map(str, excellent_items['문항'].unique()))} " + 
                                  f"(총 {len(excellent_items)}개) - 명확한 문항 수준 판정, 정상적인 정답률 패턴")
                    
                    if len(reversal_items) > 0:
                        st.info(f"📋 **패턴관찰 문항**: {', '.join(map(str, reversal_items['문항'].unique()))} " +
                                   f"(총 {len(reversal_items)}개) - 성취수준 간 정답률 패턴 관찰. 아래 상세 분석 참조.")
                    
                    if len(unclear_items) > 0:
                        st.info(f"📊 **판정 불가 문항**: {', '.join(map(str, unclear_items['문항'].unique()))} " +
                                f"(총 {len(unclear_items)}개) - 성취수준별 패턴이 불명확. 문항 내용 확인 권장.")
                
                # 양호도 판정 표 표시
                st.markdown("##### 📋 문항별 양호도 판정표")
                
                # 문항 선택 기능 추가
                st.caption("📌 특정 문항을 선택하면 상세 분석을 볼 수 있습니다.")
                
                def style_quality(val):
                    status_colors = {
                        # 선발형
                        '✅ 높은 변별도 (0.4+)': '#90EE90',
                        '📋 초고난도/역변별': '#FF6B6B',
                        '📊 낮은 변별도 (0.2 미만)': '#FFB6C6',
                        '📝 매우 쉬움': '#ADD8E6',
                        '📚 초고난도': '#DDA0DD',
                        '📚 어려움': '#DDA0DD',
                        '📋 보통': '#FFF9E6',
                        # 성취평가제 (패턴 기반)
                        '✅ A수준+변별': '#90EE90',
                        '✅ B수준+변별': '#90EE90',
                        '✅ C수준+변별': '#90EE90',
                        '✅ D수준+변별': '#90EE90',
                        '✅ E수준+변별': '#90EE90',
                        '✅ A수준': '#B8E6B8',
                        '✅ B수준': '#B8E6B8',
                        '✅ C수준': '#B8E6B8',
                        '✅ D수준': '#B8E6B8',
                        '✅ E수준': '#B8E6B8',
                        '📋 A수준(패턴)': '#D4E6F1',
                        '📋 B수준(패턴)': '#D4E6F1',
                        '📋 C수준(패턴)': '#D4E6F1',
                        '📋 D수준(패턴)': '#D4E6F1',
                        '📋 E수준(패턴)': '#D4E6F1',
                        '📋 판정불가(패턴)': '#D4E6F1',
                        '📊 판정불가': '#FFB6C6',
                        '✅ 기초달성': '#B8E6B8',
                        '📋 학습미달': '#FF6B6B',
                        # 기타
                        '미평가': '#CCCCCC'
                    }
                    color = status_colors.get(val, '#FFFFFF')
                    return f'background-color: {color}; text-align: center; font-weight: bold;'
                
                # 분류 컬럼은 표시하지 않음 (해석 컬럼은 표시)
                display_quality_df = quality_df[['문항', '정답률(%)', '변별도', '양호도', '해석']]
                
                styled_quality = display_quality_df.style \
                    .map(style_quality, subset=['양호도']) \
                    .format(precision=1) \
                    .set_table_styles([
                        {'selector': 'th', 'props': [('text-align', 'center'), ('font-weight', 'bold'), ('background-color', '#f8f9fa')]},
                        {'selector': 'td', 'props': [('text-align', 'center'), ('vertical-align', 'middle')]}
                    ], overwrite=False) \
                    .hide(axis='index')
                
                st.dataframe(styled_quality, use_container_width=True)
                
                # 성취평가제: 문항 수준 판정 결과 표시
                if test_type == "성취평가제 (절대평가)" and len(item_level_data) > 0:
                    st.markdown("##### 📊 문항 수준 판정 (성취평가제)")
                    st.caption(f"각 문항이 어느 성취수준을 목표로 하는지 판정합니다. (기준: {criterion_rate}% 이상 정답)")
                    
                    level_df = pd.DataFrame(item_level_data)
                    
                    # 판정 결과 요약
                    level_counts = level_df['판정수준'].value_counts()
                    col1, col2, col3, col4, col5, col6 = st.columns(6)
                    
                    with col1:
                        st.metric("A수준", level_counts.get('A', 0))
                    with col2:
                        st.metric("B수준", level_counts.get('B', 0))
                    with col3:
                        st.metric("C수준", level_counts.get('C', 0))
                    with col4:
                        st.metric("D수준", level_counts.get('D', 0))
                    with col5:
                        st.metric("E수준", level_counts.get('E', 0))
                    with col6:
                        st.metric("판정불가", level_counts.get('판정불가', 0))
                    
                    # 정답률 패턴 관찰
                    pattern_items = level_df[level_df['패턴관찰'] > 0]
                    if len(pattern_items) > 0:
                        st.info(f"📋 **정답률 패턴 관찰**: {len(pattern_items)}개 문항에서 정답률 패턴이 관찰되었습니다.")
                        
                        for _, item in pattern_items.iterrows():
                            with st.expander(f"문항 {item['문항']} - 패턴 상세", expanded=False):
                                for obs in item['패턴상세']:
                                    st.info(
                                        f"{obs['icon']} **{obs['title']}**: {obs['upper']}수준({obs['upper_rate']:.1f}%) < {obs['lower']}수준({obs['lower_rate']:.1f}%) "
                                        f"(차이: {obs['abs_gap']:.1f}%p)\n\n{obs['message']}\n\n{obs['disclaimer']}"
                                    )
                                
                                # 성취수준별 정답률 표시
                                st.write("**성취수준별 정답률:**")
                                ach_rates = item['성취수준별정답률']
                                rate_cols = st.columns(len(ach_rates))
                                for idx, (level, rate) in enumerate(ach_rates.items()):
                                    with rate_cols[idx]:
                                        st.metric(level, f"{rate:.1f}%")
                    else:
                        st.success("✅ 모든 문항에서 단조성이 유지됩니다. (정상적인 패턴)")
                    
                    # 문항 수준 판정 표
                    display_level_df = level_df[['문항', '판정수준', '명확성', '패턴관찰']].rename(columns={'패턴관찰': '패턴'})
                    st.dataframe(display_level_df, use_container_width=True, hide_index=True)
                
                # 문항 상세 분석 (드릴다운)
                st.markdown("##### 🔍 문항별 상세 분석 (Drill-down)")
                selected_item = st.selectbox(
                    "분석할 문항 선택",
                    options=quality_df['문항'].tolist(),
                    format_func=lambda x: f"문항 {x} - {quality_df[quality_df['문항']==x]['양호도'].values[0]}"
                )
                
                if selected_item:
                    with st.expander(f"📊 문항 {selected_item} 상세 분석", expanded=True):
                        sel_item_data = quality_df[quality_df['문항'] == selected_item].iloc[0]
                        
                        # 기본 정보
                        col_a, col_b, col_c = st.columns(3)
                        with col_a:
                            st.metric("정답률", f"{sel_item_data['정답률(%)']}%")
                        with col_b:
                            st.metric("변별도", f"{sel_item_data['변별도']:.3f}")
                        with col_c:
                            st.markdown(f"**양호도**: {sel_item_data['양호도']}")
                        
                        st.info(f"💡 **해석**: {sel_item_data['해석']}")
                        
                        # 상위/하위 집단별 답지 반응 분석
                        st.markdown("**📈 상위/하위 집단별 답지 선택 패턴**")
                        st.caption("상위 27%와 하위 27% 학생들의 선택 패턴을 비교합니다.")
                        
                        col_name = f'Item_{selected_item}'
                        if col_name in main_df.columns:
                            # Total_Score 기준 상위/하위 27% 학생 구분
                            total_students = len(main_df)
                            top_n = int(total_students * 0.27)
                            
                            # Total_Score 컬럼이 있는지 확인
                            if 'Total_Score' in main_df.columns:
                                sorted_df = main_df.sort_values('Total_Score', ascending=False)
                            else:
                                # Total_Score가 없으면 선택형 문항 합산으로 대체
                                item_cols = [f'Item_{i}' for i in range(1, 17) if f'Item_{i}' in main_df.columns]
                                main_df['_temp_score'] = main_df[item_cols].apply(
                                    lambda row: sum([1 if str(val) == '.' else 0 for val in row]), axis=1
                                )
                                sorted_df = main_df.sort_values('_temp_score', ascending=False)
                            
                            top_group = sorted_df.head(top_n)
                            bottom_group = sorted_df.tail(top_n)
                            
                            # 답지별 선택 비율
                            choices = ['1', '2', '3', '4', '5', '.']  # .은 정답
                            top_dist = top_group[col_name].astype(str).value_counts(normalize=True) * 100
                            bottom_dist = bottom_group[col_name].astype(str).value_counts(normalize=True) * 100
                            
                            # 정답 확인
                            try:
                                correct_ans = str(int(info_df[info_df['No'] == selected_item]['Correct_Ans'].values[0]))
                            except:
                                correct_ans = ''
                            
                            # 데이터프레임 생성
                            distractor_data = []
                            for choice in choices:
                                if choice == '.':
                                    label = f"정답 ({correct_ans}번)" if correct_ans else "정답"
                                else:
                                    label = f"{choice}번"
                                
                                top_pct = top_dist.get(choice, 0)
                                bottom_pct = bottom_dist.get(choice, 0)
                                
                                distractor_data.append({
                                    '선택지': label,
                                    '상위 27%': f"{top_pct:.1f}%",
                                    '하위 27%': f"{bottom_pct:.1f}%",
                                    '차이': f"{top_pct - bottom_pct:+.1f}%p"
                                })
                            
                            distractor_df = pd.DataFrame(distractor_data)
                            st.dataframe(distractor_df, use_container_width=True, hide_index=True)
                            
                            # 패턴 진단
                            st.markdown("**🔬 패턴 진단**")
                            
                            # 정답 선택률
                            correct_top = top_dist.get('.', 0)
                            correct_bottom = bottom_dist.get('.', 0)
                            
                            if correct_top > 80 and correct_bottom < 40:
                                st.success("✅ **양호한 패턴**: 상위권은 정답에 집중, 하위권은 분산. 변별력이 우수합니다.")
                            elif correct_top < correct_bottom:
                                st.error("⚠️ **역변별 패턴**: 하위권이 상위권보다 정답률이 높습니다. 문항에 오류가 있거나 출제 의도와 다른 해석이 가능할 수 있습니다.")
                            elif correct_top < 50 and correct_bottom < 50:
                                # 오답 매력도 체크
                                max_distractor = max([top_dist.get(c, 0) for c in ['1', '2', '3', '4', '5']])
                                if max_distractor > correct_top:
                                    st.warning(f"⚠️ **매력적 오답 존재**: 특정 오답이 정답보다 많이 선택되었습니다. 오개념을 유발하는 요인이 있을 수 있습니다.")
                                else:
                                    st.info("ℹ️ **난이도 과다**: 상위/하위 모두 정답률이 낮습니다. 문항이 너무 어렵거나 학습이 부족할 수 있습니다.")
                            else:
                                st.info("ℹ️ **일반적 패턴**: 상위권의 정답률이 높습니다.")
                        else:
                            st.warning("해당 문항의 데이터를 찾을 수 없습니다.")
                
                st.divider()
                
                # ============================================================
                # 🔍 5가지 추가 분석 기능
                # ============================================================
                
                # [1️⃣] 예상 vs 실제 난이도 비교
                st.subheader("🎯 예상 vs 실제 난이도 비교")
                st.caption("출제 시 예상한 난이도와 실제 정답률을 비교하여 문항 타당성을 검증합니다.")
                
                # 난이도 매핑: 예상난이도 -> 예상 정답률 범위
                difficulty_expectation = {
                    '하': (0.7, 1.0, "매우 쉬운 문항 - 정답률 70% 이상 예상"),
                    '중': (0.4, 0.7, "중간 난이도 - 정답률 40~70% 예상"),
                    '상': (0.0, 0.4, "어려운 문항 - 정답률 40% 미만 예상")
                }
                
                # 일치도 판정
                def check_difficulty_match(actual_p, exp_diff):
                    if pd.isna(actual_p) or exp_diff == '-':
                        return '미정', 'gray'
                    
                    p_val = float(actual_p) / 100 if actual_p > 1 else float(actual_p)
                    exp_low, exp_high, _ = difficulty_expectation.get(exp_diff, (0, 1, '미정'))
                    
                    # 일치 기준: 정답률이 예상 범위 ±10% 이내
                    if exp_low - 0.1 <= p_val <= exp_high + 0.1:
                        return '✅ 일치', '#90EE90'  # 라이트 그린
                    elif p_val < exp_low - 0.1:
                        return '↑ 낮음', '#FFB6C6'  # 라이트 레드
                    else:
                        return '↓ 높음', '#FFE4B5'  # 몬테카를로
                
                # 예상 vs 실제 데이터 생성
                difficulty_match_data = []
                for idx, row in res_df.iterrows():
                    item_no = int(row['No'])
                    exp_diff = row['Exp_Diff']
                    actual_p = row['정답률(P)']
                    
                    match_status, color = check_difficulty_match(actual_p, exp_diff)
                    exp_low, exp_high, description = difficulty_expectation.get(exp_diff, (None, None, '-'))
                    
                    difficulty_match_data.append({
                        '문항': item_no,
                        '예상난이도': exp_diff,
                        '예상정답률(%)': f"{exp_low*100:.0f}~{exp_high*100:.0f}%" if exp_low is not None else "-",
                        '실제정답률(%)': round(actual_p * 100, 1),
                        '일치도': match_status,
                        '_color': color
                    })
                
                difficulty_match_df = pd.DataFrame(difficulty_match_data)
                
                # 스타일링
                def style_match_status(val):
                    try:
                        color_map = {
                            '✅ 일치': '#90EE90',
                            '↑ 낮음': '#FFB6C6',
                            '↓ 높음': '#FFE4B5',
                            '미정': '#FFFFFF'
                        }
                        color = color_map.get(val, '#FFFFFF')
                        return f'background-color: {color}; text-align: center; font-weight: bold;'
                    except:
                        return ''
                
                styled_match = difficulty_match_df.drop(columns=['_color']).style \
                    .map(style_match_status, subset=['일치도']) \
                    .format(precision=1) \
                    .set_table_styles([
                        {'selector': 'th', 'props': [('text-align', 'center'), ('font-weight', 'bold'), ('background-color', '#f8f9fa')]},
                        {'selector': 'td', 'props': [('text-align', 'center'), ('vertical-align', 'middle')]}
                    ], overwrite=False) \
                    .hide(axis='index')
                
                st.dataframe(styled_match, use_container_width=True)
                
                # 통계 요약
                match_count = (difficulty_match_df['일치도'] == '✅ 일치').sum()
                low_count = (difficulty_match_df['일치도'] == '↑ 낮음').sum()
                high_count = (difficulty_match_df['일치도'] == '↓ 높음').sum()
                
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("예상과 일치", match_count, delta=f"{match_count/len(difficulty_match_df)*100:.1f}%")
                with col2:
                    st.metric("예상보다 낮음", low_count, delta=f"정답률 부족")
                with col3:
                    st.metric("예상보다 높음", high_count, delta=f"정답률 초과")
                
                st.divider()
                
                # [2️⃣] 오답 매력도 분석
                st.subheader("🎯 오답 매력도 분석")
                st.caption("각 오답지가 선택된 비율을 분석합니다. 정답보다 많이 선택된 오답이 있으면 문항 내용 확인을 권장합니다.")
                
                # 선택형 문항별 오답 매력도
                attractiveness_data = []
                for i in range(1, 17):
                    col = f'Item_{i}'
                    
                    try:
                        ans_val = info_df[info_df['No'] == i]['Correct_Ans'].values[0]
                        correct_ans = str(int(ans_val))
                    except:
                        correct_ans = ''
                    
                    # 각 선택지 응답 비율
                    item_responses = main_df[col].astype(str).value_counts(normalize=True) * 100
                    correct_mark_count = item_responses.get('.', 0)
                    
                    choice_rates = {}
                    for j in range(1, 6):
                        count = item_responses.get(str(j), 0)
                        choice_rates[j] = count
                    
                    # 정답 비율 계산
                    if correct_ans in choice_rates:
                        correct_rate = choice_rates[int(correct_ans)] + correct_mark_count
                    else:
                        correct_rate = correct_mark_count
                    
                    # 오답 중 가장 높은 비율
                    max_incorrect_rate = 0
                    max_incorrect_no = 0
                    for j in [1, 2, 3, 4, 5]:
                        if str(j) != correct_ans and choice_rates[j] > max_incorrect_rate:
                            max_incorrect_rate = choice_rates[j]
                            max_incorrect_no = j
                    
                    # 매력도 판정
                    if max_incorrect_rate >= 5:  # 5% 이상 선택된 오답
                        if max_incorrect_rate > correct_rate:
                            attractiveness = '🚨 문제'
                            color = '#FFB6C6'
                        else:
                            attractiveness = '⚠️ 검토'
                            color = '#FFE4B5'
                    else:
                        attractiveness = '✅ 양호'
                        color = '#90EE90'
                    
                    attractiveness_data.append({
                        '문항': i,
                        '정답지(%)': round(correct_rate, 1),
                        '최고오답지(%)': round(max_incorrect_rate, 1),
                        '최고오답번호': max_incorrect_no if max_incorrect_rate > 0 else '-',
                        '평가': attractiveness,
                        '_color': color
                    })
                
                attr_df = pd.DataFrame(attractiveness_data)
                
                def style_attractiveness(val):
                    status_colors = {
                        '🚨 문제': '#FFB6C6',
                        '⚠️ 검토': '#FFE4B5',
                        '✅ 양호': '#90EE90'
                    }
                    color = status_colors.get(val, '#FFFFFF')
                    return f'background-color: {color}; text-align: center; font-weight: bold;'
                
                styled_attr = attr_df.drop(columns=['_color']).style \
                    .map(style_attractiveness, subset=['평가']) \
                    .format(precision=1) \
                    .set_table_styles([
                        {'selector': 'th', 'props': [('text-align', 'center'), ('font-weight', 'bold'), ('background-color', '#f8f9fa')]},
                        {'selector': 'td', 'props': [('text-align', 'center'), ('vertical-align', 'middle')]}
                    ], overwrite=False) \
                    .hide(axis='index')
                
                st.dataframe(styled_attr, use_container_width=True)
                
                # 문제 문항 리스트
                problem_items = attr_df[attr_df['평가'] == '🚨 문제']
                review_items = attr_df[attr_df['평가'] == '⚠️ 검토']
                
                if len(problem_items) > 0:
                    st.error(f"📋 **긴급**: {', '.join(map(str, problem_items['문항'].tolist()))} - 정답 선택률보다 오답 선택률이 높습니다. 문항 내용 확인 권장!")
                
                if len(review_items) > 0:
                    st.warning(f"🔍 **검토**: {', '.join(map(str, review_items['문항'].tolist()))} - 높은 오답 매력도가 있습니다. 문항 검토 권장.")
                
                st.divider()
                
                # [3️⃣] 성취수준별 정답률 패턴 분석
                st.subheader("🔄 성취수준별 정답률 패턴 분석")
                st.caption("성취수준이 높을수록 정답률이 높아야 합니다. 정답률 패턴이 관찰되면 문항의 특성을 분석해 보세요.")
                
                # 정답률 패턴 검사
                reversal_data = []
                for i in range(1, 17):
                    col = f'Item_{i}'
                    item_no = i
                    
                    # 성취수준별 정답률
                    level_rates = {}
                    for level in available_levels:
                        level_data = main_df[main_df['Achievement'] == level]
                        if len(level_data) > 0:
                            lv_rate = (level_data[col].astype(str) == '.').mean() * 100
                            level_rates[level] = lv_rate
                        else:
                            level_rates[level] = np.nan
                    
                    # 패턴 관찰 검사 (단조성 검사)
                    reversals = []
                    level_order = ['A', 'B', 'C', 'D', 'E', '미도달'] if '미도달' in available_levels else ['A', 'B', 'C', 'D', 'E']
                    valid_levels = [l for l in level_order if l in available_levels]
                    
                    for j in range(len(valid_levels) - 1):
                        curr_level = valid_levels[j]
                        next_level = valid_levels[j + 1]
                        
                        curr_rate = level_rates.get(curr_level, np.nan)
                        next_rate = level_rates.get(next_level, np.nan)
                        
                        if not pd.isna(curr_rate) and not pd.isna(next_rate):
                            if curr_rate < next_rate:  # 패턴 관찰
                                reversals.append(f"{curr_level}({curr_rate:.1f}%) > {next_level}({next_rate:.1f}%)")
                    
                    # 평가
                    if len(reversals) == 0:
                        status = '✅ 정상'
                        color = '#90EE90'
                    elif len(reversals) == 1:
                        status = '⚠️ 경미'
                        color = '#FFE4B5'
                    else:
                        status = '🚨 심각'
                        color = '#FFB6C6'
                    
                    reversal_data.append({
                        '문항': item_no,
                        'A정답률': round(level_rates.get('A', np.nan), 1),
                        'B정답률': round(level_rates.get('B', np.nan), 1),
                        'C정답률': round(level_rates.get('C', np.nan), 1),
                        'D정답률': round(level_rates.get('D', np.nan), 1),
                        'E정답률': round(level_rates.get('E', np.nan), 1),
                        '미도달정답률': round(level_rates.get('미도달', np.nan), 1) if '미도달' in available_levels else np.nan,
                        '패턴관찰': ' / '.join(reversals) if reversals else '-',
                        '평가': status,
                        '_color': color
                    })
                
                reversal_df = pd.DataFrame(reversal_data)
                
                # 불필요한 열 제거
                if reversal_df['미도달정답률'].isna().all():
                    reversal_df = reversal_df.drop(columns=['미도달정답률'])
                
                def style_reversal_status(val):
                    status_colors = {
                        '✅ 정상': '#90EE90',
                        '⚠️ 경미': '#FFE4B5',
                        '🚨 심각': '#FFB6C6'
                    }
                    color = status_colors.get(val, '#FFFFFF')
                    return f'background-color: {color}; text-align: center; font-weight: bold;'
                
                styled_reversal = reversal_df.drop(columns=['_color']).style \
                    .map(style_reversal_status, subset=['평가']) \
                    .format(precision=1) \
                    .set_table_styles([
                        {'selector': 'th', 'props': [('text-align', 'center'), ('font-weight', 'bold'), ('background-color', '#f8f9fa'), ('font-size', '12px')]},
                        {'selector': 'td', 'props': [('text-align', 'center'), ('vertical-align', 'middle'), ('font-size', '11px')]}
                    ], overwrite=False) \
                    .hide(axis='index')
                
                st.dataframe(styled_reversal, use_container_width=True)
                
                # 패턴 통계
                normal = (reversal_df['평가'] == '✅ 정상').sum()
                minor = (reversal_df['평가'] == '⚠️ 경미').sum()
                severe = (reversal_df['평가'] == '🚨 심각').sum()
                
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("정상", normal)
                with col2:
                    st.metric("5%p 이내 차이", minor)
                with col3:
                    st.metric("10%p 이상 차이", severe)
                
                st.divider()
                
                # [4️⃣] 문항 특성 종합 분석
                st.subheader("📊 문항 특성 종합 분석")
                st.caption("5가지 지표를 종합하여 각 문항의 특성을 분류합니다. 이는 참고용이며, 최종 판단은 문항 내용 검토와 함께 이루어져야 합니다.")
                
                def determine_grade(item_no, quality_df, difficulty_match_df, attr_df, reversal_df):
                    """5개 지표를 종합한 등급 판정"""
                    
                    # 1. 양호도 (0~50점)
                    quality_status = quality_df[quality_df['문항'] == item_no]['양호도'].values
                    if len(quality_status) > 0:
                        q_score = {
                            # 선발형
                            '✅ 높은 변별도 (0.4+)': 50,
                            '📝 매우 쉬움': 25,
                            '📚 초고난도': 20,
                            '📚 어려움': 25,
                            '📊 낮은 변별도 (0.2 미만)': 10,
                            '📋 초고난도/역변별': 5,
                            '📋 보통': 30,
                            # 성취평가제 (패턴 기반)
                            '✅ A수준+변별': 50,
                            '✅ B수준+변별': 50,
                            '✅ C수준+변별': 50,
                            '✅ D수준+변별': 50,
                            '✅ E수준+변별': 50,
                            '✅ A수준': 48,
                            '✅ B수준': 48,
                            '✅ C수준': 48,
                            '✅ D수준': 48,
                            '✅ E수준': 48,
                            '📋 A수준(패턴)': 35,
                            '📋 B수준(패턴)': 35,
                            '📋 C수준(패턴)': 35,
                            '📋 D수준(패턴)': 35,
                            '📋 E수준(패턴)': 35,
                            '📋 판정불가(패턴)': 30,
                            '📊 판정불가': 15,
                            '✅ 기초달성': 40,
                            '📋 학습미달': 10,
                        }.get(quality_status[0], 25)
                    else:
                        q_score = 25
                    
                    # 2. 난이도 일치도 (0~30점)
                    match_status = difficulty_match_df[difficulty_match_df['문항'] == item_no]['일치도'].values
                    if len(match_status) > 0:
                        d_score = {'✅ 일치': 30, '↑ 낮음': 15, '↓ 높음': 15, '미정': 0}.get(match_status[0], 15)
                    else:
                        d_score = 15
                    
                    # 3. 오답 매력도 (0~10점)
                    attr_status = attr_df[attr_df['문항'] == item_no]['평가'].values
                    if len(attr_status) > 0:
                        a_score = {'✅ 양호': 10, '⚠️ 검토': 5, '🚨 문제': 0}.get(attr_status[0], 5)
                    else:
                        a_score = 5
                    
                    # 4. 패턴 관찰 (0~10점)
                    reversal_status = reversal_df[reversal_df['문항'] == item_no]['평가'].values
                    if len(reversal_status) > 0:
                        r_score = {'✅ 정상': 10, '⚠️ 경미': 5, '🚨 심각': 0}.get(reversal_status[0], 5)
                    else:
                        r_score = 5
                    
                    total_score = q_score + d_score + a_score + r_score  # 총 100점
                    
                    # 등급 판정 (점수 범위별)
                    if total_score >= 85:
                        grade = 'A'
                        description = '5개 지표 모두 높은 범위 (85점 이상)'
                        color = '#90EE90'
                    elif total_score >= 70:
                        grade = 'B'
                        description = '대부분 지표 양호 (70-84점)'
                        color = '#ADD8E6'
                    elif total_score >= 50:
                        grade = 'C'
                        description = '중간 범위 (50-69점)'
                        color = '#FFE4B5'
                    elif total_score >= 30:
                        grade = 'D'
                        description = '여러 지표 낮음 (30-49점)'
                        color = '#FFCCCC'
                    else:
                        grade = 'F'
                        description = '대부뵘8 지표 낮음 (29점 이하)'
                        color = '#FFB6C6'
                    
                    return grade, total_score, description, color, q_score, d_score, a_score, r_score
                
                # 모든 문항의 등급 계산
                final_grades = []
                for item_no in range(1, 17):
                    grade, score, desc, color, q, d, a, r = determine_grade(
                        item_no, quality_df, difficulty_match_df, attr_df, reversal_df
                    )
                    final_grades.append({
                        '문항': item_no,
                        '등급': grade,
                        '총점': score,
                        '양호도': q,
                        '난이도': d,
                        '오답': a,
                        '패턴': r,
                        '평가': desc,
                        '_color': color
                    })
                
                grade_df = pd.DataFrame(final_grades)
                
                def style_final_grade(val):
                    grade_colors = {
                        'A': '#90EE90',
                        'B': '#ADD8E6',
                        'C': '#FFE4B5',
                        'D': '#FFCCCC',
                        'F': '#FFB6C6'
                    }
                    color = grade_colors.get(val, '#FFFFFF')
                    return f'background-color: {color}; text-align: center; font-weight: bold; font-size: 16px;'
                
                styled_grade = grade_df.drop(columns=['_color']).style \
                    .map(style_final_grade, subset=['등급']) \
                    .format(precision=0, subset=['등급', '총점', '양호도', '난이도', '오답', '패턴']) \
                    .set_table_styles([
                        {'selector': 'th', 'props': [('text-align', 'center'), ('font-weight', 'bold'), ('background-color', '#f8f9fa')]},
                        {'selector': 'td', 'props': [('text-align', 'center'), ('vertical-align', 'middle')]}
                    ], overwrite=False) \
                    .hide(axis='index')
                
                st.dataframe(styled_grade, use_container_width=True)
                
                # 등급별 통계
                grade_counts = grade_df['등급'].value_counts().sort_index()
                
                col1, col2, col3, col4, col5 = st.columns(5)
                gradeColors = {'A': '#90EE90', 'B': '#ADD8E6', 'C': '#FFE4B5', 'D': '#FFCCCC', 'F': '#FFB6C6'}
                
                for idx, (grade, count) in enumerate(grade_counts.items()):
                    with [col1, col2, col3, col4, col5][idx]:
                        color = gradeColors.get(grade, '#FFF')
                        st.markdown(
                            f'<div style="background-color:{color}; padding: 15px; border-radius: 8px; text-align: center;">'
                            f'<div style="font-size:24px; font-weight:bold; margin-bottom:5px;">{grade}</div>'
                            f'<div style="font-size:18px; font-weight:bold;">{count}개</div>'
                            f'</div>',
                            unsafe_allow_html=True
                        )
                
                st.divider()
                
                # 등급별 상세 분석
                st.subheader("📊 등급별 상세 분석")
                
                tabs_grade = st.tabs(['A등급 (85+)', 'B등급 (70-84)', 'C등급 (50-69)', 'D등급 (30-49)', 'F등급 (~29)'])
                
                grade_info = {
                    'A': ('5개 지표 모두 높은 범위 (85점 이상)', grade_df[grade_df['등급'] == 'A']),
                    'B': ('대부분 지표 양호 (70-84점)', grade_df[grade_df['등급'] == 'B']),
                    'C': ('중간 범위 (50-69점)', grade_df[grade_df['등급'] == 'C']),
                    'D': ('여러 지표 낮음 (30-49점)', grade_df[grade_df['등급'] == 'D']),
                    'F': ('대부분 지표 낮음 (29점 이하)', grade_df[grade_df['등급'] == 'F'])
                }
                
                for tab, (grade_letter, (desc, df_grade)) in zip(tabs_grade, grade_info.items()):
                    with tab:
                        if len(df_grade) > 0:
                            st.write(f"**{desc}**")
                            st.write(f"해당 문항: {', '.join(map(str, df_grade['문항'].tolist()))}")
                            
                            # 상세 정보 표시
                            display_df = df_grade.drop(columns=['_color'] if '_color' in df_grade.columns else [])
                            st.dataframe(display_df, use_container_width=True, hide_index=True)
                        else:
                            st.info(f"해당 등급의 문항이 없습니다.")
                
                st.divider()
                
                # 서답형 문항 분석 (테스트 모드에서 생성)
                # 정답률이 없거나 모두 0이면 테스트 모드
                select_score_exists = (main_df['Select_Score'] > 0).any() if 'Select_Score' in main_df.columns else False
                
                if not select_score_exists:
                    st.subheader("📋 서답형 문항 분석 (테스트 데이터)")
                    st.info("🎨 **서답형 문항의 색상 범례**: 1/2 이상은 흰색, 1/2 미만은 회색입니다.")
                    
                    # 테스트 모드: 서답형 문항 임의 생성 (서답형 1-3)
                    np.random.seed(43)
                    essay_items_count = 3
                    
                    # 각 학생별 서답형 문항 응답 및 점수 생성
                    essay_item_data = []
                    for item_num in range(1, essay_items_count + 1):
                        # 정답률 임의 생성 (30-90%)
                        correct_rate = np.random.uniform(0.3, 0.9)
                        num_correct = int(len(main_df) * correct_rate)
                        
                        # 학생 응답 생성 (정답: 'O', 오답: 'X')
                        response = ['O'] * num_correct + ['X'] * (len(main_df) - num_correct)
                        np.random.shuffle(response)
                        main_df[f'Essay_{item_num}'] = response
                        
                        # 각 학생의 문항별 점수 계산 (정답 10점, 오답 0-7점 랜덤)
                        item_scores = []
                        for idx, row_data in main_df.iterrows():
                            if row_data[f'Essay_{item_num}'] == 'O':
                                item_scores.append(10)
                            else:
                                item_scores.append(np.random.randint(0, 8))
                        
                        # 통계 계산
                        row = {
                            ('', '문항'): f'서답형{item_num}',
                            ('', '예상난이도'): np.random.choice(['상', '중', '하']),
                            ('통계', '최소값'): float(min(item_scores)),
                            ('통계', '최댓값'): float(max(item_scores)),
                            ('통계', '평균'): float(np.mean(item_scores)),
                            ('통계', '표준편차'): float(np.std(item_scores, ddof=1)),
                            ('', '정답률(%)'): correct_rate * 100,
                            ('', '변별도'): np.random.uniform(0.1, 0.8)
                        }
                        
                        # 성취수준별 정답률
                        for level in available_levels:
                            level_df = main_df[main_df['Achievement'] == level]
                            if len(level_df) > 0:
                                level_correct_rate = (level_df[f'Essay_{item_num}'] == 'O').mean() * 100
                            else:
                                level_correct_rate = np.nan
                            row[('성취수준별 정답률(%)', level)] = level_correct_rate
                            
                        essay_item_data.append(row)
                    
                    essay_analysis = pd.DataFrame(essay_item_data)
                    
                    # 컬럼 순서
                    # [수정] 1행처럼 보이게 하기 위해 첫번째 레벨 빈 문자열 처리
                    ecols_basic = [('', '문항'), ('', '예상난이도')]
                    ecols_stat = [('통계', '최소값'), ('통계', '최댓값'), ('통계', '평균'), ('통계', '표준편차')]
                    ecols_res = [('', '정답률(%)'), ('', '변별도')]
                    ecols_lv = [('성취수준별 정답률(%)', lv) for lv in available_levels]
                    
                    efinal_cols = ecols_basic + ecols_stat + ecols_res + ecols_lv
                    # 존재하는 컬럼만 필터링 (만에하나)
                    efinal_cols = [c for c in efinal_cols if c in essay_analysis.columns]
                    essay_analysis = essay_analysis[efinal_cols]
                    
                    # 서답형 데이터프레임 처리
                    # 기존 DataFrame의 컬럼이 `('', '문항')` 형태이므로 `('문항', '문항')`으로 변경
                    rename_map = {
                        ('', '문항'): ('문항', '문항'),
                        ('문항', '문항'): ('문항', '문항'),
                        ('', '예상난이도'): ('예상난이도', '예상난이도'),
                        ('예상난이도', '예상난이도'): ('예상난이도', '예상난이도'),
                        ('', '정답률(%)'): ('정답률', '정답률'), # [변경]
                        ('정답률', '정답률(%)'): ('정답률', '정답률'),
                        ('', '변별도'): ('변별도', '변별도'),
                        ('변별도', '변별도'): ('변별도', '변별도')
                    }
                    
                    new_cols = []
                    for c in essay_analysis.columns:
                        if c in rename_map:
                            new_cols.append(rename_map[c])
                        else:
                            new_cols.append(c)
                    essay_analysis.columns = new_cols

                    # [중요] 컬럼을 MultiIndex로 명시적 변환
                    essay_analysis.columns = pd.MultiIndex.from_tuples(
                        essay_analysis.columns, 
                        names=[None, None]
                    )

                    # [디버깅] 스타일 적용을 위해 숫자형으로 변환 보장 및 소수점 1자리 반올림
                    essay_analysis[('정답률', '정답률')] = pd.to_numeric(essay_analysis[('정답률', '정답률')], errors='coerce').round(1)
                    essay_analysis[('변별도', '변별도')] = pd.to_numeric(essay_analysis[('변별도', '변별도')], errors='coerce').round(1)
                    
                    # 스타일 함수 (재사용)
                    e_level_cols = [c for c in essay_analysis.columns if c[0] == '성취수준별 정답률(%)']
                    
                    e_styler = essay_analysis.style.format(precision=1, na_rep='') \
                        .hide(axis='index') \
                        .format(precision=2, subset=[c for c in ecols_stat if c in essay_analysis.columns]) \
                        .format(precision=1, subset=[('정답률', '정답률')]) \
                        .format(precision=1, subset=[('변별도', '변별도')]) \
                        .map(lambda x: style_background_level_v2(x, 50.0), subset=e_level_cols) \
                        .map(lambda x: custom_bar_style(x, 50.0), subset=[('정답률', '정답률')]) \
                        .bar(subset=[('변별도', '변별도')], color='#a5d6a7', vmin=-0.2, vmax=1.0) \
                        .set_table_styles([
                            {'selector': 'th', 'props': [
                                ('text-align', 'center'), 
                                ('font-weight', 'bold'), 
                                ('color', 'black'), 
                                ('vertical-align', 'middle'), 
                                ('border', '1px solid #e0e0e0'),
                                ('background-color', '#f8f9fa')
                            ]},
                            {'selector': 'td', 'props': [
                                ('text-align', 'center'), 
                                ('vertical-align', 'middle'), 
                                ('border', '1px solid #e0e0e0')
                            ]}
                        ], overwrite=False)
                        
                    html_e = e_styler.to_html(escape=False)
                    
                    # [HTML 후처리] 헤더 병합 (서답형용)
                    # 전역 함수 merge_headers 사용
                    html_e = merge_headers(html_e, ['문항', '예상난이도', '정답률', '변별도'])

                    # [DataTables 렌더링 호출]
                    render_datatables(html_e, unique_id='essay_analysis')
                    
                    st.caption("📌 테스트 데이터로 생성된 서답형 문항 분석입니다.")
                
                st.divider()
                
                # 서답형 전체 성적 분석
                st.subheader("📋 서답형 전체 성적 분석")
                st.caption("서답형은 전체 점수로만 수집되어 개별 문항 분석이 불가합니다.")
                
                # 서답형 데이터 (Essay_Score가 있으면 표시)
                essay_df = main_df.copy()
                
                # 테스트 모드: Essay_Score가 없거나 모두 0이면 임의 데이터 생성
                test_mode = False
                if 'Essay_Score' not in essay_df.columns or (essay_df['Essay_Score'] == 0).all():
                    np.random.seed(42)
                    # 서답형 점수 임의 생성 (0-20점 범위)
                    essay_df['Essay_Score'] = np.random.randint(0, 21, size=len(essay_df))
                    
                    # Semester_Score도 없으면 생성
                    if 'Semester_Score' not in essay_df.columns:
                        essay_df['Semester_Score'] = (essay_df['Essay_Score'] * 1.5).round(1)
                    
                    st.info("📌 서답형 데이터 없음 - **테스트 데이터를 임의로 생성하여 표시**합니다.")
                    test_mode = True
                
                if 'Essay_Score' in essay_df.columns and (essay_df['Essay_Score'] > 0).any():
                    # 필요한 컬럼만 선택 (없는 컬럼은 제외)
                    cols_to_select = ['강의실', 'ID', 'Name', 'Essay_Score']
                    if 'Semester_Score' in essay_df.columns:
                        cols_to_select.append('Semester_Score')
                    if 'Achievement' in essay_df.columns:
                        cols_to_select.append('Achievement')
                    
                    essay_display = essay_df[cols_to_select]
                    
                    # 컬럼명 변환
                    rename_map = {
                        '강의실': '수강반',
                        'ID': '학번',
                        'Name': '이름',
                        'Essay_Score': '서답형점수',
                        'Semester_Score': '학기말점수',
                        'Achievement': '성취수준'
                    }
                    # 실제 존재하는 컬럼만 rename
                    rename_map = {k: v for k, v in rename_map.items() if k in essay_display.columns}
                    essay_display = essay_display.rename(columns=rename_map)
                    
                    essay_display = essay_display[essay_display['서답형점수'] > 0]
                    
                    # 수치 포맷팅
                    if '서답형점수' in essay_display.columns:
                        essay_display['서답형점수'] = essay_display['서답형점수'].round(1)
                    if '학기말점수' in essay_display.columns:
                        essay_display['학기말점수'] = essay_display['학기말점수'].round(1)
                    
                    # 동적 column_config 생성
                    col_config = {
                        "수강반": st.column_config.TextColumn("수강반", width="small"),
                        "학번": st.column_config.TextColumn("학번", width="small"),
                        "이름": st.column_config.TextColumn("이름", width="small"),
                        "서답형점수": st.column_config.NumberColumn("서답형점수", format="%.1f", width="small"),
                        "학기말점수": st.column_config.NumberColumn("학기말점수", format="%.1f", width="small"),
                        "성취수준": st.column_config.TextColumn("성취수준", width="small")
                    }
                    # 존재하는 컬럼에 대해서만 config 생성
                    col_config = {k: v for k, v in col_config.items() if k in essay_display.columns}
                    
                    st.dataframe(
                        essay_display,
                        use_container_width=True,
                        height=400,
                        hide_index=True,
                        column_config=col_config
                    )
                else:
                    st.info("서답형 데이터가 없습니다.")

                # --- 데이터 검증 섹션 (tab_item 내) ---
                st.divider()
                with st.expander("🔍 성취수준별 정답률 검증", expanded=False):
                    st.subheader("성취수준별 정답률 계산 검증")
                    
                    # 검증할 문항 선택
                    verify_item = st.selectbox("검증할 문항 선택", options=range(1, 17), key="verify_item_select")
                    verify_column = f'Item_{verify_item}'
                    
                    if verify_column in main_df.columns:
                        st.write(f"**문항 {verify_item} 정답률 검증**")
                        
                        # 각 성취수준별로 계산
                        verify_data = []
                        for level in available_levels:
                            level_students = main_df[main_df['Achievement'] == level]
                            total_count = len(level_students)
                            
                            if total_count > 0:
                                correct_count = (level_students[verify_column].astype(str) == '.').sum()
                                correct_rate = (correct_count / total_count) * 100
                                
                                verify_data.append({
                                    '성취수준': level,
                                    '총 학생 수': total_count,
                                    '정답 학생 수': correct_count,
                                    '정답 학생 비율(%)': round(correct_rate, 1),
                                    '오답 학생 수': total_count - correct_count
                                })
                        
                        if verify_data:
                            verify_df = pd.DataFrame(verify_data)
                            st.dataframe(verify_df, use_container_width=True, hide_index=True)
                            
                            # 계산 과정 표시
                            st.write("**계산 과정 예시:**")
                            first_level = verify_data[0]
                            st.code(
                                f"""성취수준 '{first_level['성취수준']}' 정답률:
= 정답 학생 수 / 총 학생 수 × 100
= {int(first_level['정답 학생 수'])} / {first_level['총 학생 수']} × 100
= {first_level['정답 학생 비율(%)']}%""",
                                language="text"
                            )
                        else:
                            st.warning(f"성취수준 '{level}'에 해당하는 학생이 없습니다.")
                    else:
                        st.warning(f"문항 {verify_item} 데이터를 찾을 수 없습니다.")
                
                # 개별 문항 상세 분석 섹션 (2컬럼 레이아웃)
                st.divider()
                with st.expander("🔍 개별 문항 상세 분석 (명확한 판단 vs 참고 정보)", expanded=False):
                    st.markdown("**선택한 문항의 상세 분석 결과를 2가지 관점으로 제시합니다.**")
                    
                    # 분석할 문항 선택
                    selected_item_detail = st.selectbox("분석할 문항 번호", options=range(1, 17), key="detail_item_select")
                    
                    # 문항 정보 수집
                    item_col = f'Item_{selected_item_detail}'
                    
                    # 문항 정보 가져오기
                    item_info = info_df[info_df['No'] == selected_item_detail]
                    exp_diff_val = item_info['Exp_Diff'].values[0] if len(item_info) > 0 and 'Exp_Diff' in item_info.columns else '-'
                    
                    try:
                        correct_ans = str(int(item_info['Correct_Ans'].values[0]))
                    except:
                        correct_ans = '-'
                    
                    # 정답률 계산
                    correct_rate = (main_df[item_col].astype(str) == '.').mean() * 100
                    
                    # 변별도 (상위 27% vs 하위 27%)
                    sorted_df = main_df.sort_values('Total_Score' if 'Total_Score' in main_df.columns else main_df.columns[-1], ascending=False)
                    cutoff = max(1, len(sorted_df) // 4)
                    top_group = sorted_df.head(cutoff)
                    bottom_group = sorted_df.tail(cutoff)
                    
                    top_correct = (top_group[item_col].astype(str) == '.').mean() * 100
                    bottom_correct = (bottom_group[item_col].astype(str) == '.').mean() * 100
                    discrimination = (top_correct - bottom_correct) / 100
                    
                    # 성취수준별 정답률
                    achievement_rates = {}
                    student_counts = {}
                    for level in available_levels:
                        level_data = main_df[main_df['Achievement'] == level]
                        student_counts[level] = len(level_data)
                        if len(level_data) > 0:
                            achievement_rates[level] = (level_data[item_col].astype(str) == '.').mean() * 100
                        else:
                            achievement_rates[level] = 0
                    
                    # 선택지별 응답분포
                    item_responses = main_df[item_col].astype(str).value_counts()
                    response_dist = {}
                    for j in range(1, 6):
                        response_dist[j] = item_responses.get(str(j), 0)
                    response_dist['noResponse'] = item_responses.get('nan', 0) + item_responses.get('', 0)
                    
                    # 응답분포를 비율(%)로 변환
                    total_students = len(main_df)
                    response_dist_pct = {k: (v / total_students * 100) if total_students > 0 else 0 for k, v in response_dist.items()}
                    
                    # =========== 2컬럼 레이아웃 시작 ===========
                    col_definite, col_reference = st.columns(2)
                    
                    # ─── 왼쪽 컬럼: ✅ 명확한 판단 ───
                    with col_definite:
                        st.markdown("### ✅ 명확한 판단")
                        st.markdown("*객관적 기준에 따른 판정 결과*")
                        st.markdown("---")
                        
                        # 성취평가제에서 문항 수준 판정
                        if st.session_state.get('eval_type') == 'achievement':
                            criterion = st.session_state.get('criterion_rate', 66.7)
                            item_level_result = determine_item_level(achievement_rates, criterion)
                            
                            st.markdown(f"**📊 문항 수준: {item_level_result['level']}수준 문항**")
                            
                            meets_str = ", ".join([f"{l}({achievement_rates.get(l, 0):.1f}%)✓" for l in item_level_result['meets']]) 
                            below_str = ", ".join([f"{l}({achievement_rates.get(l, 0):.1f}%)" for l in item_level_result['below']])
                            
                            st.markdown(f"- 기준 충족: {meets_str}")
                            if below_str:
                                st.markdown(f"- 기준 미달: {below_str}")
                            st.markdown("")
                        
                        # 난이도 판정
                        difficulty_result = evaluate_difficulty(correct_rate)
                        st.markdown(f"**📈 난이도: {difficulty_result['level']}** ({correct_rate:.1f}%)")
                        st.caption(f"{difficulty_result['description']}")
                        st.markdown("")
                        
                        # 변별도 판정
                        discrimination_result = evaluate_discrimination(discrimination)
                        st.markdown(f"**📉 변별도: {discrimination_result['level']}** ({discrimination:.2f})")
                        st.caption(f"{discrimination_result['description']}")
                        st.markdown("")
                        
                        # 오답 매력도 분석
                        st.markdown("**🔢 오답 매력도**")
                        distractor_results = analyze_distractor(response_dist_pct, int(correct_ans) if correct_ans != '-' else 0)
                        
                        for result in distractor_results:
                            if result['type'] == 'correct':
                                st.markdown(f"- {result['option']}번: {result['rate']:.1f}% (정답) ✓")
                            elif result['type'] == 'no_response':
                                st.markdown(f"- 무응답: {result['rate']:.1f}%")
                            elif result['type'] == 'functional':
                                st.markdown(f"- {result['option']}번: {result['rate']:.1f}% - 적절 ○")
                            else:
                                st.markdown(f"- {result['option']}번: {result['rate']:.1f}% - **낮음** △")
                    
                    # ─── 오른쪽 컬럼: 📋 참고 정보 ───
                    with col_reference:
                        st.markdown("### 📋 참고 정보")
                        st.markdown("*해석이 필요한 관찰 사항*")
                        st.markdown("---")
                        
                        if st.session_state.get('eval_type') == 'achievement':
                            pattern_result = analyze_achievement_pattern(achievement_rates, student_counts)
                            
                            if pattern_result['overall_status'] == 'good':
                                st.success(pattern_result['overall'])
                            else:
                                st.markdown(f"**💡 관찰된 패턴**")
                                st.markdown(pattern_result['overall'])
                                
                                for obs in pattern_result['observations']:
                                    with st.container():
                                        st.markdown(f"""
                                        **{obs['icon']} {obs['title']}**: {obs['upper']}-{obs['lower']} 구간
                                        
                                        {obs['upper']}수준({obs['upper_rate']:.1f}%)과 {obs['lower']}수준({obs['lower_rate']:.1f}%) 간 
                                        **{obs['abs_gap']:.1f}%p** 차이가 관찰되었습니다.{obs['count_info']}
                                        
                                        {obs['message']}
                                        """)
                                        
                                        with st.expander("가능한 원인", expanded=False):
                                            st.markdown("""
                                            • 해당 수준의 학생 수가 적어 발생하는 통계적 변동
                                            • 문항이 측정하는 특정 개념에서의 집단별 차이
                                            • 선택형 문항에서의 추측 효과
                                            • 자연스러운 응답 패턴의 일부
                                            """)
                                        
                                        st.caption(obs['disclaimer'])
                                        st.markdown("---")
                        else:  # 선발형
                            st.markdown("**💡 상대평가 분석 시 참고 사항**")
                            
                            target = st.session_state.get('target_rate', 70)
                            diff_from_target = correct_rate - target
                            
                            if abs(diff_from_target) <= 10:
                                st.success(f"정답률({correct_rate:.1f}%)이 목표({target}%)에 근접합니다.")
                            elif diff_from_target > 10:
                                st.info(f"정답률({correct_rate:.1f}%)이 목표({target}%)보다 {diff_from_target:.1f}%p 높습니다.")
                            else:
                                st.info(f"정답률({correct_rate:.1f}%)이 목표({target}%)보다 {abs(diff_from_target):.1f}%p 낮습니다.")

            # --- [Tab 4] 성취수준별 답지반응 ---
            with tab_dist:
                st.subheader("문항별 반응 상세 분석")
                col_sel, col_desc = st.columns([1, 2])
                with col_sel:
                    sel_item = st.selectbox("분석할 문항 번호를 선택하세요", options=range(1, 17))
                with col_desc:
                    # 성취기준 없을 경우 대비
                    std_text = info_df.loc[info_df['No']==sel_item, 'Standard']
                    std_val = std_text.values[0] if not std_text.empty else "(성취기준 정보 없음)"
                    st.info(f"📌 **성취기준**: {std_val}")
                
                # 수준별 정답률 곡선
                # 안전한 Groupby 계산
                if sel_item:
                    level_perf = main_df.groupby('Achievement')[f'Item_{sel_item}'].apply(
                        lambda x: (x.astype(str) == '.').mean() * 100
                    ).reindex(['A','B','C','D','E','미도달']).fillna(0)
                    
                    # 성취평가제일 경우 패턴 분석
                    achievement_rates_chart = dict(level_perf)
                    student_counts_chart = main_df.groupby('Achievement').size().to_dict()
                    
                    fig_curve = go.Figure()
                    
                    # 성취평가제: 패턴 관찰 구간 하이라이트
                    if st.session_state.get('eval_type') == 'achievement':
                        pattern_result = analyze_achievement_pattern(achievement_rates_chart, student_counts_chart)
                        
                        # 패턴이 있는 구간 표시 (연한 파란색)
                        if pattern_result['has_observations']:
                            levels_list = ['A', 'B', 'C', 'D', 'E', '미도달']
                            for obs in pattern_result['observations']:
                                try:
                                    upper_idx = levels_list.index(obs['upper'])
                                    fig_curve.add_vrect(
                                        x0=upper_idx - 0.3,
                                        x1=upper_idx + 1.3,
                                        fillcolor="rgba(52, 152, 219, 0.1)",  # 연한 파란색
                                        layer="below",
                                        line_width=0,
                                        annotation_text="📊 관찰됨",
                                        annotation_position="top left",
                                        annotation_font_size=10,
                                        annotation_font_color="rgba(52, 152, 219, 0.8)"
                                    )
                                except:
                                    pass
                        
                        # 기준선 (66.7%)
                        criterion = st.session_state.get('criterion_rate', 66.7)
                        fig_curve.add_hline(
                            y=criterion,
                            line_dash="dash",
                            line_color="orange",
                            line_width=2,
                            annotation_text=f"기준 {criterion}%",
                            annotation_position="right",
                            annotation_font_color="orange",
                            annotation_font_size=11
                        )
                    
                    # 메인 라인 차트
                    fig_curve.add_trace(go.Scatter(
                        x=level_perf.index, y=level_perf.values, 
                        mode='lines+markers+text', 
                        text=[f"{v:.1f}%" for v in level_perf.values],
                        textposition="top center",
                        name='정답률', 
                        line=dict(color='#3498db', width=3),
                        marker=dict(size=10, color='#3498db')
                    ))
                    
                    fig_curve.update_layout(
                        title=f"<b>{sel_item}번 문항: 성취수준별 정답률 추이</b>", 
                        xaxis_title="성취수준", 
                        yaxis_title="정답률 (%)",
                        yaxis=dict(range=[0, 105]),
                        paper_bgcolor="rgba(0,0,0,0)",
                        plot_bgcolor="rgba(240,242,246,0.3)",
                        font_family="Pretendard",
                        height=400,
                        font=dict(size=12),
                        hovermode='x unified',
                        showlegend=False
                    )
                    st.plotly_chart(fig_curve, use_container_width=True)

            # --- [Tab 5] 성취기준 분석 결과 ---
            with tab_std:
                st.subheader("📜 성취기준별 분석 결과")
                
                # 성취기준별 그룹화
                std_stats = res_df.groupby('Standard').agg({
                    'No': 'count',
                    'Score': 'sum',
                    '정답률(P)': 'mean',
                    '변별도(D)': 'mean'
                }).reset_index()
                
                std_stats.columns = ['성취기준', '문항 수', '배점 합계', '평균 정답률', '평균 변별도']
                
                # 스타일링 (matplotlib 없이)
                st.dataframe(
                    std_stats.style.format({
                        '평균 정답률': '{:.2f}',
                        '평균 변별도': '{:.2f}'
                    }),
                    use_container_width=True,
                    hide_index=True
                )

            # --- [Tab 6] 분석 리포트 ---
            with tab_report:
                st.subheader("📝 AI 자동 분석 리포트 및 출제 제언")
                
                good_items = res_df[res_df['변별도(D)'] >= 0.4]['No'].tolist()
                bad_items = res_df[res_df['변별도(D)'] < 0.2]['No'].tolist()
                
                st.markdown(f"""
                #### 1. 평가 도구 종합 진단
                - **신뢰도:** 본 검사의 신뢰도는 **{alpha:.3f}**로, { '높은 일관성(안정적)' if alpha >= 0.8 else '중간 수준' }을 보입니다.
                - **변별도:** 전체 문항 중 **{len(good_items)}개** 문항이 상위권과 하위권을 명확히 구분합니다 (변별도 0.4 이상).
                
                #### 2. 문항 참고 정보
                - **🌟 높은 변별도 문항:** {', '.join(map(str, good_items[:5]))}번 등 (변별력이 높아 향후 유사 유형 출제 시 참고)
                - **🔧 검토 권장:** {', '.join(map(str, bad_items)) if bad_items else '없음'}번 등 (변별도가 낮음. 발문이나 선택지 검토 시 참고)
                
                #### 3. 차후 출제 시 참고사항
                > **Tip:** 정답률이 지나치게 높거나 낮은 문항은 수업 중 강조점을 확인하거나, 난이도 조정을 고려해 볼 수 있습니다.
                """)
    
    else:
        # 데이터 로드 실패시
        st.error("⚠️ **데이터 로드에 실패했습니다.**")
        st.info("""
        다음을 확인하세요:
        
        **정기고사 (분할점수 기반):**
        - 문항정보표와 학생 정오표 파일이 올바른 NEIS 양식인지 확인
        - 파일명에 특수문자가 없는지 확인
        
        **정기고사 (학기말 성취도 기반):**
        - 위의 확인사항 + 성적일람표가 성취도 정보를 포함하고 있는지 확인
        
        **수행평가:**
        - 평가기준표와 성적일람표가 올바른 양식인지 확인
        """)

else:
    # 데이터 미업로드 시 초기 화면
    st.container()
    st.info("👈 **시작하려면 왼쪽 사이드바에서 필요한 파일을 업로드하세요.**")
    
    if analysis_basis == "분할점수 기반":
        st.success("""
        #### 분석 방식: 입력 분할점수 기반 자동 판정
        
        왼쪽 설정에서 입력한 분할점수를 기준으로 학생 성적을 분석하여 **자동으로 성취도를 판정**합니다.
        """)
        if exam_category == "정기고사":
            st.write(f"### 📝 {exam_name} 준비물")
            st.write("""
            1. **📑 문항정보표** - NEIS에서 다운로드
            2. **✍️ 학생 정오표** - NEIS에서 다운로드 (여러 반 가능)
            """)
        else:
            st.write("### 📝 수행평가 준비물")
            st.write("""
            1. **📑 평가기준표** - 수행평가 항목과 배점 정보
            """)
    else:  # 학기말 성취도 기반
        st.success("""
        #### 분석 방식: 성적일람표 성취도 기준
        
        성적일람표에 이미 판정되어 있는 성취도를 **그대로 사용**합니다.
        """)
        if exam_category == "정기고사":
            st.write(f"### 📝 {exam_name} 준비물")
            st.write("""
            1. **📑 문항정보표** - NEIS에서 다운로드
            2. **✍️ 학생 정오표** - NEIS에서 다운로드 (여러 반 가능)
            3. **📊 성적일람표** - 성취도 정보 포함 (여러 반 가능)
            """)
        else:
            st.write("### 📝 수행평가 준비물")
            st.write("""
            1. **📑 평가기준표** - 수행평가 항목과 배점 정보
            2. **📊 성적일람표** - 수행평가 점수와 성취도 정보
            """)
    
    st.markdown("---")
    st.caption("🔒 데이터가 서버로 전송되지 않고 브라우저에서 안전하게 처리됩니다.")